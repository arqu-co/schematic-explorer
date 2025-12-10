"""Adaptive extraction logic for any tower format.

This extractor analyzes the spreadsheet structure from first principles:
1. Merged cells define visual blocks - these are the primary structure
2. Spatial proximity determines relationships
3. Content patterns infer field types
4. No assumptions about specific labels or column positions
"""

import re
from typing import Any

from openpyxl.utils import get_column_letter, range_boundaries

from .blocks import Block, classify_blocks
from .carriers import _is_non_carrier, _looks_like_policy_number
from .proximity import (
    calculate_block_proximity,
    detect_summary_columns,
    get_column_range,
    is_block_relevant,
    match_currency_block,
    match_participation_block,
)
from .types import BILLION, CarrierEntry, LayerSummary, parse_excess_notation, parse_limit_value
from .utils import build_merged_cell_map, get_cell_color, get_cell_value

# =============================================================================
# Constants
# =============================================================================

# Proximity thresholds for spatial matching
MAX_ROW_SEARCH_DISTANCE = 10  # How many rows below carrier to search for data
LAYER_ROW_PROXIMITY = 2  # Distance threshold for "nearby" layer rows

# Column limits
MAX_LAYER_LIMIT_COLUMN = 2  # Layer limits should be in columns A or B
MAX_HEADER_SCAN_ROW = 10  # How many rows to scan for headers

# Confidence thresholds
MIN_CARRIER_CONFIDENCE = 0.5  # Minimum confidence for carrier block inclusion

# =============================================================================
# Main Entry Point
# =============================================================================


def extract_adaptive(ws) -> tuple[list[CarrierEntry], list[LayerSummary]]:
    """Extract tower data by analyzing merged cell structure and spatial patterns.

    Returns:
        tuple: (carrier_entries, layer_summaries)
            - carrier_entries: List of per-carrier participation data
            - layer_summaries: List of layer-level aggregate data (for cross-checking)
    """

    # Step 1: Find all blocks (merged regions and significant single cells)
    blocks = _find_all_blocks(ws)

    # Step 2: Classify blocks by content
    classify_blocks(blocks)

    # Step 3: Detect summary/aggregate columns to exclude from carrier extraction
    # These columns contain layer-level data like "Annualized Layer Rate" not per-carrier data
    summary_info = detect_summary_columns(ws)
    summary_cols = summary_info["columns"]

    # Step 4: Find layer structure by looking for large limit values
    layers = _identify_layers(blocks, ws)

    # Step 5: For each layer, find carrier blocks and their associated data
    entries = []
    layer_summaries = []
    for layer in layers:
        layer_entries = _extract_layer_data(ws, blocks, layer, summary_cols)
        entries.extend(layer_entries)

        # Extract layer summary data from summary columns (for cross-checking)
        summary = _extract_layer_summary(ws, layer, summary_info)
        if summary:
            layer_summaries.append(summary)

    return entries, layer_summaries


# =============================================================================
# Block Discovery
# =============================================================================


def _find_all_blocks(ws) -> list[Block]:
    """Find all merged cell regions and significant single cells.

    Uses pre-built merged cell map for O(1) lookups instead of O(n) iteration
    for each cell, significantly improving performance on large spreadsheets.
    """
    blocks = []
    processed = set()

    # Pre-build merged cell map for O(1) lookups
    merged_map = build_merged_cell_map(ws)

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if (row, col) in processed:
                continue

            val = get_cell_value(ws, row, col, merged_map)
            if val is None or (isinstance(val, str) and not val.strip()):
                processed.add((row, col))
                continue

            # Check for merged range using O(1) lookup
            merged = merged_map.get((row, col))
            if merged:
                min_c, min_r, max_c, max_r = range_boundaries(merged)
                # Only process from top-left
                if row != min_r or col != min_c:
                    processed.add((row, col))
                    continue

                block = Block(
                    row=min_r, col=min_c, value=val, rows=max_r - min_r + 1, cols=max_c - min_c + 1
                )
                # Mark all cells in merge as processed
                for r in range(min_r, max_r + 1):
                    for c in range(min_c, max_c + 1):
                        processed.add((r, c))
            else:
                block = Block(row=row, col=col, value=val)
                processed.add((row, col))

            blocks.append(block)

    return blocks


# =============================================================================
# Layer Identification
# =============================================================================


def _has_conflicting_label(block: Block, other: Block) -> bool:
    """Check if another block indicates this block is not a layer limit.

    Returns True if the other block contains patterns that indicate the
    current block's row contains data, not layer definitions.
    """
    # Check same-row labels that indicate data rows
    if other.field_type == "label" and other.row == block.row:
        val_lower = str(other.value).lower()
        # "LIMIT" in column A means per-carrier limits, not layer limit
        # "Premium" or "Policy" means this row has data, not layer definitions
        if "premium" in val_lower or val_lower == "limit" or "policy" in val_lower:
            return True

    # Check for year patterns in same row - "2019 Bound", "2018 Marketing" etc.
    # These indicate summary/historical rows, not layer definitions
    if other.row == block.row:
        val_str = str(other.value)
        if re.match(r"^20\d{2}\b", val_str):  # Starts with year like 2019, 2018
            return True

    # Check the row ABOVE - if it contains "Premium" or "Share Premium" labels,
    # this row contains premium data, not layer limits
    if other.field_type == "label" and other.row == block.row - 1:
        val_lower = str(other.value).lower()
        if "premium" in val_lower or "participation" in val_lower:
            return True

    return False


def _should_skip_large_number_block(block: Block, blocks: list[Block]) -> bool:
    """Check if a large_number block should be skipped as a layer limit candidate."""
    return any(_has_conflicting_label(block, other) for other in blocks)


def _is_valid_limit_block(block: Block, blocks: list[Block]) -> bool:
    """Check if a block is a valid layer limit candidate."""
    # Must be limit or large_number type with sufficient confidence
    if block.field_type not in ("limit", "large_number") or block.confidence < 0.7:
        return False

    # Only consider blocks in leftmost columns (A or B) for layer limits
    if block.col > MAX_LAYER_LIMIT_COLUMN:
        return False

    # Filter out aggregate totals - numbers > $1B are almost never layer limits
    if isinstance(block.value, int | float) and block.value > BILLION:
        return False

    # For large_number type, check for conflicting labels nearby
    if block.field_type == "large_number" and _should_skip_large_number_block(block, blocks):
        return False

    return True


def _identify_layers(blocks: list[Block], ws) -> list[dict]:
    """Identify layer boundaries from limit blocks."""
    # Find blocks that look like layer limits
    # Layer limits should be:
    # 1. In column A or B (leftmost columns)
    # 2. Formatted as dollar amounts with M/K suffix, OR
    # 3. Large numbers that are explicitly labeled as limits
    limit_blocks = [b for b in blocks if _is_valid_limit_block(b, blocks)]

    # Sort by row
    limit_blocks.sort(key=lambda b: b.row)

    # Filter to keep only "primary" limits (leftmost in their row region)
    primary_limits = _filter_primary_limits(limit_blocks)

    # Build layer info from primary limits
    return _build_layer_info(primary_limits, ws.max_row)


def _is_dominated_by_existing(block: Block, existing_limits: list[Block]) -> bool:
    """Check if a block is dominated by an existing limit in its row region."""
    return any(
        abs(block.row - existing.row) <= LAYER_ROW_PROXIMITY and block.col > existing.col
        for existing in existing_limits
    )


def _filter_primary_limits(limit_blocks: list[Block]) -> list[Block]:
    """Filter limit blocks to keep only the primary (leftmost) in each row region."""
    primary_limits = []
    for block in limit_blocks:
        if not _is_dominated_by_existing(block, primary_limits):
            primary_limits.append(block)
    return primary_limits


def _build_layer_info(primary_limits: list[Block], max_row: int) -> list[dict]:
    """Build layer info dictionaries from primary limit blocks."""
    layers = []
    for i, block in enumerate(primary_limits):
        end_row = primary_limits[i + 1].row - 1 if i + 1 < len(primary_limits) else max_row
        layers.append(
            {
                "limit": _format_limit(block.value),
                "limit_row": block.row,
                "limit_col": block.col,
                "start_row": block.row,
                "end_row": end_row,
            }
        )
    return layers


def _format_limit(value) -> str:
    """Format a limit value for display."""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, int | float):
        return parse_limit_value(value)
    return str(value)


# =============================================================================
# Layer Summary Extraction
# =============================================================================


def _extract_layer_summary(ws, layer: dict, summary_info: dict) -> LayerSummary | None:
    """Extract layer-level summary data from summary columns.

    This data is used for cross-checking that carrier premiums sum to layer totals.

    Args:
        ws: The worksheet
        layer: Layer dict with limit, start_row, end_row
        summary_info: Dict with bound_premium_col, layer_target_col, layer_rate_col

    Returns:
        LayerSummary or None if no summary data found for this layer
    """
    bound_premium_col = summary_info.get("bound_premium_col")
    layer_target_col = summary_info.get("layer_target_col")
    layer_rate_col = summary_info.get("layer_rate_col")

    # If no summary columns detected, skip
    if not any([bound_premium_col, layer_target_col, layer_rate_col]):
        return None

    start_row = layer["start_row"]
    end_row = layer["end_row"]

    layer_bound_premium = None
    layer_target = None
    layer_rate = None
    excel_range = None

    # Look for values in summary columns within this layer's row range
    # Summary values are typically in a specific row for each layer
    for row in range(start_row, end_row + 1):
        if bound_premium_col:
            val = get_cell_value(ws, row, bound_premium_col)
            if val is not None and isinstance(val, int | float) and val > 0:
                layer_bound_premium = float(val)
                excel_range = f"{get_column_letter(bound_premium_col)}{row}"

        if layer_target_col:
            val = get_cell_value(ws, row, layer_target_col)
            if val is not None and isinstance(val, int | float) and val > 0:
                layer_target = float(val)

        if layer_rate_col:
            val = get_cell_value(ws, row, layer_rate_col)
            if val is not None and isinstance(val, int | float):
                # Rate could be very small (e.g., 0.0038)
                layer_rate = float(val)

    # Only return if we found at least one value
    if layer_bound_premium is not None or layer_target is not None or layer_rate is not None:
        return LayerSummary(
            layer_limit=layer["limit"],
            layer_target=layer_target,
            layer_rate=layer_rate,
            layer_bound_premium=layer_bound_premium,
            excel_range=excel_range,
        )

    return None


# =============================================================================
# Layer Data Extraction
# =============================================================================


def _split_multiline_carrier(carrier_block: Block) -> list[tuple[Block, str]]:
    """Split a carrier block containing multiple carriers (newlines) into separate blocks.

    Returns list of tuples: (block, original_cell_ref)
    The original_cell_ref preserves the actual Excel cell for traceability.
    """
    value = carrier_block.value
    original_cell = f"{get_column_letter(carrier_block.col)}{carrier_block.row}"

    if not isinstance(value, str):
        return [(carrier_block, original_cell)]

    # Split on newlines
    lines = [line.strip() for line in value.split("\n") if line.strip()]

    # If only one line, return original
    if len(lines) <= 1:
        return [(carrier_block, original_cell)]

    # Create a new block for each line
    split_blocks = []
    for i, line in enumerate(lines):
        # Skip if it looks like a non-carrier (policy number, label, etc.)
        if _looks_like_policy_number(line):
            continue
        if _is_non_carrier(line):
            continue

        # Create new block with adjusted row for data matching
        # but preserve original cell reference for traceability
        new_block = Block(
            row=carrier_block.row + i,  # Adjust row for data proximity matching
            col=carrier_block.col,
            value=line,
            rows=1,
            cols=carrier_block.cols,
            field_type="carrier",
            confidence=carrier_block.confidence,
        )
        split_blocks.append((new_block, original_cell))

    return split_blocks if split_blocks else [(carrier_block, original_cell)]


def _extract_layer_data(
    ws, blocks: list[Block], layer: dict, summary_cols: set[int] = None
) -> list[CarrierEntry]:
    """Extract carrier entries from a layer region.

    Args:
        ws: The worksheet
        blocks: All classified blocks
        layer: Layer dict with start_row, end_row, limit
        summary_cols: Set of column numbers to exclude (contain aggregate data, not per-carrier)
    """
    entries = []
    start_row = layer["start_row"]
    end_row = layer["end_row"]
    summary_cols = summary_cols or set()

    # Find all blocks in this layer's row range
    layer_blocks = [b for b in blocks if start_row <= b.row <= end_row]

    # Find column headers for this layer
    column_headers = _find_column_headers(ws, blocks, layer)

    # Find row labels for this layer (Premium vs % Premium distinction)
    row_labels = _find_row_labels(ws, blocks, layer)

    # Find carrier blocks, excluding those in summary columns
    carrier_blocks = [
        b
        for b in layer_blocks
        if b.field_type == "carrier"
        and b.confidence >= MIN_CARRIER_CONFIDENCE
        and b.col not in summary_cols
    ]

    # Split multi-line carrier blocks into individual carriers
    # Returns list of (block, original_cell_ref) tuples
    expanded_carriers = []
    for carrier in carrier_blocks:
        expanded_carriers.extend(_split_multiline_carrier(carrier))

    # Find data blocks (percentages, currency, terms)
    # Note: large_number is included because premium values can be > $1M
    # Exclude data from summary columns as well
    data_blocks = [
        b
        for b in layer_blocks
        if b.field_type
        in (
            "percentage",
            "percentage_or_number",
            "currency",
            "currency_string",
            "large_number",
            "zero",
            "terms",
            "layer_description",
        )
        and b.col not in summary_cols
    ]

    # Try to match carriers with their data using spatial proximity
    for carrier, original_cell in expanded_carriers:
        entry = _build_entry_from_proximity(
            ws, carrier, data_blocks, layer, column_headers, row_labels, original_cell
        )
        if entry:
            entries.append(entry)

    return entries


# =============================================================================
# Header and Label Detection
# =============================================================================


def _filter_label_blocks(
    blocks: list[Block],
    row_start: int,
    row_end: int,
    col_min: int | None = None,
    col_max: int | None = None,
) -> list[Block]:
    """Filter blocks to label blocks within row/column bounds."""
    result = []
    for block in blocks:
        if block.row < row_start or block.row > row_end:
            continue
        if block.field_type != "label":
            continue
        if col_min is not None and block.col < col_min:
            continue
        if col_max is not None and block.col > col_max:
            continue
        result.append(block)
    return result


def _find_column_headers(ws, blocks: list[Block], layer: dict) -> dict:
    """Find column header positions for Premium, Limit, Rate, TIV, etc.

    Looks in two places:
    1. Rows near/above the layer start (global headers)
    2. Within the layer (sub-headers for complex schematics)
    """
    headers = {}

    # First pass: look in rows near or above the layer start (traditional layout)
    search_start = max(1, layer["start_row"] - 5)
    search_end = layer["start_row"] + 3
    header_blocks = _filter_label_blocks(blocks, search_start, search_end)
    for block in header_blocks:
        val_lower = str(block.value).lower().strip()
        _classify_column_header(val_lower, block.col, headers)

    # Second pass: look for sub-headers within the layer (for complex schematics)
    # Look for specific headers that appear mid-layer as data table column headers
    # These are typically in columns beyond A/B (which contain row labels)
    layer_blocks = _filter_label_blocks(blocks, layer["start_row"], layer["end_row"], col_min=3)
    for block in layer_blocks:
        val_lower = str(block.value).lower().strip()
        _classify_sub_header(val_lower, block.col, headers)

    return headers


def _classify_sub_header(val_lower: str, col: int, headers: dict) -> None:
    """Classify a sub-header within a layer."""
    if val_lower == "rate" and "rate_col" not in headers:
        headers["rate_col"] = col
    elif val_lower == "premium" and col > 2:
        headers["premium_col"] = col
    elif ("tiv" in val_lower or val_lower == "updated tiv") and "tiv_col" not in headers:
        headers["tiv_col"] = col
        if "tiv_data_col" not in headers:
            headers["tiv_data_col"] = col + 1


def _classify_column_header(val_lower: str, col: int, headers: dict) -> None:
    """Classify a column header and update headers dict."""
    if "premium" in val_lower and "limit" not in val_lower:
        if "% premium" in val_lower or "share" in val_lower:
            if "premium_share_col" not in headers:
                headers["premium_share_col"] = col
        else:
            if "premium_col" not in headers:
                headers["premium_col"] = col
    elif "limit" in val_lower:
        if "limit_col" not in headers:
            headers["limit_col"] = col
    elif "participation" in val_lower or "% share" in val_lower or val_lower == "share":
        if "participation_col" not in headers:
            headers["participation_col"] = col
    elif val_lower == "rate":
        if "rate_col" not in headers:
            headers["rate_col"] = col
    elif "tiv" in val_lower or val_lower == "updated tiv":
        if "tiv_col" not in headers:
            headers["tiv_col"] = col
            if "tiv_data_col" not in headers:
                headers["tiv_data_col"] = col + 1


def _find_row_labels(ws, blocks: list[Block], layer: dict) -> dict:
    """Find row label positions for Premium, % Premium, etc.

    Many schematics use row labels in column A to indicate what each row contains.
    Some schematics (like Super Hard) repeat labels per carrier column.
    This helps distinguish between "Premium" row and "% Premium" row.
    """
    labels = {}
    start_row = layer["start_row"]
    end_row = layer["end_row"]

    # First pass: look for labels in columns A or B (traditional layout)
    ab_blocks = _filter_label_blocks(blocks, start_row, end_row, col_max=2)
    for block in ab_blocks:
        val_lower = str(block.value).lower().strip()
        _classify_row_label(val_lower, block.row, labels)

    # Second pass: if we didn't find key labels, look in any column
    # (for schematics where labels repeat per carrier column)
    if "premium_row" not in labels or "participation_row" not in labels:
        all_blocks = _filter_label_blocks(blocks, start_row, end_row)
        for block in all_blocks:
            val_lower = str(block.value).lower().strip()
            _classify_row_label(val_lower, block.row, labels)

    return labels


def _classify_row_label(val_lower: str, row: int, labels: dict) -> None:
    """Classify a row label and update the labels dict if not already set."""
    # Identify row types by labels
    if "premium" in val_lower:
        if "% premium" in val_lower or val_lower.startswith("%"):
            if "percent_premium_row" not in labels:
                labels["percent_premium_row"] = row
        elif val_lower == "premium" or val_lower == "share premium" or "layer premium" in val_lower:
            if "premium_row" not in labels:
                labels["premium_row"] = row
    elif val_lower == "limit":
        # "LIMIT" row often contains premium/limit values per carrier
        if "limit_row" not in labels:
            labels["limit_row"] = row
    elif "participation" in val_lower or "% share" in val_lower or val_lower == "share":
        if "participation_row" not in labels:
            labels["participation_row"] = row
    elif "carrier" in val_lower:
        if "carrier_row" not in labels:
            labels["carrier_row"] = row
    elif "layer" in val_lower and "premium" not in val_lower:
        if "layer_row" not in labels:
            labels["layer_row"] = row
    elif "terms" in val_lower:
        if "terms_row" not in labels:
            labels["terms_row"] = row
    elif "policy" in val_lower:
        # Track policy number row so we don't mistake it for premium
        if "policy_row" not in labels:
            labels["policy_row"] = row


# =============================================================================
# Entry Building
# =============================================================================


def _build_entry_from_proximity(
    ws,
    carrier: Block,
    data_blocks: list[Block],
    layer: dict,
    column_headers: dict = None,
    row_labels: dict = None,
    original_cell: str = None,
) -> CarrierEntry:
    """Build a CarrierEntry by finding spatially related data blocks.

    Args:
        ws: The worksheet
        carrier: The carrier block to build an entry for
        data_blocks: List of data blocks to search for related data
        layer: Layer dict with limit, start_row, end_row
        column_headers: Dict of column header positions
        row_labels: Dict of row label positions
        original_cell: The actual Excel cell reference (for multi-line carriers)

    Returns:
        CarrierEntry with extracted data
    """
    # Define search ranges
    carrier_col_range = get_column_range(carrier)
    row_range = range(carrier.row, min(carrier.row + MAX_ROW_SEARCH_DISTANCE, layer["end_row"] + 1))

    # Initialize extracted values
    participation = None
    premium = None
    premium_share = None
    terms = None
    layer_desc = None

    # Get rate column for filtering (Rate != participation)
    rate_col = column_headers.get("rate_col") if column_headers else None

    # Sort data blocks by proximity to carrier
    sorted_blocks = sorted(
        data_blocks,
        key=lambda b: calculate_block_proximity(b, carrier, carrier_col_range),
    )

    # Process each block
    for block in sorted_blocks:
        # Skip blocks that aren't relevant to this carrier
        if not is_block_relevant(block, carrier, carrier_col_range, row_range):
            continue

        # Match percentage blocks (participation)
        if block.field_type in ("percentage", "percentage_or_number") and participation is None:
            matched = match_participation_block(block, row_labels, rate_col)
            if matched is not None:
                participation = matched

        # Match currency blocks (premium)
        elif block.field_type in ("currency", "currency_string", "large_number", "zero"):
            premium, premium_share = match_currency_block(
                block, column_headers, row_labels, premium, premium_share
            )

        # Match terms
        elif block.field_type == "terms" and terms is None:
            terms = str(block.value).strip()

        # Match layer description
        elif block.field_type == "layer_description" and layer_desc is None:
            layer_desc = str(block.value).strip()

    # Build cell reference
    cell_ref = original_cell if original_cell else f"{get_column_letter(carrier.col)}{carrier.row}"

    # Parse attachment point from carrier name or layer description
    carrier_name = str(carrier.value).strip()
    _, attachment_point = parse_excess_notation(carrier_name)
    if not attachment_point and layer_desc:
        _, attachment_point = parse_excess_notation(layer_desc)

    return CarrierEntry(
        layer_limit=layer["limit"],
        layer_description=layer_desc or "",
        carrier=carrier_name,
        participation_pct=participation,
        premium=premium,
        premium_share=premium_share,
        terms=terms,
        policy_number=None,
        excel_range=cell_ref,
        col_span=carrier.cols,
        row_span=carrier.rows,
        fill_color=get_cell_color(ws, carrier.row, carrier.col),
        attachment_point=attachment_point,
    )


# =============================================================================
# Workbook Loading
# =============================================================================


def _load_workbook(filepath: str, sheet_name: str | None = None) -> Any:
    """Load workbook and return worksheet, with proper error handling.

    Args:
        filepath: Path to Excel file
        sheet_name: Optional sheet name

    Returns:
        Worksheet object (openpyxl.worksheet.worksheet.Worksheet)

    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If sheet name is invalid or file format is unsupported
    """
    from pathlib import Path

    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException

    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    if path.suffix.lower() not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        raise ValueError(
            f"Unsupported file format: {path.suffix}. Expected .xlsx, .xlsm, .xltx, or .xltm"
        )

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except InvalidFileException as e:
        raise ValueError(f"Invalid or corrupted Excel file: {filepath}") from e

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {available}")
        return wb[sheet_name]
    return wb.active


# =============================================================================
# Public API
# =============================================================================


def extract_schematic(filepath: str, sheet_name: str | None = None) -> list[dict]:
    """Extract insurance tower schematic data from an Excel file.

    This is the main entry point for the library. It analyzes the spreadsheet
    structure to extract carrier participation data from insurance tower diagrams.

    Args:
        filepath: Path to the Excel file (.xlsx)
        sheet_name: Optional sheet name (uses active sheet if not specified)

    Returns:
        List of carrier entry dictionaries, each containing:
        - layer_limit: Layer limit (e.g., "$250M")
        - layer_description: Layer description if any
        - carrier: Carrier name
        - participation_pct: Participation percentage (0-1 scale)
        - premium: Premium amount
        - excel_range: Source cell reference (e.g., "H47")
        - col_span, row_span: Cell span information
        - fill_color: Cell background color if any
        - attachment_point: Parsed from "xs." notation

    Raises:
        FileNotFoundError: If the file doesn't exist
        ValueError: If the file format is invalid or sheet name not found

    Example:
        >>> from schematic_explorer import extract_schematic
        >>> entries = extract_schematic("tower.xlsx")
        >>> for entry in entries:
        ...     print(f"{entry['carrier']}: {entry['participation_pct']}")
    """
    ws = _load_workbook(filepath, sheet_name)
    entries, _ = extract_adaptive(ws)
    return [entry.to_dict() for entry in entries]


def extract_schematic_with_summaries(
    filepath: str, sheet_name: str | None = None
) -> tuple[list[dict], list[dict]]:
    """Extract schematic data along with layer summaries for cross-checking.

    Args:
        filepath: Path to the Excel file
        sheet_name: Optional sheet name

    Returns:
        Tuple of (carrier_entries, layer_summaries)

    Raises:
        FileNotFoundError: If the file doesn't exist
        ValueError: If the file format is invalid or sheet name not found
    """
    ws = _load_workbook(filepath, sheet_name)
    entries, summaries = extract_adaptive(ws)
    return ([entry.to_dict() for entry in entries], [summary.to_dict() for summary in summaries])

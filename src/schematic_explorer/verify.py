"""AI-powered verification of extracted tower data using Gemini."""

import json
import logging
import os
from pathlib import Path

import google.generativeai as genai
from dotenv import load_dotenv
from PIL import Image

from .types import CarrierEntry, LayerSummary, VerificationError, VerificationResult

# =============================================================================
# Configuration Constants
# =============================================================================

# Default model for verification (flash-lite has thinking OFF, so use flash)
DEFAULT_GEMINI_MODEL = "gemini-2.5-flash"

# Generation temperature (0 for deterministic output)
GENERATION_TEMPERATURE = 0

# Cross-check thresholds
LAYER_MISSING_PREMIUM_THRESHOLD = 10000  # Flag if layer shows > $10K but no carriers
LAYER_DISCREPANCY_THRESHOLD = 2.0  # Flag if > 200% difference between carrier sum and layer total
MAX_LAYER_PENALTY = 0.15  # Maximum score reduction for layer discrepancies
PENALTY_PER_DISCREPANCY = 0.05  # Score penalty per discrepancy found

# Merged cell display limit (in Excel-to-text conversion)
MAX_MERGED_CELLS_DISPLAY = 20

# Library logging follows best practice: NullHandler by default
# CLI tools/users can configure logging to see output
logger = logging.getLogger(__name__)
logger.addHandler(logging.NullHandler())

# Output directory for snapshots
OUTPUT_DIR = Path(__file__).parent.parent.parent.parent / "output"


# Schema for structured output (Gemini-compatible format)
VERIFICATION_SCHEMA = {
    "type": "object",
    "properties": {
        "score": {"type": "number", "description": "Accuracy score from 0.0 to 1.0"},
        "summary": {
            "type": "string",
            "description": "Brief 1-2 sentence summary of extraction quality",
        },
        "issues": {
            "type": "array",
            "items": {"type": "string"},
            "description": "List of specific issues found",
        },
        "suggestions": {
            "type": "array",
            "items": {"type": "string"},
            "description": "List of improvement suggestions",
        },
    },
    "required": ["score", "summary", "issues", "suggestions"],
}

SNAPSHOT_VERIFICATION_SCHEMA = {
    "type": "object",
    "properties": {
        "score": {
            "type": "number",
            "description": "Accuracy score from 0.0 to 1.0 based on visual comparison",
        },
        "summary": {"type": "string", "description": "Brief assessment based on visual comparison"},
        "visual_issues": {
            "type": "array",
            "items": {"type": "string"},
            "description": "Discrepancies between image and extracted data",
        },
        "missing_from_extraction": {
            "type": "array",
            "items": {"type": "string"},
            "description": "Data visible in image but not in extraction",
        },
        "false_positives": {
            "type": "array",
            "items": {"type": "string"},
            "description": "Data in extraction that doesn't appear in image",
        },
    },
    "required": ["score", "summary", "visual_issues", "missing_from_extraction", "false_positives"],
}

# Schema for cross-validation pass
CROSS_VALIDATION_SCHEMA = {
    "type": "object",
    "properties": {
        "adjusted_score": {
            "type": "number",
            "description": "Revised accuracy score from 0.0 to 1.0 after reviewing issues",
        },
        "summary": {
            "type": "string",
            "description": "Brief summary of the cross-validation findings",
        },
        "confirmed_issues": {
            "type": "array",
            "items": {"type": "string"},
            "description": "Issues from the first pass that are confirmed as real problems",
        },
        "dismissed_issues": {
            "type": "array",
            "items": {"type": "string"},
            "description": "Issues from the first pass that were false positives (with brief reason)",
        },
        "new_issues": {
            "type": "array",
            "items": {"type": "string"},
            "description": "Any additional issues discovered during cross-validation",
        },
        "suggestions": {
            "type": "array",
            "items": {"type": "string"},
            "description": "Refined suggestions for improving extraction",
        },
    },
    "required": [
        "adjusted_score",
        "summary",
        "confirmed_issues",
        "dismissed_issues",
        "new_issues",
        "suggestions",
    ],
}


def _get_client() -> genai.GenerativeModel:
    """Initialize Gemini client."""
    load_dotenv()
    api_key = os.getenv("GEMINI_API_KEY")
    model_id = os.getenv("GEMINI_MODEL_ID", DEFAULT_GEMINI_MODEL)

    if not api_key:
        raise ValueError("GEMINI_API_KEY not found in environment")

    genai.configure(api_key=api_key)
    return genai.GenerativeModel(model_id)


# Generation config with temperature=0 for deterministic, consistent output
GENERATION_CONFIG = {
    "temperature": GENERATION_TEMPERATURE,
    "response_mime_type": "application/json",
}


def _load_workbook_for_verification(filepath: str, sheet_name: str | None = None):
    """Load workbook and return worksheet for verification.

    Args:
        filepath: Path to Excel file
        sheet_name: Optional sheet name

    Returns:
        Tuple of (worksheet, filename)

    Raises:
        FileNotFoundError: If the file doesn't exist
        ValueError: If the file format is invalid or sheet name not found
    """
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException

    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except InvalidFileException as e:
        raise ValueError(f"Invalid or corrupted Excel file: {filepath}") from e

    if sheet_name and sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {available}")

    ws = wb[sheet_name] if sheet_name else wb.active
    return ws, path.name


def _format_cell_rows(ws) -> list[str]:
    """Format worksheet cell contents as text lines.

    Args:
        ws: Worksheet to format

    Returns:
        List of formatted row strings
    """
    from openpyxl.utils import get_column_letter

    lines = []
    # DO NOT truncate or limit - the LLM needs ALL cell contents to verify accuracy
    for row in range(1, ws.max_row + 1):
        row_cells = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                val_str = str(val).replace("\n", " | ")  # Preserve newlines as pipes
                row_cells.append(f"{get_column_letter(col)}{row}={val_str}")
        if row_cells:
            lines.append(f"Row {row}: {' | '.join(row_cells)}")
    return lines


def _format_merged_cells(ws) -> list[str]:
    """Format merged cell information as text lines.

    Args:
        ws: Worksheet with merged cells

    Returns:
        List of formatted merged cell strings
    """
    lines = []
    if ws.merged_cells.ranges:
        lines.append("")
        lines.append(f"Merged cells: {len(list(ws.merged_cells.ranges))}")
        for mr in list(ws.merged_cells.ranges)[:MAX_MERGED_CELLS_DISPLAY]:
            lines.append(f"  {mr}")
    return lines


def _excel_to_text(filepath: str, sheet_name: str | None = None) -> str:
    """Convert Excel file to text representation for Gemini.

    Raises:
        FileNotFoundError: If the file doesn't exist
        ValueError: If the file format is invalid or sheet name not found
    """
    ws, filename = _load_workbook_for_verification(filepath, sheet_name)

    # Build header
    lines = [f"Excel File: {filename}"]
    if sheet_name:
        lines.append(f"Sheet: {sheet_name}")
    lines.append(f"Dimensions: {ws.dimensions}")
    lines.append("")

    # Add cell contents
    lines.extend(_format_cell_rows(ws))

    # Add merged cell info
    lines.extend(_format_merged_cells(ws))

    return "\n".join(lines)


def _entries_to_text(entries: list[CarrierEntry]) -> str:
    """Convert extracted entries to text for comparison."""
    lines = ["Extracted Data:"]
    lines.append(f"Total entries: {len(entries)}")
    lines.append("")

    # Group by layer
    layers = {}
    for e in entries:
        if e.layer_limit not in layers:
            layers[e.layer_limit] = []
        layers[e.layer_limit].append(e)

    for limit, layer_entries in layers.items():
        lines.append(f"Layer {limit}:")
        for e in layer_entries:
            # Output participation as decimal (same format as Excel) to avoid conversion confusion
            # Note: use "is not None" to handle 0 values correctly (0 is falsy but valid)
            pct = f"{e.participation_pct}" if e.participation_pct is not None else "N/A"
            premium = f"{e.premium}" if e.premium is not None else "N/A"
            # Include excel_range for precise cell-level verification
            cell_ref = f" [cell:{e.excel_range}]" if e.excel_range else ""
            lines.append(f"  - {e.carrier}{cell_ref}: participation={pct}, premium={premium}")
            if e.layer_description:
                lines.append(f"    Description: {e.layer_description}")
        lines.append("")

    return "\n".join(lines)


VERIFICATION_PROMPT = """Compare extracted insurance data against Excel source.

## Excel Data
{excel_content}

## Extracted Data
{extracted_content}

## CRITICAL: Cell Reference Verification
Each extracted entry has a cell reference like [cell:H47].
ALWAYS verify by checking the SAME COLUMN in Excel data.
Do NOT compare values from different columns or different layers.

## VALUE EQUIVALENCE - DO NOT FLAG THESE AS ISSUES:
- Percentages: 0.1 = 10% = 0.10 = 0.1000
- Currency: 100000.0 = 100000 = $100,000 = $100,000.0
- Decimals: 3500000.0 = 3500000 = $3,500,000

## LAYER ISOLATION
Same carrier name may appear in multiple layers with DIFFERENT values.
Only compare within the SAME layer. Different layers = different entries.

## HIDDEN ROWS
Hidden rows contain grouping metadata. Carriers in hidden rows have no participation.

## ONLY REPORT GENUINE ISSUES:
- Mathematically different values (0.1 vs 0.2)
- Missing data that should exist
- Wrong carrier-layer assignments

Score: 0.0 to 1.0"""


VERIFICATION_PROMPT_TEXT_ONLY = """Compare extracted insurance data against Excel source.

## Excel Data
{excel_content}

## Extracted Data
{extracted_content}

## CRITICAL: Cell Reference Verification
Each extracted entry has a cell reference like [cell:H47].
ALWAYS verify by checking the SAME COLUMN in Excel data.
Do NOT compare values from different columns or different layers.

## VALUE EQUIVALENCE - DO NOT FLAG THESE AS ISSUES:
- Percentages: 0.1 = 10% = 0.10 = 0.1000
- Currency: 100000.0 = 100000 = $100,000 = $100,000.0
- Decimals: 3500000.0 = 3500000 = $3,500,000

## LAYER ISOLATION
Same carrier name may appear in multiple layers with DIFFERENT values.
Only compare within the SAME layer. Different layers = different entries.

## HIDDEN ROWS
Hidden rows contain grouping metadata. Carriers in hidden rows have no participation.

## ONLY REPORT GENUINE ISSUES:
- Mathematically different values (0.1 vs 0.2)
- Missing data that should exist
- Wrong carrier-layer assignments

Score: 0.0 to 1.0"""


SNAPSHOT_VERIFICATION_PROMPT = """You are verifying extracted insurance tower data against this Excel spreadsheet image.

## Extracted Data
{extracted_content}

## IMPORTANT RULES

### Rule 1: Decimal = Percentage (NEVER flag as error)
- 0.1 = 10%, 0.5 = 50%, 0.25 = 25%
- If extraction shows "50%" and image shows "0.5" → CORRECT
- If extraction shows "10.0%" and image shows "0.1" → CORRECT

### Rule 2: Only flag ACTUAL differences
- 10% vs 20% → ERROR (different numbers)
- 10% vs 0.1 → NOT AN ERROR (same number)

### Rule 3: Don't flag identical values
If extraction matches Excel exactly, it's NOT an issue.

## Your Task
Look at the image and verify the extracted data. Only report genuine discrepancies where values are mathematically different or data is missing/wrong.

Score: 0.0 (completely wrong) to 1.0 (perfect)"""


CROSS_VALIDATION_PROMPT = """Second-pass cross-validation of extracted insurance data.

## Excel Data
{excel_content}

## Extracted Data
{extracted_content}

## First-Pass Findings
Score: {initial_score:.0%}
Summary: {initial_summary}

Issues identified:
{issues_list}

## CRITICAL VERIFICATION RULES

### Rule 1: Cell Reference Matching (MOST IMPORTANT)
Each extracted entry includes a cell reference like [cell:H47].
To verify a carrier's participation:
1. Find the carrier's cell reference (e.g., H47)
2. Look at the SAME COLUMN in the % SHARE row (e.g., H48)
3. Compare ONLY that specific cell's value
4. Do NOT compare to values in different columns (J48, K48, etc.)

### Rule 2: Layer Isolation
Insurance towers have multiple layers ($250M, $150M, $100M, $50M, $25M).
- The SAME carrier may appear in MULTIPLE layers with DIFFERENT values
- "Lexington" in $25M layer (H47/H48) is DIFFERENT from "Lexington" in $250M layer (J7/J8)
- ONLY compare values within the SAME layer
- DISMISS any issue comparing a carrier across different layers

### Rule 3: Hidden Rows
Hidden rows (often rows 60+) contain grouping metadata, not carrier data.
- Carriers in hidden rows should have NULL/N/A participation
- DISMISS any issue about hidden row carriers having wrong values

### Rule 4: Value Equivalence - ALWAYS DISMISS
- 0.1 = 10% = 0.10 = 0.1000 (ALL EQUIVALENT)
- 0.225 = 22.5% = 0.2250 (ALL EQUIVALENT)
- 100000 = 100000.0 (EQUIVALENT)
- DISMISS any issue comparing equivalent values

### Rule 5: Column Alignment Check
If an issue claims "carrier X has 0.2 but Excel shows 0.09":
1. Find carrier X's cell reference in extracted data
2. Verify the "Excel shows" value comes from the SAME column
3. If comparing different columns, DISMISS the issue

## VERIFICATION PROTOCOL
For EACH issue:
1. Identify the carrier and its cell reference [cell:XX]
2. Find that exact cell's participation (same column, % SHARE row)
3. Compare: extracted value vs that specific cell
4. DISMISS if: values match, different columns, different layers, or equivalent formats

Be AGGRESSIVE about dismissing false positives. When in doubt, DISMISS."""


def _get_snapshot_path(filepath: str) -> Path | None:
    """Get the snapshot image path for an Excel file."""
    excel_path = Path(filepath)
    snapshot_path = OUTPUT_DIR / f"{excel_path.stem}.png"
    if snapshot_path.exists():
        return snapshot_path
    return None


def _parse_json_response(raw_response: str) -> dict:
    """Parse JSON from Gemini response, handling common issues."""
    import re

    text = raw_response.strip()

    # Handle markdown code blocks
    if text.startswith("```"):
        parts = text.split("```")
        if len(parts) >= 2:
            text = parts[1]
            if text.startswith("json"):
                text = text[4:]
    text = text.strip()

    # Try direct parse first
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # Try to fix common escape issues
    # Replace problematic backslashes that aren't valid escapes
    text_fixed = re.sub(r'\\(?!["\\/bfnrtu])', r"\\\\", text)
    try:
        return json.loads(text_fixed)
    except json.JSONDecodeError:
        pass

    # Try to extract just the JSON object
    match = re.search(r"\{[\s\S]*\}", text)
    if match:
        json_str = match.group()
        # Fix escapes in extracted JSON
        json_str = re.sub(r'\\(?!["\\/bfnrtu])', r"\\\\", json_str)
        try:
            return json.loads(json_str)
        except json.JSONDecodeError:
            pass

    # Last resort: try to build a minimal valid response
    score_match = re.search(r'"score"\s*:\s*([\d.]+)', text)
    summary_match = re.search(r'"summary"\s*:\s*"([^"]*)"', text)

    if score_match:
        return {
            "score": float(score_match.group(1)),
            "summary": summary_match.group(1) if summary_match else "Partial parse",
            "issues": [],
            "suggestions": [],
        }

    raise ValueError("Could not parse JSON response")


def _make_gemini_request(
    model,
    prompt: str,
    schema: dict,
    image: Image.Image | None = None,
    context: str = "request",
) -> tuple[dict, str, dict]:
    """Make a Gemini request with structured output and fallback parsing.

    Args:
        model: Gemini model instance
        prompt: The prompt text
        schema: JSON schema for structured output
        image: Optional PIL Image for multimodal requests
        context: Context string for logging (e.g., "verify_extraction")

    Returns:
        Tuple of (parsed_data, raw_response, metadata)

    Raises:
        Exception: If both structured output and fallback parsing fail
    """
    generation_config = {
        "temperature": GENERATION_TEMPERATURE,
        "response_mime_type": "application/json",
        "response_schema": schema,
    }

    content = [prompt, image] if image else prompt
    structured_error_msg = None

    # Try structured output first
    try:
        response = model.generate_content(content, generation_config=generation_config)
        raw_response = response.text
        data = json.loads(raw_response)
        logger.info("%s: structured output parsed successfully", context)
        return data, raw_response, {"parsing_method": "structured", "fallback_used": False}
    except Exception as e:
        structured_error_msg = str(e)
        logger.warning(
            "%s: structured output failed (%s), using fallback parser",
            context,
            structured_error_msg,
        )

    # Fallback: try without schema enforcement
    try:
        response = model.generate_content(content if image else prompt)
        raw_response = response.text
        data = _parse_json_response(raw_response)
        logger.info("%s: fallback parser succeeded", context)
        return (
            data,
            raw_response,
            {
                "parsing_method": "fallback",
                "fallback_used": True,
                "structured_error": structured_error_msg,
            },
        )
    except Exception as fallback_error:
        logger.error("%s: fallback parser also failed (%s)", context, str(fallback_error))
        raise VerificationError(
            f"Both structured output and fallback parsing failed: {structured_error_msg}, {fallback_error}"
        ) from fallback_error


def verify_extraction(
    filepath: str, entries: list[CarrierEntry], sheet_name: str | None = None
) -> VerificationResult:
    """
    Verify extracted data against original Excel using Gemini with structured output.

    Args:
        filepath: Path to the original Excel file
        entries: Extracted CarrierEntry objects
        sheet_name: Optional sheet name

    Returns:
        VerificationResult with score and details
    """
    model = _get_client()

    excel_content = _excel_to_text(filepath, sheet_name)
    extracted_content = _entries_to_text(entries)

    # Check if we have a snapshot image
    snapshot_path = _get_snapshot_path(filepath)

    # Build prompt based on whether we have an image
    if snapshot_path:
        prompt = VERIFICATION_PROMPT.format(
            excel_content=excel_content, extracted_content=extracted_content
        )
        image = Image.open(snapshot_path)
    else:
        prompt = VERIFICATION_PROMPT_TEXT_ONLY.format(
            excel_content=excel_content, extracted_content=extracted_content
        )
        image = None

    try:
        data, raw_response, metadata = _make_gemini_request(
            model, prompt, VERIFICATION_SCHEMA, image, "verify_extraction"
        )
        return VerificationResult(
            score=float(data.get("score", 0)),
            summary=data.get("summary", "No summary"),
            issues=data.get("issues", []),
            suggestions=data.get("suggestions", []),
            raw_response=raw_response,
            metadata=metadata,
        )
    except Exception as e:
        return VerificationResult(
            score=0.0,
            summary=f"Verification failed: {e}",
            issues=[str(e)],
            suggestions=[],
            raw_response=str(e),
            metadata={"parsing_method": "error", "fallback_used": True, "error": str(e)},
        )


def _convert_snapshot_issues(data: dict) -> list[str]:
    """Convert snapshot-specific fields to standard issue format."""
    issues = list(data.get("visual_issues", []))
    issues.extend([f"Missing: {m}" for m in data.get("missing_from_extraction", [])])
    issues.extend([f"False positive: {f}" for f in data.get("false_positives", [])])
    return issues


def verify_snapshot(
    filepath: str,
    entries: list[CarrierEntry],
) -> VerificationResult | None:
    """
    Verify extracted data against the visual snapshot only using structured output.

    Args:
        filepath: Path to the original Excel file
        entries: Extracted CarrierEntry objects

    Returns:
        VerificationResult or None if no snapshot available
    """
    snapshot_path = _get_snapshot_path(filepath)
    if not snapshot_path:
        return None

    model = _get_client()
    extracted_content = _entries_to_text(entries)
    prompt = SNAPSHOT_VERIFICATION_PROMPT.format(extracted_content=extracted_content)
    image = Image.open(snapshot_path)

    try:
        data, raw_response, metadata = _make_gemini_request(
            model, prompt, SNAPSHOT_VERIFICATION_SCHEMA, image, "verify_snapshot"
        )
        issues = _convert_snapshot_issues(data)
        return VerificationResult(
            score=float(data.get("score", 0)),
            summary=data.get("summary", "Visual verification complete"),
            issues=issues,
            suggestions=[],
            raw_response=raw_response,
            metadata=metadata,
        )
    except Exception as e:
        return VerificationResult(
            score=0.0,
            summary=f"Snapshot verification failed: {e}",
            issues=[str(e)],
            suggestions=[],
            raw_response=str(e),
            metadata={"parsing_method": "error", "fallback_used": True, "error": str(e)},
        )


def cross_validate(
    filepath: str,
    entries: list[CarrierEntry],
    initial_result: VerificationResult,
    sheet_name: str | None = None,
) -> VerificationResult:
    """
    Cross-validate the first-pass verification to filter false positives.

    Args:
        filepath: Path to the original Excel file
        entries: Extracted CarrierEntry objects
        initial_result: Result from first verification pass
        sheet_name: Optional sheet name

    Returns:
        VerificationResult with refined findings
    """
    snapshot_path = _get_snapshot_path(filepath)
    if not snapshot_path:
        # No snapshot available, return initial result as-is
        return initial_result

    model = _get_client()
    excel_content = _excel_to_text(filepath, sheet_name)
    extracted_content = _entries_to_text(entries)

    # Format issues and suggestions for the prompt
    issues_list = (
        "\n".join(f"- {issue}" for issue in initial_result.issues)
        if initial_result.issues
        else "None identified"
    )
    suggestions_list = (
        "\n".join(f"- {s}" for s in initial_result.suggestions)
        if initial_result.suggestions
        else "None"
    )

    prompt = CROSS_VALIDATION_PROMPT.format(
        excel_content=excel_content,
        extracted_content=extracted_content,
        initial_score=initial_result.score,
        initial_summary=initial_result.summary,
        issues_list=issues_list,
        suggestions_list=suggestions_list,
    )

    generation_config = {
        "temperature": GENERATION_TEMPERATURE,
        "response_mime_type": "application/json",
        "response_schema": CROSS_VALIDATION_SCHEMA,
    }

    try:
        image = Image.open(snapshot_path)
        response = model.generate_content([prompt, image], generation_config=generation_config)
        raw_response = response.text
        data = json.loads(raw_response)

        # Combine confirmed and new issues
        all_issues = list(data.get("confirmed_issues", []))
        all_issues.extend(data.get("new_issues", []))

        # Build summary noting how many issues were dismissed
        dismissed_count = len(data.get("dismissed_issues", []))
        summary = data.get("summary", "Cross-validation complete")
        if dismissed_count > 0:
            summary += f" ({dismissed_count} false positives filtered)"

        return VerificationResult(
            score=float(data.get("adjusted_score", initial_result.score)),
            summary=summary,
            issues=all_issues,
            suggestions=data.get("suggestions", []),
            raw_response=f"INITIAL:\n{initial_result.raw_response}\n\nCROSS-VALIDATION:\n{raw_response}",
        )
    except Exception as e:
        # If cross-validation fails, return initial result
        return VerificationResult(
            score=initial_result.score,
            summary=f"{initial_result.summary} (cross-validation failed: {e})",
            issues=initial_result.issues,
            suggestions=initial_result.suggestions,
            raw_response=initial_result.raw_response,
        )


def _calculate_discrepancy_pct(expected: float, actual: float) -> float:
    """Calculate discrepancy percentage between expected and actual values."""
    if expected > 0:
        return abs(expected - actual) / expected
    return 0.0 if actual == 0 else 1.0


def _check_missing_carriers(
    layer_limit: str, summary: LayerSummary, actual: float
) -> str | None:
    """Check if layer has no carriers but should have data.

    Args:
        layer_limit: The layer limit identifier
        summary: Layer summary with expected values
        actual: Actual carrier premium total

    Returns:
        Issue string or None if no issue
    """
    expected = summary.layer_bound_premium
    if actual == 0 and expected > LAYER_MISSING_PREMIUM_THRESHOLD:
        return (
            f"Layer {layer_limit}: No carrier premiums extracted but "
            f"summary shows ${expected:,.0f} (cell {summary.excel_range}) - possible extraction gap"
        )
    return None


def _check_extreme_discrepancy(
    layer_limit: str, summary: LayerSummary, actual: float, discrepancy_pct: float
) -> str | None:
    """Check for extreme discrepancy between carrier sum and layer total.

    Args:
        layer_limit: The layer limit identifier
        summary: Layer summary with expected values
        actual: Actual carrier premium total
        discrepancy_pct: Calculated discrepancy percentage

    Returns:
        Issue string or None if no issue
    """
    expected = summary.layer_bound_premium
    if discrepancy_pct > LAYER_DISCREPANCY_THRESHOLD:
        return (
            f"Layer {layer_limit}: Carrier premiums ${actual:,.0f} vs "
            f"summary ${expected:,.0f} (cell {summary.excel_range}) - "
            f"{discrepancy_pct:.0%} difference (may be prior year data)"
        )
    return None


def _build_carrier_totals_by_layer(entries: list[CarrierEntry]) -> dict[str, float]:
    """Group carrier entries by layer and sum premiums.

    Args:
        entries: List of carrier entries

    Returns:
        Dict mapping layer limit to total premium
    """
    totals = {}
    for entry in entries:
        layer = entry.layer_limit
        if layer not in totals:
            totals[layer] = 0.0
        if entry.premium is not None:
            totals[layer] += entry.premium
    return totals


def cross_check_layer_totals(
    entries: list[CarrierEntry], layer_summaries: list[LayerSummary], result: VerificationResult
) -> VerificationResult:
    """
    Cross-check extracted carrier premiums against layer summary totals.

    For each layer with a summary, sum the carrier premiums and compare
    to the layer_bound_premium. Large discrepancies indicate extraction issues.

    Args:
        entries: Extracted carrier entries
        layer_summaries: Layer-level summary data from summary columns
        result: Current verification result to augment

    Returns:
        Updated VerificationResult with any layer total issues
    """
    issues = list(result.issues)
    suggestions = list(result.suggestions)

    # Build lookup structures
    summary_by_layer = {s.layer_limit: s for s in layer_summaries}
    carrier_totals_by_layer = _build_carrier_totals_by_layer(entries)

    # Cross-check each layer
    discrepancies_found = 0

    for layer_limit, summary in summary_by_layer.items():
        if summary.layer_bound_premium is None:
            continue

        actual = carrier_totals_by_layer.get(layer_limit, 0.0)
        discrepancy_pct = _calculate_discrepancy_pct(summary.layer_bound_premium, actual)

        # Check for missing carriers
        issue = _check_missing_carriers(layer_limit, summary, actual)
        if issue:
            discrepancies_found += 1
            issues.append(issue)
            continue

        # Check for extreme discrepancy
        issue = _check_extreme_discrepancy(layer_limit, summary, actual, discrepancy_pct)
        if issue:
            discrepancies_found += 1
            issues.append(issue)

    # Adjust score only for severe issues
    score = result.score
    if discrepancies_found > 0:
        penalty = min(PENALTY_PER_DISCREPANCY * discrepancies_found, MAX_LAYER_PENALTY)
        score = max(0.0, score - penalty)
        suggestions.append(
            f"Review {discrepancies_found} layer(s) with significant carrier/summary differences "
            "(note: summary columns may show prior year data)"
        )

    return VerificationResult(
        score=score,
        summary=result.summary,
        issues=issues,
        suggestions=suggestions,
        raw_response=result.raw_response,
    )


def verify_file(filepath: str, sheet_name: str | None = None) -> VerificationResult:
    """
    Extract and verify a single file using two-pass cross-validation.

    Pass 1: Initial verification comparing extracted data to source
    Pass 2: Cross-validation to filter false positives and confirm real issues
    Pass 3: Layer total cross-check using summary columns (if available)

    Args:
        filepath: Path to Excel file
        sheet_name: Optional sheet name

    Returns:
        VerificationResult with cross-validated findings
    """
    from .extractor import extract_schematic_with_summaries

    entries, layer_summaries = extract_schematic_with_summaries(filepath, sheet_name)
    if not entries:
        return VerificationResult(
            score=0.0,
            summary="No data extracted from file",
            issues=["Extraction returned empty results"],
            suggestions=["Check if file format is supported"],
            raw_response="",
        )

    # Pass 1: Initial verification (text + image if available)
    initial_result = verify_extraction(filepath, entries, sheet_name)

    # Pass 2: Cross-validation to filter false positives
    final_result = cross_validate(filepath, entries, initial_result, sheet_name)

    # Pass 3: Cross-check layer totals against summary columns (if available)
    if layer_summaries:
        final_result = cross_check_layer_totals(entries, layer_summaries, final_result)

    return final_result

"""Utility functions for Excel cell operations."""

from typing import TYPE_CHECKING, Any

from openpyxl.utils import range_boundaries

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


def build_merged_cell_map(ws: "Worksheet") -> dict[tuple[int, int], str]:
    """Pre-build a map of (row, col) -> merged_range_string for O(1) lookups.

    This is much more efficient than iterating through all merged ranges
    for every cell lookup.

    Args:
        ws: The worksheet to build the map for

    Returns:
        Dict mapping (row, col) tuples to merged range strings
    """
    merged_map: dict[tuple[int, int], str] = {}
    for merged_range in ws.merged_cells.ranges:
        range_str = str(merged_range)
        min_col, min_row, max_col, max_row = range_boundaries(range_str)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_map[(row, col)] = range_str
    return merged_map


def find_merged_range_at(
    ws: "Worksheet", row: int, col: int, merged_map: dict[tuple[int, int], str] | None = None
) -> str | None:
    """Find if a cell is part of a merged range, return the range string.

    Args:
        ws: The worksheet
        row: Row number
        col: Column number
        merged_map: Optional pre-built merged cell map for O(1) lookups.
                   If not provided, falls back to O(n) iteration.
    """
    # Use pre-built map if available (O(1) lookup)
    if merged_map is not None:
        return merged_map.get((row, col))

    # Fallback to iteration (O(n) where n = number of merged ranges)
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return str(merged_range)
    return None


def get_cell_value(
    ws: "Worksheet", row: int, col: int, merged_map: dict[tuple[int, int], str] | None = None
) -> Any:
    """Get cell value, handling merged cells (value is in top-left).

    Args:
        ws: The worksheet
        row: Row number
        col: Column number
        merged_map: Optional pre-built merged cell map for O(1) lookups
    """
    merged = find_merged_range_at(ws, row, col, merged_map)
    if merged:
        min_col, min_row, _, _ = range_boundaries(merged)
        return ws.cell(row=min_row, column=min_col).value
    return ws.cell(row=row, column=col).value


def get_cell_color(
    ws: "Worksheet", row: int, col: int, merged_map: dict[tuple[int, int], str] | None = None
) -> str | None:
    """Get background color of a cell as hex string.

    Args:
        ws: The worksheet
        row: Row number
        col: Column number
        merged_map: Optional pre-built merged cell map for O(1) lookups
    """
    merged = find_merged_range_at(ws, row, col, merged_map)
    if merged:
        min_col, min_row, _, _ = range_boundaries(merged)
        cell = ws.cell(row=min_row, column=min_col)
    else:
        cell = ws.cell(row=row, column=col)

    if cell.fill and cell.fill.fgColor:
        color = cell.fill.fgColor
        if color.type == "rgb" and color.rgb and color.rgb not in ("00000000", "000000"):
            return color.rgb
    return None

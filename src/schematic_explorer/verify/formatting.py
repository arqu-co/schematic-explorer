"""Excel and data formatting utilities for verification."""

from pathlib import Path

from ..types import CarrierEntry

# Merged cell display limit (in Excel-to-text conversion)
MAX_MERGED_CELLS_DISPLAY = 20


def load_workbook_for_verification(filepath: str, sheet_name: str | None = None):
    """Load workbook and return worksheet for verification.

    Args:
        filepath: Path to Excel file
        sheet_name: Optional sheet name

    Returns:
        Tuple of (worksheet, filename)

    Raises:
        FileNotFoundError: If the file doesn't exist
        ValueError: If the file format is invalid or sheet name not found
    """
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException

    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except InvalidFileException as e:
        raise ValueError(f"Invalid or corrupted Excel file: {filepath}") from e

    if sheet_name and sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {available}")

    ws = wb[sheet_name] if sheet_name else wb.active
    return ws, path.name


def format_cell_rows(ws) -> list[str]:
    """Format worksheet cell contents as text lines.

    Args:
        ws: Worksheet to format

    Returns:
        List of formatted row strings
    """
    from openpyxl.utils import get_column_letter

    lines = []
    # DO NOT truncate or limit - the LLM needs ALL cell contents to verify accuracy
    for row in range(1, ws.max_row + 1):
        row_cells = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                val_str = str(val).replace("\n", " | ")  # Preserve newlines as pipes
                row_cells.append(f"{get_column_letter(col)}{row}={val_str}")
        if row_cells:
            lines.append(f"Row {row}: {' | '.join(row_cells)}")
    return lines


def format_merged_cells(ws) -> list[str]:
    """Format merged cell information as text lines.

    Args:
        ws: Worksheet with merged cells

    Returns:
        List of formatted merged cell strings
    """
    lines = []
    if ws.merged_cells.ranges:
        lines.append("")
        lines.append(f"Merged cells: {len(list(ws.merged_cells.ranges))}")
        for mr in list(ws.merged_cells.ranges)[:MAX_MERGED_CELLS_DISPLAY]:
            lines.append(f"  {mr}")
    return lines


def excel_to_text(filepath: str, sheet_name: str | None = None) -> str:
    """Convert Excel file to text representation for Gemini.

    Raises:
        FileNotFoundError: If the file doesn't exist
        ValueError: If the file format is invalid or sheet name not found
    """
    ws, filename = load_workbook_for_verification(filepath, sheet_name)

    # Build header
    lines = [f"Excel File: {filename}"]
    if sheet_name:
        lines.append(f"Sheet: {sheet_name}")
    lines.append(f"Dimensions: {ws.dimensions}")
    lines.append("")

    # Add cell contents
    lines.extend(format_cell_rows(ws))

    # Add merged cell info
    lines.extend(format_merged_cells(ws))

    return "\n".join(lines)


def entries_to_text(entries: list[CarrierEntry]) -> str:
    """Convert extracted entries to text for comparison."""
    lines = ["Extracted Data:"]
    lines.append(f"Total entries: {len(entries)}")
    lines.append("")

    # Group by layer
    layers: dict[str, list[CarrierEntry]] = {}
    for e in entries:
        if e.layer_limit not in layers:
            layers[e.layer_limit] = []
        layers[e.layer_limit].append(e)

    for limit, layer_entries in layers.items():
        lines.append(f"Layer {limit}:")
        for e in layer_entries:
            # Output participation as decimal (same format as Excel) to avoid conversion confusion
            # Note: use "is not None" to handle 0 values correctly (0 is falsy but valid)
            pct = f"{e.participation_pct}" if e.participation_pct is not None else "N/A"
            premium = f"{e.premium}" if e.premium is not None else "N/A"
            # Include excel_range for precise cell-level verification
            cell_ref = f" [cell:{e.excel_range}]" if e.excel_range else ""
            lines.append(f"  - {e.carrier}{cell_ref}: participation={pct}, premium={premium}")
            if e.layer_description:
                lines.append(f"    Description: {e.layer_description}")
        lines.append("")

    return "\n".join(lines)

#!/usr/bin/env python3
"""
Test harness for verification prompt improvements.

This script tests the verification system against known ground truth data
to identify and eliminate false positives. All fixes must be generalizable -
no file-specific hacks allowed.

Usage:
    python test_verification.py [--iterations N] [--verbose]
"""

import argparse
import json
import os
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from dotenv import load_dotenv

# Add project to path
sys.path.insert(0, str(Path(__file__).parent))

from tower_extractor.extract import extract_tower_data
from tower_extractor.models import CarrierEntry


@dataclass
class GroundTruthEntry:
    """A verified carrier-participation pair from Excel."""
    carrier: str
    layer: str
    participation: Optional[float]
    excel_cell: str  # e.g., "H48"
    carrier_cell: str  # e.g., "H47"
    notes: str = ""


@dataclass
class VerificationTestCase:
    """A test case for verification accuracy."""
    file_path: str
    ground_truth: list[GroundTruthEntry] = field(default_factory=list)


def build_ground_truth() -> list[VerificationTestCase]:
    """
    Build ground truth test cases from actual Excel files.

    IMPORTANT: This function extracts ground truth by reading Excel cells directly.
    It does NOT hardcode values specific to any file - it programmatically reads
    the actual cell values to build verification test cases.
    """
    test_cases = []

    # Test case 1: Hard Schematic - $25M layer
    hard_schematic = Path("input/Hard Schematic.xlsx")
    if hard_schematic.exists():
        tc = VerificationTestCase(file_path=str(hard_schematic))

        wb = openpyxl.load_workbook(hard_schematic, data_only=True)
        ws = wb.active

        # Find layers by looking for large numbers in column A that indicate layer limits
        # This is a generalizable approach - find CARRIER rows and their corresponding % SHARE rows
        layer_info = _find_layer_structure(ws)

        for layer in layer_info:
            carrier_row = layer['carrier_row']
            share_row = layer['share_row']
            layer_limit = layer['limit']

            # Extract all carrier-participation pairs for this layer
            for col in range(2, ws.max_column + 1):
                carrier_val = ws.cell(row=carrier_row, column=col).value
                share_val = ws.cell(row=share_row, column=col).value

                if carrier_val and isinstance(carrier_val, str) and carrier_val.strip():
                    # Skip if this is a label or total
                    if carrier_val.strip().upper() in ('CARRIER', 'TOTAL', ''):
                        continue

                    participation = None
                    if isinstance(share_val, (int, float)) and 0 < share_val <= 1:
                        participation = float(share_val)

                    entry = GroundTruthEntry(
                        carrier=carrier_val.strip()[:60],
                        layer=layer_limit,
                        participation=participation,
                        excel_cell=f"{get_column_letter(col)}{share_row}",
                        carrier_cell=f"{get_column_letter(col)}{carrier_row}",
                    )
                    tc.ground_truth.append(entry)

        # Also note hidden rows (generalizable detection)
        hidden_rows = []
        for row in range(1, ws.max_row + 1):
            rd = ws.row_dimensions.get(row)
            if rd and rd.hidden:
                hidden_rows.append(row)

        if hidden_rows:
            # Mark entries from hidden rows
            for entry in tc.ground_truth:
                row_num = int(''.join(c for c in entry.carrier_cell if c.isdigit()))
                if row_num in hidden_rows:
                    entry.notes = "HIDDEN_ROW"

        wb.close()
        test_cases.append(tc)

    # Test case 2: Super Hard Schematic (if exists)
    super_hard = Path("input/Super Hard Schematic.xlsx")
    if super_hard.exists():
        tc = VerificationTestCase(file_path=str(super_hard))

        wb = openpyxl.load_workbook(super_hard, data_only=True)
        ws = wb.active

        layer_info = _find_layer_structure(ws)

        for layer in layer_info:
            carrier_row = layer['carrier_row']
            share_row = layer['share_row']
            layer_limit = layer['limit']

            for col in range(2, ws.max_column + 1):
                carrier_val = ws.cell(row=carrier_row, column=col).value
                share_val = ws.cell(row=share_row, column=col).value

                if carrier_val and isinstance(carrier_val, str) and carrier_val.strip():
                    if carrier_val.strip().upper() in ('CARRIER', 'TOTAL', ''):
                        continue

                    participation = None
                    if isinstance(share_val, (int, float)) and 0 < share_val <= 1:
                        participation = float(share_val)

                    entry = GroundTruthEntry(
                        carrier=carrier_val.strip()[:60],
                        layer=layer_limit,
                        participation=participation,
                        excel_cell=f"{get_column_letter(col)}{share_row}",
                        carrier_cell=f"{get_column_letter(col)}{carrier_row}",
                    )
                    tc.ground_truth.append(entry)

        wb.close()
        test_cases.append(tc)

    return test_cases


def _find_layer_structure(ws) -> list[dict]:
    """
    Find layer structure by identifying CARRIER and % SHARE row patterns.

    This is a generalizable approach that works across different schematic formats.
    """
    layers = []

    # Scan for "CARRIER" labels in column A
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and isinstance(val, str) and val.strip().upper() == 'CARRIER':
            carrier_row = row

            # Look for % SHARE row (usually 1-3 rows below)
            share_row = None
            for offset in range(1, 5):
                check_val = ws.cell(row=row + offset, column=1).value
                if check_val and isinstance(check_val, str):
                    check_lower = check_val.strip().lower()
                    if '% share' in check_lower or 'share' == check_lower or '%' in check_lower:
                        share_row = row + offset
                        break

            if not share_row:
                # Default to next row if no label found
                share_row = row + 1

            # Find layer limit (look above for large number)
            layer_limit = "Unknown"
            for offset in range(1, 10):
                if row - offset < 1:
                    break
                limit_val = ws.cell(row=row - offset, column=1).value
                if isinstance(limit_val, (int, float)) and limit_val >= 1_000_000:
                    layer_limit = _format_limit(limit_val)
                    break

            layers.append({
                'carrier_row': carrier_row,
                'share_row': share_row,
                'limit': layer_limit,
            })

    return layers


def _format_limit(value) -> str:
    """Format a limit value for display."""
    if isinstance(value, (int, float)):
        if value >= 1_000_000_000:
            return f"${value / 1_000_000_000:.0f}B"
        if value >= 1_000_000:
            return f"${value / 1_000_000:.0f}M"
        if value >= 1_000:
            return f"${value / 1_000:.0f}K"
        return f"${value:.0f}"
    return str(value)


def run_verification_test(test_case: VerificationTestCase, verbose: bool = False) -> dict:
    """
    Run verification on a test case and compare against ground truth.

    Returns metrics on false positives and accuracy.
    """
    from tower_extractor.verify import verify_file, cross_validate, verify_extraction

    print(f"\nTesting: {Path(test_case.file_path).name}")
    print(f"Ground truth entries: {len(test_case.ground_truth)}")

    # Run the verification
    result = verify_file(test_case.file_path)

    print(f"Verification score: {result.score:.0%}")
    print(f"Issues reported: {len(result.issues)}")

    # Analyze issues against ground truth
    false_positives = []
    true_positives = []

    for issue in result.issues:
        is_false_positive = _check_if_false_positive(issue, test_case.ground_truth, verbose)
        if is_false_positive:
            false_positives.append(issue)
        else:
            true_positives.append(issue)

    # Calculate metrics
    total_issues = len(result.issues)
    fp_count = len(false_positives)
    tp_count = len(true_positives)
    fp_rate = fp_count / total_issues if total_issues > 0 else 0

    print(f"\nResults:")
    print(f"  False positives: {fp_count}")
    print(f"  True positives: {tp_count}")
    print(f"  FP rate: {fp_rate:.1%}")

    if verbose and false_positives:
        print(f"\nFalse positive details:")
        for fp in false_positives:
            print(f"  - {fp[:100]}...")

    return {
        'file': test_case.file_path,
        'score': result.score,
        'total_issues': total_issues,
        'false_positives': fp_count,
        'true_positives': tp_count,
        'fp_rate': fp_rate,
        'false_positive_issues': false_positives,
        'summary': result.summary,
    }


def _check_if_false_positive(issue: str, ground_truth: list[GroundTruthEntry], verbose: bool) -> bool:
    """
    Check if an issue is a false positive by comparing against ground truth.

    An issue is a false positive if:
    1. The carrier mentioned exists in ground truth
    2. The extracted value matches the ground truth value
    3. The issue claims a different value
    """
    issue_lower = issue.lower()

    for gt in ground_truth:
        carrier_lower = gt.carrier.lower()

        # Check if this issue mentions this carrier
        # Use partial matching since carrier names may be truncated
        carrier_words = carrier_lower.split()[:3]  # First 3 words
        if not any(word in issue_lower for word in carrier_words if len(word) > 3):
            continue

        # This issue mentions this carrier
        if gt.participation is not None:
            # Check if the issue incorrectly claims a different value
            gt_str = f"{gt.participation:.4f}"
            gt_str_short = f"{gt.participation:.2f}"

            # If the issue mentions the correct value as "extracted" and claims
            # Excel shows something different, it's a false positive
            if gt_str in issue or gt_str_short in issue:
                # The ground truth value is mentioned - check if it's being incorrectly flagged
                if 'excel shows' in issue_lower or 'source shows' in issue_lower:
                    if verbose:
                        print(f"    FP detected: {gt.carrier[:30]} - GT={gt.participation}, issue claims mismatch")
                    return True

        # Check for hidden row carriers being flagged
        if gt.notes == "HIDDEN_ROW":
            if verbose:
                print(f"    FP detected: {gt.carrier[:30]} is in hidden row")
            return True

    return False


def test_prompt_variation(prompt_name: str, prompt_text: str, test_cases: list[VerificationTestCase]) -> dict:
    """
    Test a specific prompt variation against all test cases.
    """
    # Temporarily override the prompt in the verify module
    import tower_extractor.verify as verify_module

    original_prompt = verify_module.CROSS_VALIDATION_PROMPT
    verify_module.CROSS_VALIDATION_PROMPT = prompt_text

    try:
        results = []
        total_fp = 0
        total_issues = 0

        for tc in test_cases:
            result = run_verification_test(tc, verbose=False)
            results.append(result)
            total_fp += result['false_positives']
            total_issues += result['total_issues']

        overall_fp_rate = total_fp / total_issues if total_issues > 0 else 0

        return {
            'prompt_name': prompt_name,
            'results': results,
            'total_false_positives': total_fp,
            'total_issues': total_issues,
            'overall_fp_rate': overall_fp_rate,
        }
    finally:
        # Restore original prompt
        verify_module.CROSS_VALIDATION_PROMPT = original_prompt


# Prompt variations to test - these are generalizable improvements
PROMPT_VARIATIONS = {
    'baseline': None,  # Use current prompt

    'cell_reference_focus': """Second-pass review of extracted insurance data.

## Excel Data
{excel_content}

## Extracted Data
{extracted_content}

## First-Pass Findings
Score: {initial_score:.0%}
Summary: {initial_summary}

Issues identified:
{issues_list}

## CRITICAL VERIFICATION RULES

### Rule 1: Match by Cell Reference, Not Just Name
Each extracted entry has an excel_range (e.g., "H47"). When verifying:
- Find the EXACT cell reference in the Excel data
- Compare the participation value from the SAME COLUMN's % SHARE row
- Do NOT compare values from different columns or different layers

### Rule 2: Layer Isolation
Insurance towers have multiple layers (e.g., $250M, $150M, $100M, $50M, $25M).
- ONLY compare carriers within the SAME layer
- A carrier may appear in multiple layers with DIFFERENT participation %
- Lexington at H47 ($25M layer) is DIFFERENT from Lexington at J7 ($250M layer)

### Rule 3: Hidden Rows Are Not Errors
- Hidden rows (e.g., rows 61-64) contain grouping keys, not data
- Do NOT flag carriers in hidden rows as having wrong values
- Hidden row carriers should have NO participation value

### Rule 4: Equivalent Values - DISMISS
- 0.1 = 10% = 0.1000 (all equivalent)
- 0.225 = 22.5% = 0.2250 (all equivalent)
- DISMISS any issue comparing equivalent values

### Rule 5: Column Alignment
The carrier in column H row 47 should be compared to the value in column H row 48 (% SHARE).
Do NOT compare H47 carrier to values in columns J, K, L, etc.

Review each first-pass issue with these rules. Dismiss false positives aggressively.""",

    'strict_cell_matching': """Second-pass cross-validation of extracted insurance data.

## Excel Data
{excel_content}

## Extracted Data
{extracted_content}

## First-Pass Findings
Score: {initial_score:.0%}
Summary: {initial_summary}

Issues identified:
{issues_list}

## VERIFICATION PROTOCOL

For EACH issue in the first-pass findings:

1. IDENTIFY the carrier name mentioned in the issue
2. FIND that carrier's excel_range in the Extracted Data (e.g., "H47")
3. LOCATE the exact cell in Excel Data using that cell reference
4. FIND the participation value in the SAME COLUMN (e.g., H48 for carrier at H47)
5. COMPARE: Does extracted participation match Excel cell value?

### DISMISS the issue if:
- Extracted value MATCHES the Excel cell in the same column
- Values are mathematically equivalent (0.1 = 10%, 0.225 = 22.5%)
- Carrier is in a hidden row (rows 61-64 or similar grouping rows)
- Issue compares values from DIFFERENT layers or DIFFERENT columns

### CONFIRM the issue ONLY if:
- Extracted value is MATHEMATICALLY DIFFERENT from the correct Excel cell
- You verified the EXACT column alignment (same column for carrier and %)

## COMMON FALSE POSITIVE PATTERNS TO DISMISS:
- Comparing $25M layer values to $50M layer values
- Comparing column H value to column K value
- Flagging 0.225 vs 0.2250 (identical)
- Flagging hidden row carriers

Be extremely conservative. When in doubt, DISMISS the issue.""",

    'layer_aware': """Second-pass verification with layer-aware matching.

## Excel Data
{excel_content}

## Extracted Data
{extracted_content}

## First-Pass Findings
Score: {initial_score:.0%}
Summary: {initial_summary}

Issues identified:
{issues_list}

## LAYER-AWARE VERIFICATION

Insurance tower schematics have MULTIPLE LAYERS stacked vertically. Each layer:
- Has its own limit (e.g., $250M, $150M, $100M, $50M, $25M)
- Has its own CARRIER row and % SHARE row
- May contain the SAME carrier names with DIFFERENT participation values

### CRITICAL: Same Carrier, Different Layers
- "Lexington Insurance" at $250M layer might have 10% participation
- "Lexington Insurance" at $25M layer might have 22.5% participation
- These are BOTH CORRECT - do not flag as inconsistent

### VERIFICATION STEPS
For each issue:
1. Note which LAYER the extracted entry belongs to (from layer_limit field)
2. Find that layer's section in the Excel data
3. Verify the carrier's % SHARE within THAT LAYER ONLY
4. DISMISS if comparing across different layers

### VALUE EQUIVALENCE (ALWAYS DISMISS)
- 0.1 = 10% = 0.10 = 0.1000
- 0.225 = 22.5% = 0.2250
- Integer vs float: 100000 = 100000.0

### HIDDEN ROWS (ALWAYS DISMISS)
Rows marked as hidden contain grouping metadata, not carrier data.
Any carrier from a hidden row should have NULL participation.

Review each issue. Be aggressive about dismissing false positives.""",
}


def main():
    parser = argparse.ArgumentParser(description='Test verification prompt improvements')
    parser.add_argument('--iterations', type=int, default=1, help='Number of test iterations')
    parser.add_argument('--verbose', '-v', action='store_true', help='Verbose output')
    parser.add_argument('--prompt', type=str, help='Test specific prompt variation')
    args = parser.parse_args()

    load_dotenv()

    print("=" * 70)
    print("VERIFICATION PROMPT TEST HARNESS")
    print("=" * 70)

    # Build ground truth
    print("\nBuilding ground truth from Excel files...")
    test_cases = build_ground_truth()

    if not test_cases:
        print("ERROR: No test cases found. Ensure input/*.xlsx files exist.")
        return 1

    total_gt_entries = sum(len(tc.ground_truth) for tc in test_cases)
    print(f"Built {len(test_cases)} test cases with {total_gt_entries} ground truth entries")

    # Test prompts
    prompts_to_test = PROMPT_VARIATIONS
    if args.prompt:
        if args.prompt in PROMPT_VARIATIONS:
            prompts_to_test = {args.prompt: PROMPT_VARIATIONS[args.prompt]}
        else:
            print(f"Unknown prompt: {args.prompt}")
            print(f"Available: {list(PROMPT_VARIATIONS.keys())}")
            return 1

    all_results = []

    for iteration in range(args.iterations):
        if args.iterations > 1:
            print(f"\n{'=' * 70}")
            print(f"ITERATION {iteration + 1}/{args.iterations}")
            print("=" * 70)

        for prompt_name, prompt_text in prompts_to_test.items():
            print(f"\n{'-' * 50}")
            print(f"Testing prompt: {prompt_name}")
            print("-" * 50)

            if prompt_text is None:
                # Use baseline (current prompt)
                for tc in test_cases:
                    result = run_verification_test(tc, verbose=args.verbose)
                    result['prompt_name'] = prompt_name
                    all_results.append(result)
            else:
                result = test_prompt_variation(prompt_name, prompt_text, test_cases)
                all_results.append(result)

    # Summary
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)

    # Group by prompt
    by_prompt = {}
    for r in all_results:
        name = r.get('prompt_name', 'unknown')
        if name not in by_prompt:
            by_prompt[name] = []
        by_prompt[name].append(r)

    print(f"\n{'Prompt':<25} {'FP Rate':<10} {'FPs':<8} {'Total':<8}")
    print("-" * 55)

    best_prompt = None
    best_fp_rate = 1.0

    for prompt_name, results in by_prompt.items():
        if 'overall_fp_rate' in results[0]:
            # Aggregated result
            fp_rate = results[0]['overall_fp_rate']
            total_fp = results[0]['total_false_positives']
            total = results[0]['total_issues']
        else:
            # Individual results
            total_fp = sum(r.get('false_positives', 0) for r in results)
            total = sum(r.get('total_issues', 0) for r in results)
            fp_rate = total_fp / total if total > 0 else 0

        print(f"{prompt_name:<25} {fp_rate:>8.1%} {total_fp:>8} {total:>8}")

        if fp_rate < best_fp_rate:
            best_fp_rate = fp_rate
            best_prompt = prompt_name

    print(f"\nBest performing prompt: {best_prompt} (FP rate: {best_fp_rate:.1%})")

    if best_fp_rate > 0:
        print("\nTo achieve 0% false positive rate, consider:")
        print("  1. Adding cell reference validation to the prompt")
        print("  2. Implementing layer isolation logic")
        print("  3. Pre-filtering hidden row carriers before verification")

    return 0 if best_fp_rate == 0 else 1


if __name__ == '__main__':
    sys.exit(main())

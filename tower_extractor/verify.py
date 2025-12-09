"""AI-powered verification of extracted tower data using Gemini."""

import json
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import google.generativeai as genai
from dotenv import load_dotenv
from PIL import Image

from .models import CarrierEntry

# Output directory for snapshots
OUTPUT_DIR = Path(__file__).parent.parent / "output"


# Schema for structured output (Gemini-compatible format)
VERIFICATION_SCHEMA = {
    "type": "object",
    "properties": {
        "score": {
            "type": "number",
            "description": "Accuracy score from 0.0 to 1.0"
        },
        "summary": {
            "type": "string",
            "description": "Brief 1-2 sentence summary of extraction quality"
        },
        "issues": {
            "type": "array",
            "items": {"type": "string"},
            "description": "List of specific issues found"
        },
        "suggestions": {
            "type": "array",
            "items": {"type": "string"},
            "description": "List of improvement suggestions"
        }
    },
    "required": ["score", "summary", "issues", "suggestions"]
}

SNAPSHOT_VERIFICATION_SCHEMA = {
    "type": "object",
    "properties": {
        "score": {
            "type": "number",
            "description": "Accuracy score from 0.0 to 1.0 based on visual comparison"
        },
        "summary": {
            "type": "string",
            "description": "Brief assessment based on visual comparison"
        },
        "visual_issues": {
            "type": "array",
            "items": {"type": "string"},
            "description": "Discrepancies between image and extracted data"
        },
        "missing_from_extraction": {
            "type": "array",
            "items": {"type": "string"},
            "description": "Data visible in image but not in extraction"
        },
        "false_positives": {
            "type": "array",
            "items": {"type": "string"},
            "description": "Data in extraction that doesn't appear in image"
        }
    },
    "required": ["score", "summary", "visual_issues", "missing_from_extraction", "false_positives"]
}

# Schema for cross-validation pass
CROSS_VALIDATION_SCHEMA = {
    "type": "object",
    "properties": {
        "adjusted_score": {
            "type": "number",
            "description": "Revised accuracy score from 0.0 to 1.0 after reviewing issues"
        },
        "summary": {
            "type": "string",
            "description": "Brief summary of the cross-validation findings"
        },
        "confirmed_issues": {
            "type": "array",
            "items": {"type": "string"},
            "description": "Issues from the first pass that are confirmed as real problems"
        },
        "dismissed_issues": {
            "type": "array",
            "items": {"type": "string"},
            "description": "Issues from the first pass that were false positives (with brief reason)"
        },
        "new_issues": {
            "type": "array",
            "items": {"type": "string"},
            "description": "Any additional issues discovered during cross-validation"
        },
        "suggestions": {
            "type": "array",
            "items": {"type": "string"},
            "description": "Refined suggestions for improving extraction"
        }
    },
    "required": ["adjusted_score", "summary", "confirmed_issues", "dismissed_issues", "new_issues", "suggestions"]
}


@dataclass
class VerificationResult:
    """Result of verification check."""
    score: float  # 0.0 to 1.0
    summary: str
    issues: list[str]
    suggestions: list[str]
    raw_response: str


def _get_client():
    """Initialize Gemini client."""
    load_dotenv()
    api_key = os.getenv("GEMINI_API_KEY")
    # Default to gemini-2.5-flash for better reasoning (flash-lite has thinking OFF)
    model_id = os.getenv("GEMINI_MODEL_ID", "gemini-2.5-flash")

    if not api_key:
        raise ValueError("GEMINI_API_KEY not found in environment")

    genai.configure(api_key=api_key)
    return genai.GenerativeModel(model_id)


# Generation config with temperature=0 for deterministic, consistent output
GENERATION_CONFIG = {
    "temperature": 0,
    "response_mime_type": "application/json",
}


def _excel_to_text(filepath: str, sheet_name: Optional[str] = None) -> str:
    """Convert Excel file to text representation for Gemini."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    lines = [f"Excel File: {Path(filepath).name}"]
    if sheet_name:
        lines.append(f"Sheet: {sheet_name}")
    lines.append(f"Dimensions: {ws.dimensions}")
    lines.append("")

    # Output cell contents in a readable format
    # DO NOT truncate or limit - the LLM needs ALL cell contents to verify accuracy
    for row in range(1, ws.max_row + 1):
        row_cells = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                val_str = str(val).replace('\n', ' | ')  # Preserve newlines as pipes
                row_cells.append(f"{get_column_letter(col)}{row}={val_str}")
        if row_cells:
            lines.append(f"Row {row}: {' | '.join(row_cells)}")

    # Add merged cell info
    if ws.merged_cells.ranges:
        lines.append("")
        lines.append(f"Merged cells: {len(list(ws.merged_cells.ranges))}")
        for i, mr in enumerate(list(ws.merged_cells.ranges)[:20]):
            lines.append(f"  {mr}")

    return "\n".join(lines)


def _entries_to_text(entries: list[CarrierEntry]) -> str:
    """Convert extracted entries to text for comparison."""
    lines = ["Extracted Data:"]
    lines.append(f"Total entries: {len(entries)}")
    lines.append("")

    # Group by layer
    layers = {}
    for e in entries:
        if e.layer_limit not in layers:
            layers[e.layer_limit] = []
        layers[e.layer_limit].append(e)

    for limit, layer_entries in layers.items():
        lines.append(f"Layer {limit}:")
        for e in layer_entries:
            # Output participation as decimal (same format as Excel) to avoid conversion confusion
            # Note: use "is not None" to handle 0 values correctly (0 is falsy but valid)
            pct = f"{e.participation_pct}" if e.participation_pct is not None else "N/A"
            premium = f"{e.premium}" if e.premium is not None else "N/A"
            # Include excel_range for precise cell-level verification
            cell_ref = f" [cell:{e.excel_range}]" if e.excel_range else ""
            lines.append(f"  - {e.carrier}{cell_ref}: participation={pct}, premium={premium}")
            if e.layer_description:
                lines.append(f"    Description: {e.layer_description}")
        lines.append("")

    return "\n".join(lines)


VERIFICATION_PROMPT = """Compare extracted insurance data against Excel source.

## Excel Data
{excel_content}

## Extracted Data
{extracted_content}

## CRITICAL: Cell Reference Verification
Each extracted entry has a cell reference like [cell:H47].
ALWAYS verify by checking the SAME COLUMN in Excel data.
Do NOT compare values from different columns or different layers.

## VALUE EQUIVALENCE - DO NOT FLAG THESE AS ISSUES:
- Percentages: 0.1 = 10% = 0.10 = 0.1000
- Currency: 100000.0 = 100000 = $100,000 = $100,000.0
- Decimals: 3500000.0 = 3500000 = $3,500,000

## LAYER ISOLATION
Same carrier name may appear in multiple layers with DIFFERENT values.
Only compare within the SAME layer. Different layers = different entries.

## HIDDEN ROWS
Hidden rows contain grouping metadata. Carriers in hidden rows have no participation.

## ONLY REPORT GENUINE ISSUES:
- Mathematically different values (0.1 vs 0.2)
- Missing data that should exist
- Wrong carrier-layer assignments

Score: 0.0 to 1.0"""


VERIFICATION_PROMPT_TEXT_ONLY = """Compare extracted insurance data against Excel source.

## Excel Data
{excel_content}

## Extracted Data
{extracted_content}

## CRITICAL: Cell Reference Verification
Each extracted entry has a cell reference like [cell:H47].
ALWAYS verify by checking the SAME COLUMN in Excel data.
Do NOT compare values from different columns or different layers.

## VALUE EQUIVALENCE - DO NOT FLAG THESE AS ISSUES:
- Percentages: 0.1 = 10% = 0.10 = 0.1000
- Currency: 100000.0 = 100000 = $100,000 = $100,000.0
- Decimals: 3500000.0 = 3500000 = $3,500,000

## LAYER ISOLATION
Same carrier name may appear in multiple layers with DIFFERENT values.
Only compare within the SAME layer. Different layers = different entries.

## HIDDEN ROWS
Hidden rows contain grouping metadata. Carriers in hidden rows have no participation.

## ONLY REPORT GENUINE ISSUES:
- Mathematically different values (0.1 vs 0.2)
- Missing data that should exist
- Wrong carrier-layer assignments

Score: 0.0 to 1.0"""


SNAPSHOT_VERIFICATION_PROMPT = """You are verifying extracted insurance tower data against this Excel spreadsheet image.

## Extracted Data
{extracted_content}

## IMPORTANT RULES

### Rule 1: Decimal = Percentage (NEVER flag as error)
- 0.1 = 10%, 0.5 = 50%, 0.25 = 25%
- If extraction shows "50%" and image shows "0.5" → CORRECT
- If extraction shows "10.0%" and image shows "0.1" → CORRECT

### Rule 2: Only flag ACTUAL differences
- 10% vs 20% → ERROR (different numbers)
- 10% vs 0.1 → NOT AN ERROR (same number)

### Rule 3: Don't flag identical values
If extraction matches Excel exactly, it's NOT an issue.

## Your Task
Look at the image and verify the extracted data. Only report genuine discrepancies where values are mathematically different or data is missing/wrong.

Score: 0.0 (completely wrong) to 1.0 (perfect)"""


CROSS_VALIDATION_PROMPT = """Second-pass cross-validation of extracted insurance data.

## Excel Data
{excel_content}

## Extracted Data
{extracted_content}

## First-Pass Findings
Score: {initial_score:.0%}
Summary: {initial_summary}

Issues identified:
{issues_list}

## CRITICAL VERIFICATION RULES

### Rule 1: Cell Reference Matching (MOST IMPORTANT)
Each extracted entry includes a cell reference like [cell:H47].
To verify a carrier's participation:
1. Find the carrier's cell reference (e.g., H47)
2. Look at the SAME COLUMN in the % SHARE row (e.g., H48)
3. Compare ONLY that specific cell's value
4. Do NOT compare to values in different columns (J48, K48, etc.)

### Rule 2: Layer Isolation
Insurance towers have multiple layers ($250M, $150M, $100M, $50M, $25M).
- The SAME carrier may appear in MULTIPLE layers with DIFFERENT values
- "Lexington" in $25M layer (H47/H48) is DIFFERENT from "Lexington" in $250M layer (J7/J8)
- ONLY compare values within the SAME layer
- DISMISS any issue comparing a carrier across different layers

### Rule 3: Hidden Rows
Hidden rows (often rows 60+) contain grouping metadata, not carrier data.
- Carriers in hidden rows should have NULL/N/A participation
- DISMISS any issue about hidden row carriers having wrong values

### Rule 4: Value Equivalence - ALWAYS DISMISS
- 0.1 = 10% = 0.10 = 0.1000 (ALL EQUIVALENT)
- 0.225 = 22.5% = 0.2250 (ALL EQUIVALENT)
- 100000 = 100000.0 (EQUIVALENT)
- DISMISS any issue comparing equivalent values

### Rule 5: Column Alignment Check
If an issue claims "carrier X has 0.2 but Excel shows 0.09":
1. Find carrier X's cell reference in extracted data
2. Verify the "Excel shows" value comes from the SAME column
3. If comparing different columns, DISMISS the issue

## VERIFICATION PROTOCOL
For EACH issue:
1. Identify the carrier and its cell reference [cell:XX]
2. Find that exact cell's participation (same column, % SHARE row)
3. Compare: extracted value vs that specific cell
4. DISMISS if: values match, different columns, different layers, or equivalent formats

Be AGGRESSIVE about dismissing false positives. When in doubt, DISMISS."""


def _get_snapshot_path(filepath: str) -> Optional[Path]:
    """Get the snapshot image path for an Excel file."""
    excel_path = Path(filepath)
    snapshot_path = OUTPUT_DIR / f"{excel_path.stem}.png"
    if snapshot_path.exists():
        return snapshot_path
    return None


def _parse_json_response(raw_response: str) -> dict:
    """Parse JSON from Gemini response, handling common issues."""
    import re

    text = raw_response.strip()

    # Handle markdown code blocks
    if text.startswith("```"):
        parts = text.split("```")
        if len(parts) >= 2:
            text = parts[1]
            if text.startswith("json"):
                text = text[4:]
    text = text.strip()

    # Try direct parse first
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # Try to fix common escape issues
    # Replace problematic backslashes that aren't valid escapes
    text_fixed = re.sub(r'\\(?!["\\/bfnrtu])', r'\\\\', text)
    try:
        return json.loads(text_fixed)
    except json.JSONDecodeError:
        pass

    # Try to extract just the JSON object
    match = re.search(r'\{[\s\S]*\}', text)
    if match:
        json_str = match.group()
        # Fix escapes in extracted JSON
        json_str = re.sub(r'\\(?!["\\/bfnrtu])', r'\\\\', json_str)
        try:
            return json.loads(json_str)
        except json.JSONDecodeError:
            pass

    # Last resort: try to build a minimal valid response
    score_match = re.search(r'"score"\s*:\s*([\d.]+)', text)
    summary_match = re.search(r'"summary"\s*:\s*"([^"]*)"', text)

    if score_match:
        return {
            "score": float(score_match.group(1)),
            "summary": summary_match.group(1) if summary_match else "Partial parse",
            "issues": [],
            "suggestions": []
        }

    raise ValueError("Could not parse JSON response")


def verify_extraction(
    filepath: str,
    entries: list[CarrierEntry],
    sheet_name: Optional[str] = None
) -> VerificationResult:
    """
    Verify extracted data against original Excel using Gemini with structured output.

    Args:
        filepath: Path to the original Excel file
        entries: Extracted CarrierEntry objects
        sheet_name: Optional sheet name

    Returns:
        VerificationResult with score and details
    """
    model = _get_client()

    excel_content = _excel_to_text(filepath, sheet_name)
    extracted_content = _entries_to_text(entries)

    # Check if we have a snapshot image
    snapshot_path = _get_snapshot_path(filepath)

    # Configure structured output with temperature=0 for consistency
    generation_config = {
        "temperature": 0,
        "response_mime_type": "application/json",
        "response_schema": VERIFICATION_SCHEMA,
    }

    try:
        if snapshot_path:
            # Use multimodal verification with image
            prompt = VERIFICATION_PROMPT.format(
                excel_content=excel_content,
                extracted_content=extracted_content
            )
            image = Image.open(snapshot_path)
            response = model.generate_content(
                [prompt, image],
                generation_config=generation_config
            )
        else:
            # Fall back to text-only verification
            prompt = VERIFICATION_PROMPT_TEXT_ONLY.format(
                excel_content=excel_content,
                extracted_content=extracted_content
            )
            response = model.generate_content(
                prompt,
                generation_config=generation_config
            )

        raw_response = response.text

        # Parse structured response
        data = json.loads(raw_response)

        return VerificationResult(
            score=float(data.get("score", 0)),
            summary=data.get("summary", "No summary"),
            issues=data.get("issues", []),
            suggestions=data.get("suggestions", []),
            raw_response=raw_response
        )
    except Exception as e:
        # Fallback to legacy parsing if structured output fails
        try:
            if snapshot_path:
                prompt = VERIFICATION_PROMPT.format(
                    excel_content=excel_content,
                    extracted_content=extracted_content
                )
                image = Image.open(snapshot_path)
                response = model.generate_content([prompt, image])
            else:
                prompt = VERIFICATION_PROMPT_TEXT_ONLY.format(
                    excel_content=excel_content,
                    extracted_content=extracted_content
                )
                response = model.generate_content(prompt)

            raw_response = response.text
            data = _parse_json_response(raw_response)

            return VerificationResult(
                score=float(data.get("score", 0)),
                summary=data.get("summary", "No summary provided"),
                issues=data.get("issues", []),
                suggestions=data.get("suggestions", []),
                raw_response=raw_response
            )
        except Exception as fallback_error:
            return VerificationResult(
                score=0.0,
                summary=f"Verification failed: {e}",
                issues=[str(e)],
                suggestions=[],
                raw_response=str(e)
            )


def verify_snapshot(
    filepath: str,
    entries: list[CarrierEntry],
) -> Optional[VerificationResult]:
    """
    Verify extracted data against the visual snapshot only using structured output.

    Args:
        filepath: Path to the original Excel file
        entries: Extracted CarrierEntry objects

    Returns:
        VerificationResult or None if no snapshot available
    """
    snapshot_path = _get_snapshot_path(filepath)
    if not snapshot_path:
        return None

    model = _get_client()
    extracted_content = _entries_to_text(entries)

    prompt = SNAPSHOT_VERIFICATION_PROMPT.format(
        extracted_content=extracted_content
    )

    # Configure structured output with temperature=0 for consistency
    generation_config = {
        "temperature": 0,
        "response_mime_type": "application/json",
        "response_schema": SNAPSHOT_VERIFICATION_SCHEMA,
    }

    try:
        image = Image.open(snapshot_path)
        response = model.generate_content(
            [prompt, image],
            generation_config=generation_config
        )
        raw_response = response.text

        # Parse structured response
        data = json.loads(raw_response)

        # Convert snapshot-specific fields to standard format
        issues = list(data.get("visual_issues", []))
        issues.extend([f"Missing: {m}" for m in data.get("missing_from_extraction", [])])
        issues.extend([f"False positive: {f}" for f in data.get("false_positives", [])])

        return VerificationResult(
            score=float(data.get("score", 0)),
            summary=data.get("summary", "Visual verification complete"),
            issues=issues,
            suggestions=[],
            raw_response=raw_response
        )
    except Exception as e:
        # Fallback to legacy parsing
        try:
            image = Image.open(snapshot_path)
            response = model.generate_content([prompt, image])
            raw_response = response.text
            data = _parse_json_response(raw_response)

            issues = data.get("visual_issues", [])
            issues.extend([f"Missing: {m}" for m in data.get("missing_from_extraction", [])])
            issues.extend([f"False positive: {f}" for f in data.get("false_positives", [])])

            return VerificationResult(
                score=float(data.get("score", 0)),
                summary=data.get("summary", "Visual verification complete"),
                issues=issues,
                suggestions=[],
                raw_response=raw_response
            )
        except Exception as fallback_error:
            return VerificationResult(
                score=0.0,
                summary=f"Snapshot verification failed: {e}",
                issues=[str(e)],
                suggestions=[],
                raw_response=str(e)
            )


def cross_validate(
    filepath: str,
    entries: list[CarrierEntry],
    initial_result: VerificationResult,
    sheet_name: Optional[str] = None
) -> VerificationResult:
    """
    Cross-validate the first-pass verification to filter false positives.

    Args:
        filepath: Path to the original Excel file
        entries: Extracted CarrierEntry objects
        initial_result: Result from first verification pass
        sheet_name: Optional sheet name

    Returns:
        VerificationResult with refined findings
    """
    snapshot_path = _get_snapshot_path(filepath)
    if not snapshot_path:
        # No snapshot available, return initial result as-is
        return initial_result

    model = _get_client()
    excel_content = _excel_to_text(filepath, sheet_name)
    extracted_content = _entries_to_text(entries)

    # Format issues and suggestions for the prompt
    issues_list = "\n".join(f"- {issue}" for issue in initial_result.issues) if initial_result.issues else "None identified"
    suggestions_list = "\n".join(f"- {s}" for s in initial_result.suggestions) if initial_result.suggestions else "None"

    prompt = CROSS_VALIDATION_PROMPT.format(
        excel_content=excel_content,
        extracted_content=extracted_content,
        initial_score=initial_result.score,
        initial_summary=initial_result.summary,
        issues_list=issues_list,
        suggestions_list=suggestions_list
    )

    generation_config = {
        "temperature": 0,
        "response_mime_type": "application/json",
        "response_schema": CROSS_VALIDATION_SCHEMA,
    }

    try:
        image = Image.open(snapshot_path)
        response = model.generate_content(
            [prompt, image],
            generation_config=generation_config
        )
        raw_response = response.text
        data = json.loads(raw_response)

        # Combine confirmed and new issues
        all_issues = list(data.get("confirmed_issues", []))
        all_issues.extend(data.get("new_issues", []))

        # Build summary noting how many issues were dismissed
        dismissed_count = len(data.get("dismissed_issues", []))
        summary = data.get("summary", "Cross-validation complete")
        if dismissed_count > 0:
            summary += f" ({dismissed_count} false positives filtered)"

        return VerificationResult(
            score=float(data.get("adjusted_score", initial_result.score)),
            summary=summary,
            issues=all_issues,
            suggestions=data.get("suggestions", []),
            raw_response=f"INITIAL:\n{initial_result.raw_response}\n\nCROSS-VALIDATION:\n{raw_response}"
        )
    except Exception as e:
        # If cross-validation fails, return initial result
        return VerificationResult(
            score=initial_result.score,
            summary=f"{initial_result.summary} (cross-validation failed: {e})",
            issues=initial_result.issues,
            suggestions=initial_result.suggestions,
            raw_response=initial_result.raw_response
        )


def verify_file(filepath: str, sheet_name: Optional[str] = None) -> VerificationResult:
    """
    Extract and verify a single file using two-pass cross-validation.

    Pass 1: Initial verification comparing extracted data to source
    Pass 2: Cross-validation to filter false positives and confirm real issues

    Args:
        filepath: Path to Excel file
        sheet_name: Optional sheet name

    Returns:
        VerificationResult with cross-validated findings
    """
    from .extract import extract_tower_data

    entries = extract_tower_data(filepath, sheet_name)
    if not entries:
        return VerificationResult(
            score=0.0,
            summary="No data extracted from file",
            issues=["Extraction returned empty results"],
            suggestions=["Check if file format is supported"],
            raw_response=""
        )

    # Pass 1: Initial verification (text + image if available)
    initial_result = verify_extraction(filepath, entries, sheet_name)

    # Pass 2: Cross-validation to filter false positives
    final_result = cross_validate(filepath, entries, initial_result, sheet_name)

    return final_result

"""Adaptive extraction logic for any tower format.

This extractor analyzes the spreadsheet structure from first principles:
1. Merged cells define visual blocks - these are the primary structure
2. Spatial proximity determines relationships
3. Content patterns infer field types
4. No assumptions about specific labels or column positions
"""

import re
from dataclasses import dataclass, field
from collections import defaultdict
from pathlib import Path

import yaml
from openpyxl.utils import get_column_letter, range_boundaries
from .models import CarrierEntry, LayerSummary, parse_limit_value
from .utils import get_cell_value, get_cell_color, find_merged_range_at


# Load carrier lists from YAML
_CARRIERS_FILE = Path(__file__).parent / "carriers.yml"
_KNOWN_CARRIERS = set()
_NON_CARRIERS = set()


def _normalize_for_match(s: str) -> str:
    """Normalize string for fuzzy matching - lowercase, strip punctuation."""
    return re.sub(r'[^a-z0-9\s]', '', s.lower()).strip()


def _load_carriers():
    """Load carrier lists from YAML file."""
    global _KNOWN_CARRIERS, _NON_CARRIERS
    if _KNOWN_CARRIERS:  # Already loaded
        return

    if _CARRIERS_FILE.exists():
        with open(_CARRIERS_FILE) as f:
            data = yaml.safe_load(f)
        _KNOWN_CARRIERS = {_normalize_for_match(c) for c in data.get('carriers', [])}
        _NON_CARRIERS = {_normalize_for_match(c) for c in data.get('non_carriers', [])}


def _is_known_carrier(value: str) -> bool:
    """Check if value matches a known carrier (fuzzy match)."""
    _load_carriers()
    normalized = _normalize_for_match(value)

    # Direct match
    if normalized in _KNOWN_CARRIERS:
        return True

    # Check if any known carrier is contained in the value
    # e.g., "Chubb Bermuda" contains "chubb"
    for carrier in _KNOWN_CARRIERS:
        if carrier in normalized or normalized in carrier:
            return True

    # Check for Lloyd's patterns - very common in insurance
    if 'lloyd' in normalized or 'lloyds' in normalized:
        return True

    # Check for "London" with percentages - typically indicates London market carrier
    if 'london' in normalized and '%' in value:
        return True

    return False


def _is_non_carrier(value: str) -> bool:
    """Check if value matches a known non-carrier term."""
    _load_carriers()
    normalized = _normalize_for_match(value)

    # Direct match only - don't do substring matching on non-carrier terms
    # This prevents "Chubb Bermuda" from matching "Bermuda"
    # and "London - Fidelis" from matching "London"
    if normalized in _NON_CARRIERS:
        return True

    # For compound non-carrier terms (like "RT Layer"), check if value starts with them
    # But ONLY for multi-word non-carrier terms to avoid blocking "London - Fidelis"
    for term in _NON_CARRIERS:
        # Only do prefix matching for multi-word terms
        if ' ' in term and normalized.startswith(term + ' '):
            return True

    # Additional patterns that indicate non-carrier text
    val_lower = value.lower()

    # Sentences or descriptive text (contains multiple words with common patterns)
    if any(phrase in val_lower for phrase in [
        'subject to', 'conditions', 'terms &', 'all terms',
        'tied to', 'offer capacity', 'no rp for', 'loss rating',
        'increase in', 'decrease in', 'updated', 'by year'
    ]):
        return True

    # Starts with symbols or contains mostly special characters
    if value.startswith('*') or value.startswith('#'):
        return True

    return False


def _looks_like_policy_number(value: str) -> bool:
    """Check if value looks like a policy number rather than a carrier name.

    Policy numbers typically:
    - Are alphanumeric codes (letters + numbers mixed)
    - Have specific patterns like "ABC12345", "12345678", "ABC-123-456"
    - Are relatively short (< 30 chars)
    - Have high digit-to-letter ratio or specific prefixes
    """
    if not value or len(value) > 30:
        return False

    # Pure numeric (likely policy number)
    digits_only = value.replace('-', '').replace(' ', '')
    if digits_only.isdigit() and len(digits_only) >= 6:
        return True

    # Count digits and letters
    digits = sum(1 for c in value if c.isdigit())
    letters = sum(1 for c in value if c.isalpha())

    # If mostly digits with some letters, likely policy number
    if digits >= 4 and digits > letters:
        return True

    # Common policy number patterns
    # - Starts with letters, ends with numbers: "PG2507405", "CSP00316270P-00"
    # - Has dashes/hyphens with alphanumeric segments
    val_upper = value.upper()
    if re.match(r'^[A-Z]{1,6}\d{5,}', val_upper):  # ABC12345... or RMANAH02273P03
        return True
    if re.match(r'^[A-Z]{1,6}-?\d+', val_upper) and digits >= 5:  # ABC-12345
        return True
    if re.match(r'^\d+[A-Z]+\d*', val_upper) and digits >= 5:  # 123ABC456
        return True
    # Pattern with letters at end: "RMANAH02273P03"
    if re.match(r'^[A-Z]+\d+[A-Z]*\d*$', val_upper) and digits >= 4:
        return True

    return False


@dataclass
class Block:
    """A visual block in the spreadsheet (merged cell region or single cell)."""
    row: int
    col: int
    value: any
    rows: int = 1
    cols: int = 1
    field_type: str = None
    confidence: float = 0.0


def extract_adaptive(ws) -> tuple[list[CarrierEntry], list[LayerSummary]]:
    """Extract tower data by analyzing merged cell structure and spatial patterns.

    Returns:
        tuple: (carrier_entries, layer_summaries)
            - carrier_entries: List of per-carrier participation data
            - layer_summaries: List of layer-level aggregate data (for cross-checking)
    """

    # Step 1: Find all blocks (merged regions and significant single cells)
    blocks = _find_all_blocks(ws)

    # Step 2: Classify blocks by content
    _classify_blocks(blocks)

    # Step 3: Detect summary/aggregate columns to exclude from carrier extraction
    # These columns contain layer-level data like "Annualized Layer Rate" not per-carrier data
    summary_info = _detect_summary_columns(ws)
    summary_cols = summary_info['columns']

    # Step 4: Find layer structure by looking for large limit values
    layers = _identify_layers(blocks, ws)

    # Step 5: For each layer, find carrier blocks and their associated data
    entries = []
    layer_summaries = []
    for layer in layers:
        layer_entries = _extract_layer_data(ws, blocks, layer, summary_cols)
        entries.extend(layer_entries)

        # Extract layer summary data from summary columns (for cross-checking)
        summary = _extract_layer_summary(ws, layer, summary_info)
        if summary:
            layer_summaries.append(summary)

    return entries, layer_summaries


def _detect_summary_columns(ws) -> dict:
    """Detect columns that contain summary/aggregate data rather than per-carrier data.

    These columns typically have headers like:
    - "Annualized Layer Rate"
    - "Layer Bound Premiums"
    - "Layer Rates"
    - "Layer Target"

    Returns dict with:
        - 'columns': set of column numbers to exclude from carrier extraction
        - 'bound_premium_col': column with Layer Bound Premiums (for cross-check)
        - 'layer_target_col': column with Layer Target
        - 'layer_rate_col': column with Layer Rate
    """
    result = {
        'columns': set(),
        'bound_premium_col': None,
        'layer_target_col': None,
        'layer_rate_col': None,
    }

    # Patterns that indicate a summary column (not per-carrier data)
    # Must be specific to avoid false positives on things like "Layer Premium" which is per-carrier
    summary_patterns = [
        'annualized',
        'layer rate',  # "Layer Rates" but not "Layer Premium"
        'bound premium',
        'layer target',
        'total premium',
        'aggregate',
    ]

    # Scan header rows (typically rows 1-10) for summary column indicators
    for row in range(1, 11):
        for col in range(1, min(ws.max_column + 1, 30)):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if val and isinstance(val, str):
                # Normalize whitespace for pattern matching
                val_lower = ' '.join(val.lower().split())
                for pattern in summary_patterns:
                    if pattern in val_lower:
                        result['columns'].add(col)
                        # Track specific column types for extraction
                        if 'bound premium' in val_lower:
                            result['bound_premium_col'] = col
                        elif 'target' in val_lower:
                            result['layer_target_col'] = col
                        elif 'rate' in val_lower and 'annualized' not in val_lower:
                            result['layer_rate_col'] = col
                        break

    return result


def _extract_layer_summary(ws, layer: dict, summary_info: dict) -> LayerSummary | None:
    """Extract layer-level summary data from summary columns.

    This data is used for cross-checking that carrier premiums sum to layer totals.

    Args:
        ws: The worksheet
        layer: Layer dict with limit, start_row, end_row
        summary_info: Dict with bound_premium_col, layer_target_col, layer_rate_col

    Returns:
        LayerSummary or None if no summary data found for this layer
    """
    bound_premium_col = summary_info.get('bound_premium_col')
    layer_target_col = summary_info.get('layer_target_col')
    layer_rate_col = summary_info.get('layer_rate_col')

    # If no summary columns detected, skip
    if not any([bound_premium_col, layer_target_col, layer_rate_col]):
        return None

    start_row = layer['start_row']
    end_row = layer['end_row']

    layer_bound_premium = None
    layer_target = None
    layer_rate = None
    excel_range = None

    # Look for values in summary columns within this layer's row range
    # Summary values are typically in a specific row for each layer
    for row in range(start_row, end_row + 1):
        if bound_premium_col:
            val = get_cell_value(ws, row, bound_premium_col)
            if val is not None and isinstance(val, (int, float)) and val > 0:
                layer_bound_premium = float(val)
                excel_range = f"{get_column_letter(bound_premium_col)}{row}"

        if layer_target_col:
            val = get_cell_value(ws, row, layer_target_col)
            if val is not None and isinstance(val, (int, float)) and val > 0:
                layer_target = float(val)

        if layer_rate_col:
            val = get_cell_value(ws, row, layer_rate_col)
            if val is not None and isinstance(val, (int, float)):
                # Rate could be very small (e.g., 0.0038)
                layer_rate = float(val)

    # Only return if we found at least one value
    if layer_bound_premium is not None or layer_target is not None or layer_rate is not None:
        return LayerSummary(
            layer_limit=layer['limit'],
            layer_target=layer_target,
            layer_rate=layer_rate,
            layer_bound_premium=layer_bound_premium,
            excel_range=excel_range
        )

    return None


def _find_all_blocks(ws) -> list[Block]:
    """Find all merged cell regions and significant single cells."""
    blocks = []
    processed = set()

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if (row, col) in processed:
                continue

            val = get_cell_value(ws, row, col)
            if val is None or (isinstance(val, str) and not val.strip()):
                processed.add((row, col))
                continue

            # Check for merged range
            merged = find_merged_range_at(ws, row, col)
            if merged:
                min_c, min_r, max_c, max_r = range_boundaries(merged)
                # Only process from top-left
                if row != min_r or col != min_c:
                    processed.add((row, col))
                    continue

                block = Block(
                    row=min_r, col=min_c, value=val,
                    rows=max_r - min_r + 1,
                    cols=max_c - min_c + 1
                )
                # Mark all cells in merge as processed
                for r in range(min_r, max_r + 1):
                    for c in range(min_c, max_c + 1):
                        processed.add((r, c))
            else:
                block = Block(row=row, col=col, value=val)
                processed.add((row, col))

            blocks.append(block)

    return blocks


def _classify_blocks(blocks: list[Block]):
    """Classify each block by analyzing its content."""
    for block in blocks:
        block.field_type, block.confidence = _infer_type(block.value)


def _infer_type(value) -> tuple[str, float]:
    """Infer field type and confidence from content alone."""
    if value is None:
        return None, 0.0

    # Numeric analysis
    if isinstance(value, (int, float)):
        if value == 0:
            return 'zero', 0.5
        if 0 < value <= 1:
            return 'percentage', 0.9  # High confidence - typical participation format
        if 1 < value <= 100:
            # Could be percentage as whole number, or small dollar amount
            return 'percentage_or_number', 0.6
        if value > 1_000_000:
            return 'large_number', 0.8  # Likely limit or TIV
        if value > 1_000:
            return 'currency', 0.7  # Likely premium
        return 'number', 0.5

    if not isinstance(value, str):
        return 'unknown', 0.0

    val = str(value).strip()
    if not val:
        return None, 0.0

    val_lower = val.lower()

    # Dollar amounts with M/K/B suffix (layer limits)
    if val.startswith('$'):
        # Check for complex expressions first
        if any(x in val_lower for x in ['xs', 'x/s', 'p/o', 'excess']):
            return 'layer_description', 0.95
        # Simple dollar amount with magnitude suffix
        if any(c in val.upper() for c in ['M', 'K', 'B']) and val.count('$') == 1:
            return 'limit', 0.9
        # Plain dollar amount
        rest = val[1:].replace(',', '').replace('.', '')
        if rest.isdigit():
            return 'currency_string', 0.8

    # Percentage string
    if val.endswith('%'):
        return 'percentage_string', 0.9

    # Look for terms/coverage indicators
    coverage_patterns = ['excl', 'incl', 'flood', 'earthquake', 'eq ', 'wind',
                        'terror', 'blanket', 'margin', 'ded', 'retention',
                        'all risk', 'dic', 'aop', 'named storm', 'nws']
    if any(p in val_lower for p in coverage_patterns):
        return 'terms', 0.85

    # Common label patterns
    label_patterns = ['carrier', 'participation', 'premium', 'share', 'layer',
                     'limit', 'policy', 'terms', 'coverage', 'deductible', 'total']
    if val_lower in label_patterns or any(val_lower.startswith(p) for p in label_patterns):
        return 'label', 0.9

    # Status indicators
    if val_lower in ('tbd', 'n/a', 'pending', 'incumbent', 'new', 'renewal'):
        return 'status', 0.8

    # Catch patterns like "% Premium" - labels starting with symbols
    if val.startswith('%'):
        return 'label', 0.7

    # Policy number patterns - alphanumeric codes with specific formats
    # e.g., "RMANAH02273P03", "CSP00316270P-00", "PG2507405", "61385843"
    if _looks_like_policy_number(val):
        return 'policy_number', 0.85

    # Check against known non-carrier terms (from YAML)
    if _is_non_carrier(val):
        return 'label', 0.7

    # Very short strings (1-3 chars) are unlikely to be carrier names
    # Unless they are known acronyms
    if len(val) <= 3 and not _is_known_carrier(val):
        return 'label', 0.6

    # Check against known carrier names (from YAML) - fuzzy match
    if _is_known_carrier(val):
        return 'carrier', 0.9

    # Company name heuristics for unknown carriers
    # - Reasonable length
    # - Not starting with $ or digit
    # - Contains letters
    # - May contain common suffixes
    company_suffixes = ['inc', 'llc', 'ltd', 'co', 'corp', 'company', 'ins',
                       'insurance', 'assurance', 'specialty', 'group', 're']
    if (3 <= len(val) <= 100 and
        not val[0].isdigit() and
        not val.startswith('$') and
        any(c.isalpha() for c in val)):
        # Higher confidence if has company suffix
        if any(s in val_lower for s in company_suffixes):
            return 'carrier', 0.85
        # Medium confidence for other text that looks like a name
        if len(val) >= 3 and val[0].isupper():
            return 'carrier', 0.6

    return 'text', 0.3


def _identify_layers(blocks: list[Block], ws) -> list[dict]:
    """Identify layer boundaries from limit blocks."""
    # Find blocks that look like layer limits
    # Layer limits should be:
    # 1. In column A or B (leftmost columns)
    # 2. Formatted as dollar amounts with M/K suffix, OR
    # 3. Large numbers that are explicitly labeled as limits
    limit_blocks = []
    for b in blocks:
        if b.field_type not in ('limit', 'large_number') or b.confidence < 0.7:
            continue

        # Only consider blocks in leftmost columns (A or B) for layer limits
        # This prevents premium values from being mistaken as layer limits
        if b.col > 2:
            continue

        # For large_number type, be more selective - only if it looks like a layer limit
        if b.field_type == 'large_number':
            # Check if there's a label nearby that indicates this is NOT a layer limit
            # Labels like "Premium", "LIMIT" (per-carrier), "Policy" indicate data rows
            skip_block = False
            for other in blocks:
                if other.field_type == 'label' and other.row == b.row:
                    val_lower = str(other.value).lower()
                    # "LIMIT" in column A means per-carrier limits, not layer limit
                    # "Premium" or "Policy" means this row has data, not layer definitions
                    if 'premium' in val_lower or val_lower == 'limit' or 'policy' in val_lower:
                        skip_block = True
                        break
            if skip_block:
                continue

        limit_blocks.append(b)

    # Sort by row
    limit_blocks.sort(key=lambda b: b.row)

    # Filter to keep only "primary" limits (leftmost in their row region)
    # This helps avoid treating layer descriptions as limits
    primary_limits = []
    for block in limit_blocks:
        # Check if there's already a limit in a nearby row
        dominated = False
        for existing in primary_limits:
            # If this block is in similar row range and to the right, skip it
            if abs(block.row - existing.row) <= 2 and block.col > existing.col:
                dominated = True
                break
        if not dominated:
            primary_limits.append(block)

    # Build layer info
    layers = []
    for i, block in enumerate(primary_limits):
        limit_str = _format_limit(block.value)

        # Filter out TIV (very large numbers in first few rows)
        if block.field_type == 'large_number' and block.row <= 5:
            if isinstance(block.value, (int, float)) and block.value > 1_000_000_000:
                continue

        # Determine end row
        if i + 1 < len(primary_limits):
            end_row = primary_limits[i + 1].row - 1
        else:
            end_row = ws.max_row

        layers.append({
            'limit': limit_str,
            'limit_row': block.row,
            'limit_col': block.col,
            'start_row': block.row,
            'end_row': end_row
        })

    return layers


def _format_limit(value) -> str:
    """Format a limit value for display."""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float)):
        return parse_limit_value(value)
    return str(value)


def _split_multiline_carrier(carrier_block: Block) -> list[Block]:
    """Split a carrier block containing multiple carriers (newlines) into separate blocks."""
    value = carrier_block.value
    if not isinstance(value, str):
        return [carrier_block]

    # Split on newlines
    lines = [line.strip() for line in value.split('\n') if line.strip()]

    # If only one line, return original
    if len(lines) <= 1:
        return [carrier_block]

    # Create a new block for each line
    split_blocks = []
    for i, line in enumerate(lines):
        # Skip if it looks like a non-carrier (policy number, label, etc.)
        if _looks_like_policy_number(line):
            continue
        if _is_non_carrier(line):
            continue

        # Create new block with same position but different value
        new_block = Block(
            row=carrier_block.row + i,  # Adjust row for tracking
            col=carrier_block.col,
            value=line,
            rows=1,
            cols=carrier_block.cols,
            field_type='carrier',
            confidence=carrier_block.confidence
        )
        split_blocks.append(new_block)

    return split_blocks if split_blocks else [carrier_block]


def _extract_layer_data(ws, blocks: list[Block], layer: dict, summary_cols: set[int] = None) -> list[CarrierEntry]:
    """Extract carrier entries from a layer region.

    Args:
        ws: The worksheet
        blocks: All classified blocks
        layer: Layer dict with start_row, end_row, limit
        summary_cols: Set of column numbers to exclude (contain aggregate data, not per-carrier)
    """
    entries = []
    start_row = layer['start_row']
    end_row = layer['end_row']
    summary_cols = summary_cols or set()

    # Find all blocks in this layer's row range
    layer_blocks = [b for b in blocks if start_row <= b.row <= end_row]

    # Find column headers for this layer
    column_headers = _find_column_headers(ws, blocks, layer)

    # Find row labels for this layer (Premium vs % Premium distinction)
    row_labels = _find_row_labels(ws, blocks, layer)

    # Find carrier blocks, excluding those in summary columns
    carrier_blocks = [b for b in layer_blocks
                     if b.field_type == 'carrier' and b.confidence >= 0.5
                     and b.col not in summary_cols]

    # Split multi-line carrier blocks into individual carriers
    expanded_carriers = []
    for carrier in carrier_blocks:
        expanded_carriers.extend(_split_multiline_carrier(carrier))

    # Find data blocks (percentages, currency, terms)
    # Note: large_number is included because premium values can be > $1M
    # Exclude data from summary columns as well
    data_blocks = [b for b in layer_blocks
                  if b.field_type in ('percentage', 'percentage_or_number',
                                     'currency', 'currency_string', 'large_number', 'zero',
                                     'terms', 'layer_description')
                  and b.col not in summary_cols]

    # Try to match carriers with their data using spatial proximity
    for carrier in expanded_carriers:
        entry = _build_entry_from_proximity(ws, carrier, data_blocks, layer, column_headers, row_labels)
        if entry:
            entries.append(entry)

    return entries


def _find_column_headers(ws, blocks: list[Block], layer: dict) -> dict:
    """Find column header positions for Premium, Limit, Rate, TIV, etc.

    Looks in two places:
    1. Rows near/above the layer start (global headers)
    2. Within the layer (sub-headers for complex schematics)
    """
    headers = {}

    # First pass: look in rows near or above the layer start (traditional layout)
    search_start = max(1, layer['start_row'] - 5)
    search_end = layer['start_row'] + 3

    for block in blocks:
        if block.row < search_start or block.row > search_end:
            continue
        if block.field_type != 'label':
            continue

        val_lower = str(block.value).lower().strip()
        _classify_column_header(val_lower, block.col, headers)

    # Second pass: look for sub-headers within the layer (for complex schematics)
    # Look for specific headers that appear mid-layer as data table column headers
    # These are typically in columns beyond A/B (which contain row labels)
    for block in blocks:
        if block.row < layer['start_row'] or block.row > layer['end_row']:
            continue
        if block.field_type != 'label':
            continue
        if block.col <= 2:  # Skip row labels in columns A/B
            continue

        val_lower = str(block.value).lower().strip()

        # Detect Rate headers
        if val_lower == 'rate' and 'rate_col' not in headers:
            headers['rate_col'] = block.col

        # Detect PREMIUM column header (uppercase "PREMIUM" is often a column header)
        # This takes precedence over row labels like "Share Premium"
        elif val_lower == 'premium' and block.col > 2:
            headers['premium_col'] = block.col

        # TIV detection
        elif ('tiv' in val_lower or val_lower == 'updated tiv') and 'tiv_col' not in headers:
            headers['tiv_col'] = block.col
            if 'tiv_data_col' not in headers:
                headers['tiv_data_col'] = block.col + 1

    return headers


def _classify_column_header(val_lower: str, col: int, headers: dict):
    """Classify a column header and update headers dict."""
    if 'premium' in val_lower and 'limit' not in val_lower:
        if '% premium' in val_lower or 'share' in val_lower:
            if 'premium_share_col' not in headers:
                headers['premium_share_col'] = col
        else:
            if 'premium_col' not in headers:
                headers['premium_col'] = col
    elif 'limit' in val_lower:
        if 'limit_col' not in headers:
            headers['limit_col'] = col
    elif 'participation' in val_lower or '% share' in val_lower or val_lower == 'share':
        if 'participation_col' not in headers:
            headers['participation_col'] = col
    elif val_lower == 'rate':
        if 'rate_col' not in headers:
            headers['rate_col'] = col
    elif 'tiv' in val_lower or val_lower == 'updated tiv':
        if 'tiv_col' not in headers:
            headers['tiv_col'] = col
            if 'tiv_data_col' not in headers:
                headers['tiv_data_col'] = col + 1


def _find_row_labels(ws, blocks: list[Block], layer: dict) -> dict:
    """Find row label positions for Premium, % Premium, etc.

    Many schematics use row labels in column A to indicate what each row contains.
    Some schematics (like Super Hard) repeat labels per carrier column.
    This helps distinguish between "Premium" row and "% Premium" row.
    """
    labels = {}
    start_row = layer['start_row']
    end_row = layer['end_row']

    # First pass: look for labels in columns A or B (traditional layout)
    for block in blocks:
        if block.row < start_row or block.row > end_row:
            continue
        if block.col > 2:
            continue
        if block.field_type != 'label':
            continue

        val_lower = str(block.value).lower().strip()
        _classify_row_label(val_lower, block.row, labels)

    # Second pass: if we didn't find key labels, look in any column
    # (for schematics where labels repeat per carrier column)
    if 'premium_row' not in labels or 'participation_row' not in labels:
        for block in blocks:
            if block.row < start_row or block.row > end_row:
                continue
            if block.field_type != 'label':
                continue

            val_lower = str(block.value).lower().strip()
            _classify_row_label(val_lower, block.row, labels)

    return labels


def _classify_row_label(val_lower: str, row: int, labels: dict):
    """Classify a row label and update the labels dict if not already set."""
    # Identify row types by labels
    if 'premium' in val_lower:
        if '% premium' in val_lower or val_lower.startswith('%'):
            if 'percent_premium_row' not in labels:
                labels['percent_premium_row'] = row
        elif 'premium' == val_lower or val_lower == 'share premium' or 'layer premium' in val_lower:
            if 'premium_row' not in labels:
                labels['premium_row'] = row
    elif val_lower == 'limit':
        # "LIMIT" row often contains premium/limit values per carrier
        if 'limit_row' not in labels:
            labels['limit_row'] = row
    elif 'participation' in val_lower or '% share' in val_lower or val_lower == 'share':
        if 'participation_row' not in labels:
            labels['participation_row'] = row
    elif 'carrier' in val_lower:
        if 'carrier_row' not in labels:
            labels['carrier_row'] = row
    elif 'layer' in val_lower and 'premium' not in val_lower:
        if 'layer_row' not in labels:
            labels['layer_row'] = row
    elif 'terms' in val_lower:
        if 'terms_row' not in labels:
            labels['terms_row'] = row
    elif 'policy' in val_lower:
        # Track policy number row so we don't mistake it for premium
        if 'policy_row' not in labels:
            labels['policy_row'] = row


def _build_entry_from_proximity(ws, carrier: Block, data_blocks: list[Block],
                                 layer: dict, column_headers: dict = None,
                                 row_labels: dict = None) -> CarrierEntry:
    """Build a CarrierEntry by finding spatially related data blocks."""

    # Find data in same column or within carrier's column span
    col_range = range(carrier.col, carrier.col + carrier.cols)

    # Also check rows below the carrier (common pattern)
    row_range = range(carrier.row, min(carrier.row + 10, layer['end_row'] + 1))

    participation = None
    premium = None
    premium_share = None
    terms = None
    layer_desc = None

    # Get column header info if available
    premium_col = column_headers.get('premium_col') if column_headers else None
    limit_col = column_headers.get('limit_col') if column_headers else None
    premium_share_col = column_headers.get('premium_share_col') if column_headers else None
    rate_col = column_headers.get('rate_col') if column_headers else None  # Rate != participation
    tiv_col = column_headers.get('tiv_col') if column_headers else None    # TIV != premium
    tiv_data_col = column_headers.get('tiv_data_col') if column_headers else None  # TIV data column

    # Get row label info if available (to distinguish Premium from % Premium rows)
    percent_premium_row = row_labels.get('percent_premium_row') if row_labels else None
    premium_row = row_labels.get('premium_row') if row_labels else None
    limit_row = row_labels.get('limit_row') if row_labels else None
    policy_row = row_labels.get('policy_row') if row_labels else None

    # Sort data blocks by proximity - prefer same column, then nearby rows
    def block_priority(block):
        # Check column alignment
        block_cols = range(block.col, block.col + block.cols)
        col_overlap = any(c in col_range for c in block_cols) or any(c in block_cols for c in col_range)

        # Calculate row distance
        row_dist = abs(block.row - carrier.row)

        # Priority: column-aligned blocks first, then by row distance
        # Return tuple: (not col_aligned, row_distance) for sorting
        return (not col_overlap, row_dist)

    sorted_blocks = sorted(data_blocks, key=block_priority)

    for block in sorted_blocks:
        # Check if block is in carrier's column range
        block_cols = range(block.col, block.col + block.cols)
        in_col = any(c in col_range for c in block_cols) or any(c in block_cols for c in col_range)

        # Check if block is in nearby rows (within layer)
        in_row = block.row in row_range

        # Only match if column-aligned OR (same row and close column)
        if not in_col:
            # For non-column-aligned blocks, require same row and close proximity
            if abs(block.row - carrier.row) > 1:
                continue
            if abs(block.col - carrier.col) > 3:  # Must be within 3 columns
                continue

        if block.field_type in ('percentage', 'percentage_or_number') and participation is None:
            # Skip Rate column - rates are NOT participation percentages
            if rate_col and block.col == rate_col:
                continue

            # If we have a participation_row label, only accept values from that row (or row below)
            participation_row = row_labels.get('participation_row') if row_labels else None
            if participation_row:
                if block.row == participation_row or block.row == participation_row + 1:
                    participation = _normalize_percentage(block.value)
                # Skip percentages not in the participation row
                continue
            else:
                # No participation_row label - use proximity-based matching
                participation = _normalize_percentage(block.value)

        elif block.field_type in ('currency', 'currency_string', 'large_number', 'zero'):
            # Skip TIV column - Total Insured Value is NOT premium
            if tiv_col and block.col == tiv_col:
                continue
            # Skip TIV data column (typically to the right of TIV header)
            if tiv_data_col and block.col == tiv_data_col:
                continue
            val = _parse_currency(block.value)
            # Note: val can be 0 (falsy but valid), so check for None explicitly
            if val is not None:
                # Skip policy number row - those aren't premiums!
                if policy_row and block.row == policy_row:
                    continue

                # Use column headers to distinguish Premium from Limit
                if limit_col and block.col == limit_col:
                    # This is a Limit column, skip for premium
                    continue

                # Use row labels to distinguish Premium from % Premium
                # Note: row_labels stores the LABEL row, data is often in the NEXT row
                # So we check both the label row and the row after it

                # First check for % Premium row (if it exists)
                if percent_premium_row:
                    if block.row == percent_premium_row or block.row == percent_premium_row + 1:
                        # This is % Premium row or data row below it
                        if premium_share is None:
                            premium_share = val
                        continue

                # Then check for Premium row (if it exists)
                if premium_row:
                    if block.row == premium_row or block.row == premium_row + 1:
                        # This is the Premium row or data row below it
                        if premium is None:
                            premium = val
                        continue

                # Check for LIMIT row (contains per-carrier limits, treat as premium)
                if limit_row and block.row == limit_row:
                    if premium is None:
                        premium = val
                    continue

                # If we have a premium_col column header, only accept values from that column
                if premium_col and block.col == premium_col:
                    if premium is None:
                        premium = val
                    continue
                elif premium_col:
                    # We have a premium_col but this block isn't in it - skip
                    continue

                # Fallback to column-based logic (no explicit premium_col detected)
                if premium_share_col and block.col == premium_share_col:
                    if premium_share is None:
                        premium_share = val
                elif premium is None:
                    premium = val
                elif premium_share is None:
                    premium_share = val

        elif block.field_type == 'terms' and terms is None:
            terms = str(block.value).strip()

        elif block.field_type == 'layer_description' and layer_desc is None:
            layer_desc = str(block.value).strip()

    return CarrierEntry(
        layer_limit=layer['limit'],
        layer_description=layer_desc or '',
        carrier=str(carrier.value).strip(),
        participation_pct=participation,
        premium=premium,
        premium_share=premium_share,
        terms=terms,
        policy_number=None,
        excel_range=f"{get_column_letter(carrier.col)}{carrier.row}",
        col_span=carrier.cols,
        row_span=carrier.rows,
        fill_color=get_cell_color(ws, carrier.row, carrier.col)
    )


def _normalize_percentage(value) -> float:
    """Normalize a percentage value to 0-1 range."""
    if value is None:
        return None

    if isinstance(value, str):
        val_str = value.replace('%', '').strip()
        try:
            value = float(val_str)
        except ValueError:
            return None

    if not isinstance(value, (int, float)):
        return None

    # If > 1, assume it's a whole number percentage
    if value > 1:
        return value / 100

    return float(value)


def _parse_currency(value) -> float:
    """Parse a currency value."""
    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        val_str = value.replace('$', '').replace(',', '').strip()
        try:
            return float(val_str)
        except ValueError:
            return None

    return None

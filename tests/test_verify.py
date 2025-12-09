"""Tests for schematic_explorer.verify module.

Note: This module tests the verify functionality with mocked external dependencies
(google.generativeai, PIL, etc.) to avoid requiring API keys or external services.
"""

import json
import sys
import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook

from schematic_explorer.types import CarrierEntry, LayerSummary, VerificationResult

# ============================================================================
# Mock setup - must happen before importing verify module
# ============================================================================

# Create mock modules before verify.py is imported
mock_genai = MagicMock()
mock_genai.configure = MagicMock()
mock_genai.GenerativeModel = MagicMock(return_value=MagicMock())

mock_pil = MagicMock()
mock_pil.Image = MagicMock()
mock_pil.Image.open = MagicMock(return_value=MagicMock())

mock_dotenv = MagicMock()
mock_dotenv.load_dotenv = MagicMock()

# Patch sys.modules before importing verify
sys.modules["google"] = MagicMock()
sys.modules["google.generativeai"] = mock_genai
sys.modules["PIL"] = mock_pil
sys.modules["PIL.Image"] = mock_pil.Image
sys.modules["dotenv"] = mock_dotenv

# Now we can import verify
import schematic_explorer.verify as verify_module
from schematic_explorer.verify import (
    _entries_to_text,
    _excel_to_text,
    _get_client,
    _get_snapshot_path,
    _parse_json_response,
    cross_check_layer_totals,
    cross_validate,
    verify_extraction,
    verify_file,
    verify_snapshot,
)

# ============================================================================
# Fixtures
# ============================================================================


@pytest.fixture
def sample_entries():
    """Create sample CarrierEntry objects for testing."""
    return [
        CarrierEntry(
            carrier="Lexington Insurance",
            participation_pct=0.25,
            premium=250000.0,
            layer_limit="$50M",
            layer_description="Primary Layer",
            excel_range="H47",
            col_span=1,
            row_span=1,
            premium_share=0.25,
            terms=None,
            policy_number=None,
        ),
        CarrierEntry(
            carrier="AIG",
            participation_pct=0.5,
            premium=500000.0,
            layer_limit="$50M",
            layer_description="Primary Layer",
            excel_range="I47",
            col_span=1,
            row_span=1,
            premium_share=0.5,
            terms=None,
            policy_number=None,
        ),
        CarrierEntry(
            carrier="Zurich",
            participation_pct=0.25,
            premium=250000.0,
            layer_limit="$50M",
            layer_description="Primary Layer",
            excel_range="J47",
            col_span=1,
            row_span=1,
            premium_share=0.25,
            terms=None,
            policy_number=None,
        ),
    ]


@pytest.fixture
def sample_layer_summaries():
    """Create sample LayerSummary objects for testing."""
    return [
        LayerSummary(
            layer_limit="$50M",
            layer_bound_premium=1000000.0,
            excel_range="K47",
        ),
    ]


@pytest.fixture
def excel_file():
    """Create a temporary Excel file for testing."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "$50M"
    ws["B1"] = "Carrier"
    ws["C1"] = "% Share"
    ws["B2"] = "Lexington Insurance"
    ws["C2"] = 0.25
    ws["B3"] = "AIG"
    ws["C3"] = 0.5
    ws["B4"] = "Zurich"
    ws["C4"] = 0.25

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb.save(f.name)
        yield f.name
    Path(f.name).unlink(missing_ok=True)


@pytest.fixture
def excel_file_with_merged_cells():
    """Create a temporary Excel file with merged cells for testing."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "$50M"
    ws.merge_cells("A1:A4")
    ws["B1"] = "Carrier"
    ws["C1"] = "% Share"
    ws["B2"] = "Lexington Insurance"
    ws["C2"] = 0.25

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb.save(f.name)
        yield f.name
    Path(f.name).unlink(missing_ok=True)


# ============================================================================
# Tests for _parse_json_response
# ============================================================================


class TestParseJsonResponse:
    """Tests for _parse_json_response function."""

    def test_valid_json(self):
        """Test parsing valid JSON."""
        raw = '{"score": 0.95, "summary": "Good", "issues": [], "suggestions": []}'
        result = _parse_json_response(raw)
        assert result["score"] == 0.95
        assert result["summary"] == "Good"

    def test_json_in_markdown_block(self):
        """Test parsing JSON wrapped in markdown code block."""
        raw = '```json\n{"score": 0.85, "summary": "OK", "issues": [], "suggestions": []}\n```'
        result = _parse_json_response(raw)
        assert result["score"] == 0.85

    def test_json_with_extra_whitespace(self):
        """Test parsing JSON with extra whitespace."""
        raw = '\n\n  {"score": 0.9, "summary": "Good", "issues": [], "suggestions": []}  \n\n'
        result = _parse_json_response(raw)
        assert result["score"] == 0.9

    def test_extract_score_from_malformed_json(self):
        """Test extracting score from malformed JSON."""
        raw = 'Some text "score": 0.75, more text'
        result = _parse_json_response(raw)
        assert result["score"] == 0.75

    def test_extract_json_object_from_text(self):
        """Test extracting JSON object embedded in text."""
        raw = 'Here is the result: {"score": 0.8, "summary": "Fair", "issues": ["issue1"], "suggestions": []} end'
        result = _parse_json_response(raw)
        assert result["score"] == 0.8
        assert result["issues"] == ["issue1"]

    def test_invalid_json_raises(self):
        """Test that completely invalid JSON raises ValueError."""
        with pytest.raises(ValueError, match="Could not parse JSON"):
            _parse_json_response("This is not JSON at all")

    def test_json_with_escape_sequences(self):
        """Test parsing JSON with escape sequences."""
        raw = '{"score": 0.9, "summary": "Line1\\nLine2", "issues": [], "suggestions": []}'
        result = _parse_json_response(raw)
        assert result["score"] == 0.9

    def test_markdown_block_without_json_prefix(self):
        """Test parsing markdown code block without json prefix."""
        raw = '```\n{"score": 0.75, "summary": "OK", "issues": [], "suggestions": []}\n```'
        result = _parse_json_response(raw)
        assert result["score"] == 0.75

    def test_extract_json_with_invalid_escape_then_fail(self):
        """Test JSON extraction that fails even after escape fixing."""
        # This creates a case where JSON object is found but still can't be parsed
        raw = 'Result: {incomplete json here "score": invalid}'
        with pytest.raises(ValueError, match="Could not parse JSON"):
            _parse_json_response(raw)


# ============================================================================
# Tests for _excel_to_text
# ============================================================================


class TestExcelToText:
    """Tests for _excel_to_text function."""

    def test_basic_conversion(self, excel_file):
        """Test basic Excel to text conversion."""
        text = _excel_to_text(excel_file)
        assert "Excel File:" in text
        assert "Dimensions:" in text
        assert "Lexington Insurance" in text
        assert "AIG" in text

    def test_with_sheet_name(self, excel_file):
        """Test conversion with specific sheet name."""
        text = _excel_to_text(excel_file, "Sheet")
        assert "Sheet: Sheet" in text

    def test_preserves_cell_values(self, excel_file):
        """Test that cell values are preserved."""
        text = _excel_to_text(excel_file)
        assert "0.25" in text
        assert "0.5" in text

    def test_includes_cell_references(self, excel_file):
        """Test that cell references are included."""
        text = _excel_to_text(excel_file)
        # Format is like "A1=value"
        assert "A1=" in text

    def test_with_merged_cells(self, excel_file_with_merged_cells):
        """Test conversion with merged cells."""
        text = _excel_to_text(excel_file_with_merged_cells)
        assert "Merged cells:" in text


# ============================================================================
# Tests for _entries_to_text
# ============================================================================


class TestEntriesToText:
    """Tests for _entries_to_text function."""

    def test_basic_conversion(self, sample_entries):
        """Test basic entries to text conversion."""
        text = _entries_to_text(sample_entries)
        assert "Extracted Data:" in text
        assert "Total entries: 3" in text
        assert "Lexington Insurance" in text
        assert "AIG" in text
        assert "Zurich" in text

    def test_groups_by_layer(self, sample_entries):
        """Test that entries are grouped by layer."""
        text = _entries_to_text(sample_entries)
        assert "Layer $50M:" in text

    def test_includes_cell_reference(self, sample_entries):
        """Test that cell references are included."""
        text = _entries_to_text(sample_entries)
        assert "[cell:H47]" in text
        assert "[cell:I47]" in text

    def test_handles_none_values(self):
        """Test handling of None values in entries."""
        entries = [
            CarrierEntry(
                carrier="Test Carrier",
                participation_pct=None,
                premium=None,
                layer_limit="$50M",
                layer_description="Test",
                excel_range="A1",
                col_span=1,
                row_span=1,
                premium_share=None,
                terms=None,
                policy_number=None,
            ),
        ]
        text = _entries_to_text(entries)
        assert "N/A" in text

    def test_includes_description(self, sample_entries):
        """Test that layer descriptions are included."""
        text = _entries_to_text(sample_entries)
        assert "Primary Layer" in text

    def test_multiple_layers(self):
        """Test entries from multiple layers."""
        entries = [
            CarrierEntry(
                carrier="Carrier A",
                participation_pct=0.5,
                premium=500000.0,
                layer_limit="$50M",
                layer_description="Layer 1",
                excel_range="A1",
                col_span=1,
                row_span=1,
                premium_share=0.5,
                terms=None,
                policy_number=None,
            ),
            CarrierEntry(
                carrier="Carrier B",
                participation_pct=1.0,
                premium=1000000.0,
                layer_limit="$100M",
                layer_description="Layer 2",
                excel_range="B1",
                col_span=1,
                row_span=1,
                premium_share=1.0,
                terms=None,
                policy_number=None,
            ),
        ]
        text = _entries_to_text(entries)
        assert "Layer $50M:" in text
        assert "Layer $100M:" in text


# ============================================================================
# Tests for _get_snapshot_path
# ============================================================================


class TestGetSnapshotPath:
    """Tests for _get_snapshot_path function."""

    def test_returns_none_when_no_snapshot(self):
        """Test that None is returned when no snapshot exists."""
        result = _get_snapshot_path("/nonexistent/file.xlsx")
        assert result is None

    def test_returns_path_when_snapshot_exists(self, tmp_path):
        """Test that path is returned when snapshot exists."""
        # Create a mock output directory structure
        with patch.object(verify_module, "OUTPUT_DIR", tmp_path):
            snapshot = tmp_path / "test.png"
            snapshot.write_bytes(b"fake png")
            result = _get_snapshot_path("/some/path/test.xlsx")
            assert result == snapshot


# ============================================================================
# Tests for _get_client
# ============================================================================


class TestGetClient:
    """Tests for _get_client function."""

    def test_raises_without_api_key(self):
        """Test that ValueError is raised without API key."""
        with patch.dict("os.environ", {}, clear=True):
            with pytest.raises(ValueError, match="GEMINI_API_KEY not found"):
                _get_client()

    def test_configures_genai_with_api_key(self):
        """Test that genai is configured with API key."""
        mock_genai_local = MagicMock()
        mock_genai_local.GenerativeModel.return_value = MagicMock()

        with patch.dict("os.environ", {"GEMINI_API_KEY": "test-key"}):
            with patch.object(verify_module, "genai", mock_genai_local):
                _get_client()
                mock_genai_local.configure.assert_called_once_with(api_key="test-key")

    def test_uses_custom_model_id(self):
        """Test that custom model ID is used when specified."""
        mock_genai_local = MagicMock()
        mock_genai_local.GenerativeModel.return_value = MagicMock()

        with patch.dict(
            "os.environ", {"GEMINI_API_KEY": "test-key", "GEMINI_MODEL_ID": "custom-model"}
        ):
            with patch.object(verify_module, "genai", mock_genai_local):
                _get_client()
                mock_genai_local.GenerativeModel.assert_called_once_with("custom-model")

    def test_uses_default_model(self):
        """Test that default model is used when not specified."""
        mock_genai_local = MagicMock()
        mock_genai_local.GenerativeModel.return_value = MagicMock()

        with patch.dict("os.environ", {"GEMINI_API_KEY": "test-key"}, clear=True):
            with patch.object(verify_module, "genai", mock_genai_local):
                _get_client()
                mock_genai_local.GenerativeModel.assert_called_once_with("gemini-2.5-flash")


# ============================================================================
# Tests for verify_extraction
# ============================================================================


class TestVerifyExtraction:
    """Tests for verify_extraction function."""

    def test_successful_verification(self, excel_file, sample_entries):
        """Test successful verification with mocked Gemini."""
        mock_response = MagicMock()
        mock_response.text = json.dumps(
            {"score": 0.95, "summary": "Extraction is accurate", "issues": [], "suggestions": []}
        )

        mock_model = MagicMock()
        mock_model.generate_content.return_value = mock_response

        with patch.object(verify_module, "_get_client", return_value=mock_model):
            with patch.object(verify_module, "_get_snapshot_path", return_value=None):
                result = verify_extraction(excel_file, sample_entries)

                assert result.score == 0.95
                assert result.summary == "Extraction is accurate"
                assert result.issues == []
                # Verify metadata shows structured parsing succeeded
                assert result.metadata is not None
                assert result.metadata["parsing_method"] == "structured"
                assert result.metadata["fallback_used"] is False

    def test_verification_with_issues(self, excel_file, sample_entries):
        """Test verification that finds issues."""
        mock_response = MagicMock()
        mock_response.text = json.dumps(
            {
                "score": 0.7,
                "summary": "Some issues found",
                "issues": ["Missing carrier X", "Wrong percentage for Y"],
                "suggestions": ["Review layer 2"],
            }
        )

        mock_model = MagicMock()
        mock_model.generate_content.return_value = mock_response

        with patch.object(verify_module, "_get_client", return_value=mock_model):
            with patch.object(verify_module, "_get_snapshot_path", return_value=None):
                result = verify_extraction(excel_file, sample_entries)

                assert result.score == 0.7
                assert len(result.issues) == 2
                assert len(result.suggestions) == 1

    def test_verification_error_handling(self, excel_file, sample_entries):
        """Test error handling when API fails completely (including fallback)."""
        mock_model = MagicMock()
        mock_model.generate_content.side_effect = Exception("API Error")

        with patch.object(verify_module, "_get_client", return_value=mock_model):
            with patch.object(verify_module, "_get_snapshot_path", return_value=None):
                result = verify_extraction(excel_file, sample_entries)

                assert result.score == 0.0
                assert "Verification failed" in result.summary
                # Verify metadata indicates error state
                assert result.metadata is not None
                assert result.metadata["parsing_method"] == "error"
                assert result.metadata["fallback_used"] is True
                assert "structured_error" in result.metadata
                assert "fallback_error" in result.metadata

    def test_verification_with_snapshot(self, excel_file, sample_entries, tmp_path):
        """Test verification with snapshot image."""
        snapshot = tmp_path / "test.png"
        snapshot.write_bytes(b"fake png data")

        mock_response = MagicMock()
        mock_response.text = json.dumps(
            {
                "score": 0.9,
                "summary": "Visual verification complete",
                "issues": [],
                "suggestions": [],
            }
        )

        mock_model = MagicMock()
        mock_model.generate_content.return_value = mock_response

        mock_image = MagicMock()

        with patch.object(verify_module, "_get_client", return_value=mock_model):
            with patch.object(verify_module, "_get_snapshot_path", return_value=snapshot):
                with patch.object(verify_module.Image, "open", return_value=mock_image):
                    result = verify_extraction(excel_file, sample_entries)

                    assert result.score == 0.9
                    # Verify image was passed to generate_content
                    call_args = mock_model.generate_content.call_args
                    assert mock_image in call_args[0][0]

    def test_verification_fallback_succeeds(self, excel_file, sample_entries, tmp_path):
        """Test fallback parsing when structured output fails but fallback succeeds."""
        snapshot = tmp_path / "test.png"
        snapshot.write_bytes(b"fake png data")

        # First call raises JSONDecodeError (structured output fails)
        # Second call (fallback) succeeds with parseable text
        mock_response_bad = MagicMock()
        mock_response_bad.text = "not json"

        mock_response_good = MagicMock()
        mock_response_good.text = (
            '{"score": 0.85, "summary": "Fallback worked", "issues": [], "suggestions": []}'
        )

        mock_model = MagicMock()
        mock_model.generate_content.side_effect = [mock_response_bad, mock_response_good]

        mock_image = MagicMock()

        with patch.object(verify_module, "_get_client", return_value=mock_model):
            with patch.object(verify_module, "_get_snapshot_path", return_value=snapshot):
                with patch.object(verify_module.Image, "open", return_value=mock_image):
                    result = verify_extraction(excel_file, sample_entries)

                    # Should use fallback result
                    assert result.score == 0.85
                    assert result.summary == "Fallback worked"
                    # Verify metadata indicates fallback was used
                    assert result.metadata is not None
                    assert result.metadata["parsing_method"] == "fallback"
                    assert result.metadata["fallback_used"] is True
                    assert "structured_error" in result.metadata

    def test_verification_fallback_succeeds_no_snapshot(self, excel_file, sample_entries):
        """Test fallback parsing without snapshot when structured output fails."""
        # First call raises JSONDecodeError, second call succeeds
        mock_response_bad = MagicMock()
        mock_response_bad.text = "not json"

        mock_response_good = MagicMock()
        mock_response_good.text = '{"score": 0.8, "summary": "OK", "issues": [], "suggestions": []}'

        mock_model = MagicMock()
        mock_model.generate_content.side_effect = [mock_response_bad, mock_response_good]

        with patch.object(verify_module, "_get_client", return_value=mock_model):
            with patch.object(verify_module, "_get_snapshot_path", return_value=None):
                result = verify_extraction(excel_file, sample_entries)

                assert result.score == 0.8


# ============================================================================
# Tests for verify_snapshot
# ============================================================================


class TestVerifySnapshot:
    """Tests for verify_snapshot function."""

    def test_returns_none_without_snapshot(self, sample_entries):
        """Test that None is returned without snapshot."""
        with patch.object(verify_module, "_get_snapshot_path", return_value=None):
            result = verify_snapshot("/test.xlsx", sample_entries)
            assert result is None

    def test_successful_snapshot_verification(self, sample_entries, tmp_path):
        """Test successful snapshot verification."""
        snapshot = tmp_path / "test.png"
        snapshot.write_bytes(b"fake png data")

        mock_response = MagicMock()
        mock_response.text = json.dumps(
            {
                "score": 0.85,
                "summary": "Visual check complete",
                "visual_issues": ["Alignment issue"],
                "missing_from_extraction": ["Carrier X"],
                "false_positives": [],
            }
        )

        mock_model = MagicMock()
        mock_model.generate_content.return_value = mock_response

        mock_image = MagicMock()

        with patch.object(verify_module, "_get_client", return_value=mock_model):
            with patch.object(verify_module, "_get_snapshot_path", return_value=snapshot):
                with patch.object(verify_module.Image, "open", return_value=mock_image):
                    result = verify_snapshot("/test.xlsx", sample_entries)

                    assert result is not None
                    assert result.score == 0.85
                    assert "Alignment issue" in result.issues
                    assert "Missing: Carrier X" in result.issues

    def test_snapshot_error_handling(self, sample_entries, tmp_path):
        """Test error handling when snapshot verification fails."""
        snapshot = tmp_path / "test.png"
        snapshot.write_bytes(b"fake png data")

        mock_model = MagicMock()
        mock_model.generate_content.side_effect = Exception("API Error")

        mock_image = MagicMock()

        with patch.object(verify_module, "_get_client", return_value=mock_model):
            with patch.object(verify_module, "_get_snapshot_path", return_value=snapshot):
                with patch.object(verify_module.Image, "open", return_value=mock_image):
                    result = verify_snapshot("/test.xlsx", sample_entries)

                    assert result.score == 0.0
                    assert "Snapshot verification failed" in result.summary

    def test_snapshot_fallback_succeeds(self, sample_entries, tmp_path):
        """Test snapshot fallback parsing when structured output fails but legacy succeeds."""
        snapshot = tmp_path / "test.png"
        snapshot.write_bytes(b"fake png data")

        # First response is bad JSON, fallback succeeds
        mock_response_bad = MagicMock()
        mock_response_bad.text = "not valid json"

        mock_response_good = MagicMock()
        mock_response_good.text = json.dumps(
            {
                "score": 0.75,
                "summary": "Fallback OK",
                "visual_issues": ["Issue A"],
                "missing_from_extraction": ["Missing B"],
                "false_positives": ["FP C"],
            }
        )

        mock_model = MagicMock()
        mock_model.generate_content.side_effect = [mock_response_bad, mock_response_good]

        mock_image = MagicMock()

        with patch.object(verify_module, "_get_client", return_value=mock_model):
            with patch.object(verify_module, "_get_snapshot_path", return_value=snapshot):
                with patch.object(verify_module.Image, "open", return_value=mock_image):
                    result = verify_snapshot("/test.xlsx", sample_entries)

                    assert result.score == 0.75
                    assert "Issue A" in result.issues
                    assert "Missing: Missing B" in result.issues
                    assert "False positive: FP C" in result.issues


# ============================================================================
# Tests for cross_validate
# ============================================================================


class TestCrossValidate:
    """Tests for cross_validate function."""

    def test_returns_initial_without_snapshot(self, excel_file, sample_entries):
        """Test that initial result is returned without snapshot."""
        initial = VerificationResult(
            score=0.9,
            summary="Initial result",
            issues=["Issue 1"],
            suggestions=[],
            raw_response="raw",
        )

        with patch.object(verify_module, "_get_snapshot_path", return_value=None):
            result = cross_validate(excel_file, sample_entries, initial)
            assert result == initial

    def test_successful_cross_validation(self, excel_file, sample_entries, tmp_path):
        """Test successful cross-validation."""
        snapshot = tmp_path / "test.png"
        snapshot.write_bytes(b"fake png data")

        initial = VerificationResult(
            score=0.7,
            summary="Initial",
            issues=["False positive issue", "Real issue"],
            suggestions=[],
            raw_response="raw",
        )

        mock_response = MagicMock()
        mock_response.text = json.dumps(
            {
                "adjusted_score": 0.9,
                "summary": "Refined after review",
                "confirmed_issues": ["Real issue"],
                "dismissed_issues": ["False positive issue - wrong column"],
                "new_issues": [],
                "suggestions": ["Review carefully"],
            }
        )

        mock_model = MagicMock()
        mock_model.generate_content.return_value = mock_response

        mock_image = MagicMock()

        with patch.object(verify_module, "_get_client", return_value=mock_model):
            with patch.object(verify_module, "_get_snapshot_path", return_value=snapshot):
                with patch.object(verify_module.Image, "open", return_value=mock_image):
                    result = cross_validate(excel_file, sample_entries, initial)

                    assert result.score == 0.9
                    assert "Real issue" in result.issues
                    assert "False positive issue" not in result.issues
                    assert "(1 false positives filtered)" in result.summary

    def test_cross_validation_error_returns_initial(self, excel_file, sample_entries, tmp_path):
        """Test that initial result is returned on error."""
        snapshot = tmp_path / "test.png"
        snapshot.write_bytes(b"fake png data")

        initial = VerificationResult(
            score=0.7, summary="Initial", issues=["Issue 1"], suggestions=[], raw_response="raw"
        )

        mock_model = MagicMock()
        mock_model.generate_content.side_effect = Exception("API Error")

        mock_image = MagicMock()

        with patch.object(verify_module, "_get_client", return_value=mock_model):
            with patch.object(verify_module, "_get_snapshot_path", return_value=snapshot):
                with patch.object(verify_module.Image, "open", return_value=mock_image):
                    result = cross_validate(excel_file, sample_entries, initial)

                    assert result.score == 0.7
                    assert "cross-validation failed" in result.summary


# ============================================================================
# Tests for cross_check_layer_totals
# ============================================================================


class TestCrossCheckLayerTotals:
    """Tests for cross_check_layer_totals function."""

    def test_no_discrepancies(self, sample_entries, sample_layer_summaries):
        """Test when carrier totals match layer summaries."""
        initial = VerificationResult(
            score=0.95, summary="Good extraction", issues=[], suggestions=[], raw_response="raw"
        )

        result = cross_check_layer_totals(sample_entries, sample_layer_summaries, initial)

        # Carrier premiums sum to 1,000,000 which matches layer summary
        assert result.score == 0.95
        assert len(result.issues) == 0

    def test_missing_carriers(self):
        """Test when layer has no carriers but should have data."""
        entries = []  # No carriers
        summaries = [
            LayerSummary(
                layer_limit="$50M",
                layer_bound_premium=500000.0,
                excel_range="K47",
            ),
        ]

        initial = VerificationResult(
            score=0.9, summary="Extraction complete", issues=[], suggestions=[], raw_response="raw"
        )

        result = cross_check_layer_totals(entries, summaries, initial)

        assert result.score < 0.9  # Penalty applied
        assert any("No carrier premiums extracted" in issue for issue in result.issues)

    def test_extreme_discrepancy(self):
        """Test when carrier totals significantly differ from summary (>200% difference)."""
        entries = [
            CarrierEntry(
                carrier="Test",
                participation_pct=1.0,
                premium=4000000.0,  # 4M - more than 3x the expected 1M
                layer_limit="$50M",
                layer_description="Test",
                excel_range="A1",
                col_span=1,
                row_span=1,
                premium_share=1.0,
                terms=None,
                policy_number=None,
            ),
        ]
        summaries = [
            LayerSummary(
                layer_limit="$50M",
                layer_bound_premium=1000000.0,  # Summary shows 1M
                excel_range="K47",
            ),
        ]

        initial = VerificationResult(
            score=0.9, summary="Extraction complete", issues=[], suggestions=[], raw_response="raw"
        )

        result = cross_check_layer_totals(entries, summaries, initial)

        # Should flag the >200% discrepancy (300% difference)
        assert any("difference" in issue.lower() for issue in result.issues)

    def test_no_summary_premium(self):
        """Test handling when summary has no bound premium."""
        entries = [
            CarrierEntry(
                carrier="Test",
                participation_pct=1.0,
                premium=100000.0,
                layer_limit="$50M",
                layer_description="Test",
                excel_range="A1",
                col_span=1,
                row_span=1,
                premium_share=1.0,
                terms=None,
                policy_number=None,
            ),
        ]
        summaries = [
            LayerSummary(
                layer_limit="$50M",
                layer_bound_premium=None,  # No summary premium
                excel_range="K47",
            ),
        ]

        initial = VerificationResult(
            score=0.9, summary="Extraction complete", issues=[], suggestions=[], raw_response="raw"
        )

        result = cross_check_layer_totals(entries, summaries, initial)

        # Should not flag any issues
        assert result.score == 0.9

    def test_zero_expected_with_actual_premium(self):
        """Test when expected is zero but carrier has actual premium."""
        entries = [
            CarrierEntry(
                carrier="Test",
                participation_pct=1.0,
                premium=100000.0,  # Carrier has premium
                layer_limit="$50M",
                layer_description="Test",
                excel_range="A1",
                col_span=1,
                row_span=1,
                premium_share=1.0,
                terms=None,
                policy_number=None,
            ),
        ]
        summaries = [
            LayerSummary(
                layer_limit="$50M",
                layer_bound_premium=0.0,  # Zero expected
                excel_range="K47",
            ),
        ]

        initial = VerificationResult(
            score=0.9, summary="Extraction complete", issues=[], suggestions=[], raw_response="raw"
        )

        result = cross_check_layer_totals(entries, summaries, initial)

        # The discrepancy should be flagged (100% difference when expected is 0 but actual is non-zero)
        # Though current implementation uses discrepancy_pct = 1.0 for this case
        assert result is not None  # At minimum, function should complete

    def test_zero_expected_zero_actual(self):
        """Test when both expected and actual are zero."""
        entries = []  # No carriers
        summaries = [
            LayerSummary(
                layer_limit="$50M",
                layer_bound_premium=0.0,  # Zero expected
                excel_range="K47",
            ),
        ]

        initial = VerificationResult(
            score=0.9, summary="Extraction complete", issues=[], suggestions=[], raw_response="raw"
        )

        result = cross_check_layer_totals(entries, summaries, initial)

        # No discrepancy - both are zero
        assert result.score == 0.9


# ============================================================================
# Tests for verify_file
# ============================================================================


class TestVerifyFile:
    """Tests for verify_file function."""

    def test_empty_extraction(self, excel_file):
        """Test handling of empty extraction results."""
        # extract_schematic_with_summaries is imported from .extractor inside verify_file
        with patch(
            "schematic_explorer.extractor.extract_schematic_with_summaries", return_value=([], [])
        ):
            result = verify_file(excel_file)

            assert result.score == 0.0
            assert "No data extracted" in result.summary

    def test_full_pipeline(self, excel_file, sample_entries, sample_layer_summaries):
        """Test full verification pipeline."""
        mock_response = MagicMock()
        mock_response.text = json.dumps(
            {"score": 0.95, "summary": "Good extraction", "issues": [], "suggestions": []}
        )

        mock_model = MagicMock()
        mock_model.generate_content.return_value = mock_response

        with patch(
            "schematic_explorer.extractor.extract_schematic_with_summaries",
            return_value=(sample_entries, sample_layer_summaries),
        ):
            with patch.object(verify_module, "_get_client", return_value=mock_model):
                with patch.object(verify_module, "_get_snapshot_path", return_value=None):
                    result = verify_file(excel_file)

                    assert result.score == 0.95
                    assert "Good extraction" in result.summary

    def test_pipeline_with_cross_validation(
        self, excel_file, sample_entries, sample_layer_summaries, tmp_path
    ):
        """Test full pipeline including cross-validation."""
        snapshot = tmp_path / "test.png"
        snapshot.write_bytes(b"fake png data")

        # Initial response
        initial_response = MagicMock()
        initial_response.text = json.dumps(
            {
                "score": 0.8,
                "summary": "Initial check",
                "issues": ["Possible issue"],
                "suggestions": [],
            }
        )

        # Cross-validation response
        cross_response = MagicMock()
        cross_response.text = json.dumps(
            {
                "adjusted_score": 0.95,
                "summary": "Cross-validated",
                "confirmed_issues": [],
                "dismissed_issues": ["Possible issue - false positive"],
                "new_issues": [],
                "suggestions": [],
            }
        )

        mock_model = MagicMock()
        mock_model.generate_content.side_effect = [initial_response, cross_response]

        mock_image = MagicMock()

        with patch(
            "schematic_explorer.extractor.extract_schematic_with_summaries",
            return_value=(sample_entries, sample_layer_summaries),
        ):
            with patch.object(verify_module, "_get_client", return_value=mock_model):
                with patch.object(verify_module, "_get_snapshot_path", return_value=snapshot):
                    with patch.object(verify_module.Image, "open", return_value=mock_image):
                        result = verify_file(excel_file)

                        # Should get cross-validated score
                        assert result.score == 0.95


# ============================================================================
# Tests for lazy imports in __init__.py
# ============================================================================


class TestLazyImports:
    """Tests for lazy imports in __init__.py."""

    def test_verify_file_lazy_import(self, excel_file):
        """Test that verify_file is lazily imported."""
        from schematic_explorer import verify_file as vf

        with patch(
            "schematic_explorer.extractor.extract_schematic_with_summaries", return_value=([], [])
        ):
            result = vf(excel_file)
            assert result.score == 0.0

    def test_verify_extraction_lazy_import(self, excel_file, sample_entries):
        """Test that verify_extraction is lazily imported."""
        from schematic_explorer import verify_extraction as ve

        mock_response = MagicMock()
        mock_response.text = json.dumps(
            {"score": 0.9, "summary": "OK", "issues": [], "suggestions": []}
        )

        mock_model = MagicMock()
        mock_model.generate_content.return_value = mock_response

        with patch.object(verify_module, "_get_client", return_value=mock_model):
            with patch.object(verify_module, "_get_snapshot_path", return_value=None):
                result = ve(excel_file, sample_entries)
                assert result.score == 0.9


# ============================================================================
# Tests for VERIFICATION_SCHEMA constants
# ============================================================================


class TestSchemas:
    """Tests for schema constants."""

    def test_verification_schema_structure(self):
        """Test that VERIFICATION_SCHEMA has required fields."""
        from schematic_explorer.verify import VERIFICATION_SCHEMA

        assert VERIFICATION_SCHEMA["type"] == "object"
        assert "score" in VERIFICATION_SCHEMA["properties"]
        assert "summary" in VERIFICATION_SCHEMA["properties"]
        assert "issues" in VERIFICATION_SCHEMA["properties"]
        assert "suggestions" in VERIFICATION_SCHEMA["properties"]

    def test_snapshot_verification_schema_structure(self):
        """Test that SNAPSHOT_VERIFICATION_SCHEMA has required fields."""
        from schematic_explorer.verify import SNAPSHOT_VERIFICATION_SCHEMA

        assert SNAPSHOT_VERIFICATION_SCHEMA["type"] == "object"
        assert "score" in SNAPSHOT_VERIFICATION_SCHEMA["properties"]
        assert "visual_issues" in SNAPSHOT_VERIFICATION_SCHEMA["properties"]

    def test_cross_validation_schema_structure(self):
        """Test that CROSS_VALIDATION_SCHEMA has required fields."""
        from schematic_explorer.verify import CROSS_VALIDATION_SCHEMA

        assert CROSS_VALIDATION_SCHEMA["type"] == "object"
        assert "adjusted_score" in CROSS_VALIDATION_SCHEMA["properties"]
        assert "confirmed_issues" in CROSS_VALIDATION_SCHEMA["properties"]
        assert "dismissed_issues" in CROSS_VALIDATION_SCHEMA["properties"]

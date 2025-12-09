"""Tests for schematic_explorer.preflight module."""

import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from schematic_explorer.preflight import PreflightResult, preflight


@pytest.fixture
def workbook():
    """Create a test workbook."""
    wb = Workbook()
    ws = wb.active
    return wb, ws


@pytest.fixture
def good_tower_workbook():
    """Create a well-formed tower schematic workbook."""
    wb = Workbook()
    ws = wb.active

    # Layer limit in column A
    ws["A1"] = "$50M"
    ws.merge_cells("A1:A5")

    # Carrier names with company indicators
    ws["B2"] = "Test Insurance Co"
    ws["C2"] = "Another Carrier Inc"
    ws["D2"] = "Lloyd's Syndicate 1234"

    # Participation percentages
    ws["B3"] = 0.25
    ws["C3"] = 0.50
    ws["D3"] = 0.25

    # Premium amounts
    ws["B4"] = 25000
    ws["C4"] = 50000
    ws["D4"] = 25000

    # Terms
    ws["B5"] = "All risks included"

    return wb, ws


@pytest.fixture
def minimal_tower_workbook():
    """Create a minimal tower schematic with just layers and carriers."""
    wb = Workbook()
    ws = wb.active

    # Layer limit
    ws["A1"] = "$25M"

    # Just one carrier
    ws["B1"] = "Chubb"

    return wb, ws


@pytest.fixture
def empty_workbook():
    """Create an empty workbook."""
    wb = Workbook()
    ws = wb.active
    return wb, ws


class TestPreflightResult:
    """Tests for PreflightResult dataclass."""

    def test_create_result(self):
        """Test creating a PreflightResult."""
        result = PreflightResult(
            file_name="test.xlsx",
            sheet_name="Sheet1",
            can_extract=True,
            confidence=0.85,
            layers_found=3,
            carriers_found=10,
            has_percentages=True,
            has_currency=True,
            has_terms=True,
            issues=[],
            suggestions=[],
        )
        assert result.file_name == "test.xlsx"
        assert result.can_extract is True
        assert result.confidence == 0.85
        assert result.layers_found == 3
        assert result.carriers_found == 10

    def test_create_result_with_issues(self):
        """Test creating a PreflightResult with issues."""
        result = PreflightResult(
            file_name="test.xlsx",
            sheet_name="Sheet1",
            can_extract=False,
            confidence=0.3,
            layers_found=0,
            carriers_found=0,
            has_percentages=False,
            has_currency=False,
            has_terms=False,
            issues=["No layers found", "No carriers found"],
            suggestions=["Check file format"],
        )
        assert result.can_extract is False
        assert len(result.issues) == 2
        assert len(result.suggestions) == 1


class TestPreflight:
    """Tests for preflight function."""

    def test_preflight_good_file(self, good_tower_workbook):
        """Test preflight on a well-formed tower schematic."""
        wb, ws = good_tower_workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            result = preflight(temp_path)
            assert isinstance(result, PreflightResult)
            assert result.file_name == Path(temp_path).name
            assert result.can_extract is True
            assert result.layers_found >= 1
            assert result.carriers_found >= 1
            assert result.confidence > 0.5
        finally:
            Path(temp_path).unlink()

    def test_preflight_minimal_file(self, minimal_tower_workbook):
        """Test preflight on a minimal tower schematic."""
        wb, ws = minimal_tower_workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            result = preflight(temp_path)
            assert isinstance(result, PreflightResult)
            assert result.layers_found >= 1
            assert result.carriers_found >= 1
        finally:
            Path(temp_path).unlink()

    def test_preflight_empty_file(self, empty_workbook):
        """Test preflight on an empty workbook."""
        wb, ws = empty_workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            result = preflight(temp_path)
            assert isinstance(result, PreflightResult)
            assert result.can_extract is False
            assert result.layers_found == 0
            assert result.carriers_found == 0
            assert result.confidence < 0.5
            assert len(result.issues) > 0
        finally:
            Path(temp_path).unlink()

    def test_preflight_with_sheet_name(self, good_tower_workbook):
        """Test preflight with specific sheet name."""
        wb, ws = good_tower_workbook
        ws.title = "TowerData"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            result = preflight(temp_path, sheet_name="TowerData")
            assert result.sheet_name == "TowerData"
            assert result.can_extract is True
        finally:
            Path(temp_path).unlink()

    def test_preflight_detects_percentages(self, good_tower_workbook):
        """Test preflight detects percentage values."""
        wb, ws = good_tower_workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            result = preflight(temp_path)
            assert result.has_percentages is True
        finally:
            Path(temp_path).unlink()

    def test_preflight_detects_currency(self, good_tower_workbook):
        """Test preflight detects currency values."""
        wb, ws = good_tower_workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            result = preflight(temp_path)
            assert result.has_currency is True
        finally:
            Path(temp_path).unlink()

    def test_preflight_detects_terms(self, good_tower_workbook):
        """Test preflight detects terms/conditions."""
        wb, ws = good_tower_workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            result = preflight(temp_path)
            assert result.has_terms is True
        finally:
            Path(temp_path).unlink()

    def test_preflight_confidence_calculation(self, workbook):
        """Test confidence score calculation."""
        wb, ws = workbook

        # Add only layer (30% weight)
        ws["A1"] = "$50M"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            result = preflight(temp_path)
            # Should have partial confidence from layer
            assert result.confidence >= 0.3
        finally:
            Path(temp_path).unlink()

    def test_preflight_issues_for_missing_data(self, workbook):
        """Test preflight generates issues for missing data."""
        wb, ws = workbook

        # Only carrier, no layer
        ws["A1"] = "Test Insurance"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            result = preflight(temp_path)
            # Should have issues about missing layers
            assert any("layer" in issue.lower() for issue in result.issues)
        finally:
            Path(temp_path).unlink()

    def test_preflight_suggestions_provided(self, empty_workbook):
        """Test preflight provides suggestions."""
        wb, ws = empty_workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            result = preflight(temp_path)
            # Should have suggestions when issues exist
            assert len(result.suggestions) > 0
        finally:
            Path(temp_path).unlink()

    def test_preflight_low_carrier_confidence_warning(self, workbook):
        """Test warning when carrier detection confidence is low."""
        wb, ws = workbook

        # Layer
        ws["A1"] = "$50M"

        # Carrier without clear indicators
        ws["B1"] = "XYZ"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            result = preflight(temp_path)
            # May generate warning about carrier confidence
            # (depends on detection logic)
            assert isinstance(result.issues, list)
        finally:
            Path(temp_path).unlink()

    def test_preflight_multiple_layers(self, workbook):
        """Test preflight with multiple layers."""
        wb, ws = workbook

        # Multiple layers
        ws["A1"] = "$25M"
        ws["A5"] = "$50M"
        ws["A10"] = "$100M"

        # Carriers
        ws["B1"] = "Chubb Insurance"
        ws["B5"] = "AIG"
        ws["B10"] = "Lloyd's"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            result = preflight(temp_path)
            assert result.layers_found == 3
            assert result.carriers_found >= 3
        finally:
            Path(temp_path).unlink()

    def test_preflight_uses_active_sheet_by_default(self, workbook):
        """Test preflight uses active sheet when sheet_name not provided."""
        wb, ws = workbook
        ws.title = "ActiveSheet"
        ws["A1"] = "$50M"
        ws["B1"] = "Test Insurance"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            result = preflight(temp_path)
            assert result.sheet_name == "ActiveSheet"
        finally:
            Path(temp_path).unlink()

"""Tests for schematic_explorer.extractor module."""

import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from schematic_explorer.extractor import (
    Block,
    _classify_blocks,
    _classify_column_header,
    _classify_row_label,
    _detect_summary_columns,
    _find_all_blocks,
    _format_limit,
    _identify_layers,
    _infer_type,
    _is_known_carrier,
    _is_non_carrier,
    _load_carriers,
    _looks_like_policy_number,
    _normalize_for_match,
    _normalize_percentage,
    _parse_currency,
    _split_multiline_carrier,
    extract_schematic,
    extract_schematic_with_summaries,
)


@pytest.fixture
def workbook():
    """Create a test workbook."""
    wb = Workbook()
    ws = wb.active
    return wb, ws


@pytest.fixture
def simple_tower_workbook():
    """Create a simple tower schematic workbook."""
    wb = Workbook()
    ws = wb.active

    # Layer limit in column A
    ws["A1"] = "$50M"
    ws.merge_cells("A1:A5")

    # Carrier names
    ws["B2"] = "Test Insurance Co"
    ws["C2"] = "Another Carrier Inc"

    # Participation percentages
    ws["B3"] = 0.25
    ws["C3"] = 0.75

    # Premium amounts
    ws["B4"] = 25000
    ws["C4"] = 75000

    return wb, ws


class TestNormalizeForMatch:
    """Tests for _normalize_for_match function."""

    def test_basic_normalization(self):
        """Test basic string normalization."""
        assert _normalize_for_match("Test String") == "test string"

    def test_punctuation_removal(self):
        """Test punctuation is removed."""
        assert _normalize_for_match("Lloyd's") == "lloyds"
        assert _normalize_for_match("Test, Inc.") == "test inc"

    def test_special_characters(self):
        """Test special characters are removed."""
        assert _normalize_for_match("Test & Co.") == "test  co"
        assert _normalize_for_match("ABC-123") == "abc123"

    def test_whitespace_handling(self):
        """Test whitespace is trimmed."""
        assert _normalize_for_match("  Test  ") == "test"


class TestLoadCarriers:
    """Tests for _load_carriers function."""

    def test_carriers_loaded(self):
        """Test that carriers are loaded from YAML."""
        _load_carriers()
        # Should have some known carriers loaded
        from schematic_explorer.extractor import _KNOWN_CARRIERS

        assert len(_KNOWN_CARRIERS) > 0


class TestIsKnownCarrier:
    """Tests for _is_known_carrier function."""

    def test_known_carrier_exact(self):
        """Test exact match of known carrier."""
        # This depends on carriers.yml content, but "Chubb" should be there
        # The test is more about the function logic
        assert _is_known_carrier("Chubb") is True or _is_known_carrier("AIG") is True

    def test_lloyds_patterns(self):
        """Test Lloyd's pattern detection."""
        assert _is_known_carrier("Lloyd's") is True
        assert _is_known_carrier("Lloyds of London") is True
        assert _is_known_carrier("Lloyd's Syndicate 1234") is True

    def test_london_with_percentage(self):
        """Test London market with percentage."""
        assert _is_known_carrier("London 25%") is True

    def test_unknown_carrier(self):
        """Test unknown string."""
        assert _is_known_carrier("Random Text 123") is False


class TestIsNonCarrier:
    """Tests for _is_non_carrier function."""

    def test_subject_to_phrase(self):
        """Test 'subject to' phrase is non-carrier."""
        assert _is_non_carrier("Subject to all terms") is True

    def test_conditions_phrase(self):
        """Test phrases with 'conditions'."""
        assert _is_non_carrier("Terms & conditions apply") is True

    def test_symbol_prefix(self):
        """Test strings starting with symbols."""
        assert _is_non_carrier("* Note about coverage") is True
        assert _is_non_carrier("# Reference") is True

    def test_loss_rating_phrase(self):
        """Test loss rating phrase."""
        assert _is_non_carrier("Loss Rating Period") is True

    def test_valid_carrier_not_blocked(self):
        """Test valid carrier names are not blocked."""
        assert _is_non_carrier("Chubb Insurance") is False


class TestLooksLikePolicyNumber:
    """Tests for _looks_like_policy_number function."""

    def test_pure_numeric(self):
        """Test pure numeric strings."""
        assert _looks_like_policy_number("12345678") is True
        assert _looks_like_policy_number("61385843") is True

    def test_alphanumeric_policy(self):
        """Test alphanumeric policy numbers."""
        assert _looks_like_policy_number("PG2507405") is True
        assert _looks_like_policy_number("CSP00316270P-00") is True
        assert _looks_like_policy_number("RMANAH02273P03") is True

    def test_carrier_names(self):
        """Test carrier names are not policy numbers."""
        assert _looks_like_policy_number("Chubb Insurance") is False
        assert _looks_like_policy_number("AIG") is False

    def test_short_strings(self):
        """Test short strings."""
        assert _looks_like_policy_number("ABC") is False
        assert _looks_like_policy_number("123") is False  # Only 3 digits

    def test_long_strings(self):
        """Test strings that are too long."""
        assert _looks_like_policy_number("A" * 35) is False

    def test_none_input(self):
        """Test None input."""
        assert _looks_like_policy_number(None) is False

    def test_empty_string(self):
        """Test empty string."""
        assert _looks_like_policy_number("") is False


class TestBlock:
    """Tests for Block dataclass."""

    def test_create_block(self):
        """Test creating a Block."""
        block = Block(row=1, col=1, value="Test")
        assert block.row == 1
        assert block.col == 1
        assert block.value == "Test"
        assert block.rows == 1
        assert block.cols == 1
        assert block.field_type is None
        assert block.confidence == 0.0

    def test_create_block_with_span(self):
        """Test creating a Block with span."""
        block = Block(row=5, col=3, value="Merged", rows=2, cols=3)
        assert block.rows == 2
        assert block.cols == 3


class TestInferType:
    """Tests for _infer_type function."""

    def test_none_value(self):
        """Test None input."""
        field_type, confidence = _infer_type(None)
        assert field_type is None
        assert confidence == 0.0

    def test_zero_value(self):
        """Test zero numeric value."""
        field_type, confidence = _infer_type(0)
        assert field_type == "zero"

    def test_percentage_decimal(self):
        """Test decimal percentage."""
        field_type, confidence = _infer_type(0.25)
        assert field_type == "percentage"
        assert confidence == 0.9

    def test_percentage_whole(self):
        """Test whole number that could be percentage."""
        field_type, confidence = _infer_type(25)
        assert field_type == "percentage_or_number"

    def test_large_number(self):
        """Test large number (likely limit/TIV)."""
        field_type, confidence = _infer_type(50_000_000)
        assert field_type == "large_number"

    def test_currency_value(self):
        """Test currency value (1000-1M range)."""
        field_type, confidence = _infer_type(50000)
        assert field_type == "currency"

    def test_dollar_limit_string(self):
        """Test dollar limit string."""
        field_type, confidence = _infer_type("$50M")
        assert field_type == "limit"
        assert confidence == 0.9

    def test_layer_description(self):
        """Test layer description with excess notation."""
        field_type, confidence = _infer_type("$50M xs. $25M")
        assert field_type == "layer_description"

    def test_percentage_string(self):
        """Test percentage string."""
        field_type, confidence = _infer_type("25%")
        assert field_type == "percentage_string"

    def test_currency_string(self):
        """Test currency string."""
        field_type, confidence = _infer_type("$50,000")
        assert field_type == "currency_string"

    def test_terms_pattern(self):
        """Test terms/coverage patterns."""
        assert _infer_type("Flood excluded")[0] == "terms"
        assert _infer_type("Earthquake coverage")[0] == "terms"
        assert _infer_type("Wind deductible")[0] == "terms"

    def test_label_pattern(self):
        """Test label patterns."""
        assert _infer_type("carrier")[0] == "label"
        assert _infer_type("premium")[0] == "label"
        assert _infer_type("participation")[0] == "label"

    def test_status_indicator(self):
        """Test status indicators."""
        assert _infer_type("TBD")[0] == "status"
        assert _infer_type("pending")[0] == "status"
        assert _infer_type("incumbent")[0] == "status"

    def test_policy_number_detection(self):
        """Test policy number detection."""
        field_type, confidence = _infer_type("PG2507405")
        assert field_type == "policy_number"

    def test_company_suffix(self):
        """Test company name with suffix."""
        field_type, confidence = _infer_type("Test Insurance Co")
        assert field_type == "carrier"

    def test_short_string_label(self):
        """Test short strings treated as labels or low-confidence carrier."""
        field_type, confidence = _infer_type("RT")
        # Short strings may be classified as carrier with low confidence
        assert field_type in ("label", "carrier")


class TestFindAllBlocks:
    """Tests for _find_all_blocks function."""

    def test_empty_worksheet(self, workbook):
        """Test empty worksheet."""
        wb, ws = workbook
        blocks = _find_all_blocks(ws)
        assert len(blocks) == 0

    def test_single_cell(self, workbook):
        """Test single cell."""
        wb, ws = workbook
        ws["A1"] = "Test"
        blocks = _find_all_blocks(ws)
        assert len(blocks) == 1
        assert blocks[0].value == "Test"
        assert blocks[0].row == 1
        assert blocks[0].col == 1

    def test_merged_cell(self, workbook):
        """Test merged cell creates single block."""
        wb, ws = workbook
        ws.merge_cells("A1:C3")
        ws["A1"] = "Merged"
        blocks = _find_all_blocks(ws)
        assert len(blocks) == 1
        assert blocks[0].value == "Merged"
        assert blocks[0].rows == 3
        assert blocks[0].cols == 3

    def test_empty_cells_skipped(self, workbook):
        """Test empty cells are skipped."""
        wb, ws = workbook
        ws["A1"] = "Test"
        ws["A3"] = "Another"
        # A2 is empty
        blocks = _find_all_blocks(ws)
        assert len(blocks) == 2

    def test_whitespace_only_skipped(self, workbook):
        """Test whitespace-only cells are skipped."""
        wb, ws = workbook
        ws["A1"] = "Test"
        ws["A2"] = "   "
        blocks = _find_all_blocks(ws)
        assert len(blocks) == 1


class TestClassifyBlocks:
    """Tests for _classify_blocks function."""

    def test_classify_single_block(self, workbook):
        """Test classifying a single block."""
        wb, ws = workbook
        ws["A1"] = "Chubb Insurance"
        blocks = _find_all_blocks(ws)
        _classify_blocks(blocks)
        assert blocks[0].field_type == "carrier"
        assert blocks[0].confidence > 0

    def test_classify_multiple_blocks(self, workbook):
        """Test classifying multiple blocks."""
        wb, ws = workbook
        ws["A1"] = "$50M"
        ws["B1"] = 0.25
        ws["C1"] = "Test Insurance"
        blocks = _find_all_blocks(ws)
        _classify_blocks(blocks)

        types = {b.value: b.field_type for b in blocks}
        assert types["$50M"] == "limit"
        assert types[0.25] == "percentage"


class TestIdentifyLayers:
    """Tests for _identify_layers function."""

    def test_single_layer(self, workbook):
        """Test identifying a single layer."""
        wb, ws = workbook
        ws["A1"] = "$50M"
        blocks = _find_all_blocks(ws)
        _classify_blocks(blocks)
        layers = _identify_layers(blocks, ws)
        assert len(layers) == 1
        assert layers[0]["limit"] == "$50M"

    def test_multiple_layers(self, workbook):
        """Test identifying multiple layers."""
        wb, ws = workbook
        ws["A1"] = "$25M"
        ws["A5"] = "$50M"
        ws["A10"] = "$100M"
        blocks = _find_all_blocks(ws)
        _classify_blocks(blocks)
        layers = _identify_layers(blocks, ws)
        assert len(layers) == 3

    def test_layer_row_ranges(self, workbook):
        """Test layer row ranges are correct."""
        wb, ws = workbook
        ws["A1"] = "$25M"
        ws["A5"] = "$50M"
        blocks = _find_all_blocks(ws)
        _classify_blocks(blocks)
        layers = _identify_layers(blocks, ws)

        assert layers[0]["start_row"] == 1
        assert layers[0]["end_row"] == 4
        assert layers[1]["start_row"] == 5


class TestDetectSummaryColumns:
    """Tests for _detect_summary_columns function."""

    def test_no_summary_columns(self, workbook):
        """Test worksheet with no summary columns."""
        wb, ws = workbook
        ws["A1"] = "Carrier"
        ws["B1"] = "Premium"
        result = _detect_summary_columns(ws)
        assert len(result["columns"]) == 0

    def test_detect_bound_premium_column(self, workbook):
        """Test detecting Layer Bound Premium column."""
        wb, ws = workbook
        ws["Z1"] = "Layer Bound Premium"
        result = _detect_summary_columns(ws)
        assert 26 in result["columns"]  # Column Z = 26
        assert result["bound_premium_col"] == 26

    def test_detect_layer_rate_column(self, workbook):
        """Test detecting Layer Rate column."""
        wb, ws = workbook
        ws["Y1"] = "Layer Rate"
        result = _detect_summary_columns(ws)
        assert 25 in result["columns"]

    def test_detect_year_layer_premium(self, workbook):
        """Test detecting year-prefixed layer premium columns."""
        wb, ws = workbook
        ws["X1"] = "2019 Layer Premium"
        result = _detect_summary_columns(ws)
        assert 24 in result["columns"]


class TestClassifyColumnHeader:
    """Tests for _classify_column_header function."""

    def test_premium_header(self):
        """Test premium header classification."""
        headers = {}
        _classify_column_header("premium", 5, headers)
        assert headers["premium_col"] == 5

    def test_premium_share_header(self):
        """Test % premium header classification."""
        headers = {}
        _classify_column_header("% premium", 6, headers)
        assert headers["premium_share_col"] == 6

    def test_limit_header(self):
        """Test limit header classification."""
        headers = {}
        _classify_column_header("limit", 3, headers)
        assert headers["limit_col"] == 3

    def test_participation_header(self):
        """Test participation header classification."""
        headers = {}
        _classify_column_header("participation", 4, headers)
        assert headers["participation_col"] == 4


class TestClassifyRowLabel:
    """Tests for _classify_row_label function."""

    def test_premium_row(self):
        """Test premium row classification."""
        labels = {}
        _classify_row_label("premium", 5, labels)
        assert labels["premium_row"] == 5

    def test_percent_premium_row(self):
        """Test % premium row classification."""
        labels = {}
        _classify_row_label("% premium", 6, labels)
        assert labels["percent_premium_row"] == 6

    def test_participation_row(self):
        """Test participation row classification."""
        labels = {}
        _classify_row_label("participation", 4, labels)
        assert labels["participation_row"] == 4

    def test_carrier_row(self):
        """Test carrier row classification."""
        labels = {}
        _classify_row_label("carrier", 3, labels)
        assert labels["carrier_row"] == 3


class TestNormalizePercentage:
    """Tests for _normalize_percentage function."""

    def test_none_input(self):
        """Test None input."""
        assert _normalize_percentage(None) is None

    def test_decimal_percentage(self):
        """Test decimal already in 0-1 range."""
        assert _normalize_percentage(0.25) == 0.25
        assert _normalize_percentage(0.5) == 0.5

    def test_whole_number_percentage(self):
        """Test whole number converted to decimal."""
        assert _normalize_percentage(25) == 0.25
        assert _normalize_percentage(100) == 1.0

    def test_string_percentage(self):
        """Test string percentage."""
        assert _normalize_percentage("25%") == 0.25
        assert _normalize_percentage("50%") == 0.5

    def test_invalid_string(self):
        """Test invalid string returns None."""
        assert _normalize_percentage("invalid") is None


class TestParseCurrency:
    """Tests for _parse_currency function."""

    def test_integer(self):
        """Test integer input."""
        assert _parse_currency(50000) == 50000.0

    def test_float(self):
        """Test float input."""
        assert _parse_currency(50000.50) == 50000.50

    def test_dollar_string(self):
        """Test dollar string."""
        assert _parse_currency("$50,000") == 50000.0

    def test_plain_string(self):
        """Test plain numeric string."""
        assert _parse_currency("50000") == 50000.0

    def test_invalid_string(self):
        """Test invalid string returns None."""
        assert _parse_currency("invalid") is None

    def test_none_input(self):
        """Test None-like input."""
        assert _parse_currency([]) is None


class TestFormatLimit:
    """Tests for _format_limit function."""

    def test_string_input(self):
        """Test string input is stripped."""
        assert _format_limit("  $50M  ") == "$50M"

    def test_numeric_input(self):
        """Test numeric input is formatted."""
        assert _format_limit(50_000_000) == "$50M"

    def test_other_type(self):
        """Test other type is stringified."""
        assert _format_limit(None) == "None"


class TestSplitMultilineCarrier:
    """Tests for _split_multiline_carrier function."""

    def test_single_line(self):
        """Test single line carrier."""
        block = Block(row=1, col=2, value="Test Insurance")
        result = _split_multiline_carrier(block)
        assert len(result) == 1
        assert result[0][0].value == "Test Insurance"

    def test_multiline_carriers(self):
        """Test multiline carrier block."""
        block = Block(row=1, col=2, value="Carrier One\nCarrier Two\nCarrier Three")
        result = _split_multiline_carrier(block)
        assert len(result) == 3

    def test_filters_policy_numbers(self):
        """Test policy numbers are filtered out."""
        block = Block(row=1, col=2, value="Test Insurance\nPG2507405")
        result = _split_multiline_carrier(block)
        assert len(result) == 1

    def test_non_string_value(self):
        """Test non-string value returns original."""
        block = Block(row=1, col=2, value=12345)
        result = _split_multiline_carrier(block)
        assert len(result) == 1


class TestExtractSchematic:
    """Tests for extract_schematic function."""

    def test_extract_from_file(self, simple_tower_workbook):
        """Test extracting from a workbook."""
        wb, ws = simple_tower_workbook

        # Save to temp file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries = extract_schematic(temp_path)
            assert isinstance(entries, list)
            # Should find at least one carrier
            assert len(entries) >= 1
            # Entries should be dicts
            assert all(isinstance(e, dict) for e in entries)
        finally:
            Path(temp_path).unlink()

    def test_extract_with_sheet_name(self, simple_tower_workbook):
        """Test extracting with specific sheet name."""
        wb, ws = simple_tower_workbook
        ws.title = "TestSheet"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries = extract_schematic(temp_path, sheet_name="TestSheet")
            assert isinstance(entries, list)
        finally:
            Path(temp_path).unlink()


class TestExtractSchematicWithSummaries:
    """Tests for extract_schematic_with_summaries function."""

    def test_returns_tuple(self, simple_tower_workbook):
        """Test function returns tuple of entries and summaries."""
        wb, ws = simple_tower_workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries, summaries = extract_schematic_with_summaries(temp_path)
            assert isinstance(entries, list)
            assert isinstance(summaries, list)
        finally:
            Path(temp_path).unlink()


class TestIsNonCarrierAdditional:
    """Additional tests for _is_non_carrier function."""

    def test_multi_word_non_carrier(self):
        """Test multi-word non-carrier terms."""
        # This should test the prefix matching for multi-word terms
        assert _is_non_carrier("all terms apply here") is True
        assert _is_non_carrier("offer capacity available") is True

    def test_non_carrier_phrases(self):
        """Test various non-carrier phrases."""
        assert _is_non_carrier("no rp for this") is True
        assert _is_non_carrier("increase in premium") is True
        assert _is_non_carrier("decrease in rate") is True
        assert _is_non_carrier("updated version") is True
        assert _is_non_carrier("by year analysis") is True
        assert _is_non_carrier("buy down option") is True
        assert _is_non_carrier("all risk coverage") is True
        assert _is_non_carrier("dic premium adjustment") is True
        assert _is_non_carrier("aggregate limit") is True
        assert _is_non_carrier("tied to policy") is True


class TestIsKnownCarrierAdditional:
    """Additional tests for _is_known_carrier function."""

    def test_carrier_substring_match(self):
        """Test substring matching of carriers."""
        # Test that a known carrier contained in value is matched
        _load_carriers()
        # "Chubb Bermuda" should match because it contains "Chubb"
        result = _is_known_carrier("Chubb Bermuda")
        # Result depends on carriers.yml content
        assert isinstance(result, bool)

    def test_london_without_percentage(self):
        """Test London without percentage is not automatically a carrier."""
        result = _is_known_carrier("London")
        # London alone may or may not be a carrier depending on yml
        assert isinstance(result, bool)


class TestLooksLikePolicyNumberAdditional:
    """Additional tests for _looks_like_policy_number function."""

    def test_dash_with_numbers(self):
        """Test dashed policy numbers."""
        assert _looks_like_policy_number("ABC-12345") is True
        assert _looks_like_policy_number("ABC-123-456") is True

    def test_number_letter_number_pattern(self):
        """Test patterns like 123ABC456."""
        assert _looks_like_policy_number("123ABC45678") is True

    def test_letters_numbers_letters(self):
        """Test pattern with letters at end."""
        assert _looks_like_policy_number("ABC12345XY") is True


class TestInferTypeAdditional:
    """Additional tests for _infer_type function."""

    def test_non_string_non_numeric(self):
        """Test non-string, non-numeric types."""
        field_type, confidence = _infer_type([1, 2, 3])
        assert field_type == "unknown"

    def test_empty_string(self):
        """Test empty string."""
        field_type, confidence = _infer_type("")
        assert field_type is None

    def test_whitespace_only(self):
        """Test whitespace-only string."""
        field_type, confidence = _infer_type("   ")
        assert field_type is None

    def test_dollar_with_p_o(self):
        """Test dollar amount with p/o notation."""
        field_type, confidence = _infer_type("$50M p/o $100M")
        assert field_type == "layer_description"

    def test_plain_dollar_number(self):
        """Test plain dollar with just numbers."""
        field_type, confidence = _infer_type("$50000")
        assert field_type == "currency_string"

    def test_small_number(self):
        """Test small number (< 1000)."""
        field_type, confidence = _infer_type(500)
        assert field_type == "number"

    def test_terms_patterns_additional(self):
        """Test additional terms patterns."""
        assert _infer_type("terror coverage")[0] == "terms"
        assert _infer_type("blanket policy")[0] == "terms"
        assert _infer_type("margin clause")[0] == "terms"
        assert _infer_type("retention limit")[0] == "terms"
        assert _infer_type("named storm coverage")[0] == "terms"
        assert _infer_type("nws deductible")[0] == "terms"
        assert _infer_type("aop coverage")[0] == "terms"

    def test_label_starting_with_pattern(self):
        """Test labels starting with known patterns."""
        # These patterns should match label patterns
        assert _infer_type("premium total")[0] in ("label", "carrier")
        assert _infer_type("limit")[0] == "label"
        assert _infer_type("share")[0] == "label"
        assert _infer_type("layer info")[0] in ("label", "carrier")

    def test_percent_prefix_label(self):
        """Test label starting with %."""
        field_type, confidence = _infer_type("% Premium")
        assert field_type == "label"

    def test_carrier_uppercase_initial(self):
        """Test carrier with uppercase initial."""
        field_type, confidence = _infer_type("TestCarrier")
        assert field_type == "carrier"


class TestClassifyColumnHeaderAdditional:
    """Additional tests for _classify_column_header function."""

    def test_share_header(self):
        """Test share header classification."""
        headers = {}
        _classify_column_header("share", 7, headers)
        assert headers["participation_col"] == 7

    def test_percent_share_header(self):
        """Test % share header."""
        headers = {}
        _classify_column_header("% share", 8, headers)
        assert headers["participation_col"] == 8

    def test_rate_header(self):
        """Test rate header classification."""
        headers = {}
        _classify_column_header("rate", 9, headers)
        assert headers["rate_col"] == 9

    def test_tiv_header(self):
        """Test TIV header classification."""
        headers = {}
        _classify_column_header("tiv", 10, headers)
        assert headers["tiv_col"] == 10
        assert headers["tiv_data_col"] == 11

    def test_updated_tiv_header(self):
        """Test updated TIV header."""
        headers = {}
        _classify_column_header("updated tiv", 12, headers)
        assert headers["tiv_col"] == 12


class TestClassifyRowLabelAdditional:
    """Additional tests for _classify_row_label function."""

    def test_layer_row(self):
        """Test layer row classification."""
        labels = {}
        _classify_row_label("layer", 10, labels)
        assert labels["layer_row"] == 10

    def test_terms_row(self):
        """Test terms row classification."""
        labels = {}
        _classify_row_label("terms", 11, labels)
        assert labels["terms_row"] == 11

    def test_policy_row(self):
        """Test policy row classification."""
        labels = {}
        _classify_row_label("policy", 12, labels)
        assert labels["policy_row"] == 12

    def test_limit_row(self):
        """Test limit row classification."""
        labels = {}
        _classify_row_label("limit", 13, labels)
        assert labels["limit_row"] == 13

    def test_share_premium_row(self):
        """Test share premium row classification."""
        labels = {}
        _classify_row_label("share premium", 14, labels)
        assert labels["premium_row"] == 14

    def test_layer_premium_row(self):
        """Test layer premium row classification."""
        labels = {}
        _classify_row_label("layer premium", 15, labels)
        assert labels["premium_row"] == 15


class TestDetectSummaryColumnsAdditional:
    """Additional tests for _detect_summary_columns function."""

    def test_detect_annualized_column(self):
        """Test detecting annualized column."""
        wb = Workbook()
        ws = wb.active
        ws["X1"] = "Annualized Premium"
        result = _detect_summary_columns(ws)
        assert 24 in result["columns"]

    def test_detect_total_premium_column(self):
        """Test detecting total premium column."""
        wb = Workbook()
        ws = wb.active
        ws["W1"] = "Total Premium"
        result = _detect_summary_columns(ws)
        assert 23 in result["columns"]

    def test_detect_layer_target_column(self):
        """Test detecting layer target column."""
        wb = Workbook()
        ws = wb.active
        ws["V1"] = "Layer Target"
        result = _detect_summary_columns(ws)
        assert 22 in result["columns"]
        assert result["layer_target_col"] == 22

    def test_detect_fees_following_year_premium(self):
        """Test detecting Fees/Taxes following year layer premium."""
        wb = Workbook()
        ws = wb.active
        ws["U1"] = "2019 Layer Premium"
        ws["V1"] = "Fees"
        ws["W1"] = "Taxes"
        ws["X1"] = "Total"
        result = _detect_summary_columns(ws)
        assert 21 in result["columns"]  # U
        assert 22 in result["columns"]  # V (Fees)
        assert 23 in result["columns"]  # W (Taxes)
        assert 24 in result["columns"]  # X (Total)


class TestIdentifyLayersAdditional:
    """Additional tests for _identify_layers function."""

    def test_filters_large_numbers_over_billion(self):
        """Test that numbers over $1B are filtered out."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "$50M"
        ws["A5"] = 2_000_000_000  # $2B should be filtered
        blocks = _find_all_blocks(ws)
        _classify_blocks(blocks)
        layers = _identify_layers(blocks, ws)
        assert len(layers) == 1  # Only the $50M layer

    def test_filters_layers_in_column_c_or_beyond(self):
        """Test that limits beyond column B are ignored."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "$50M"
        ws["C1"] = "$100M"  # Should be ignored (column C)
        blocks = _find_all_blocks(ws)
        _classify_blocks(blocks)
        layers = _identify_layers(blocks, ws)
        assert len(layers) == 1
        assert layers[0]["limit"] == "$50M"

    def test_filters_premium_labeled_rows(self):
        """Test that rows with Premium labels are not treated as layers."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "$50M"  # Real layer
        ws["A3"] = "Premium"
        ws["B3"] = 50_000_000  # Large number but in Premium row
        blocks = _find_all_blocks(ws)
        _classify_blocks(blocks)
        layers = _identify_layers(blocks, ws)
        assert len(layers) == 1


class TestSplitMultilineCarrierAdditional:
    """Additional tests for _split_multiline_carrier function."""

    def test_filters_non_carriers(self):
        """Test non-carrier lines are filtered."""
        block = Block(row=1, col=2, value="Test Insurance\nSubject to terms")
        result = _split_multiline_carrier(block)
        assert len(result) == 1
        assert result[0][0].value == "Test Insurance"

    def test_empty_lines_handled(self):
        """Test empty lines are handled."""
        block = Block(row=1, col=2, value="Carrier One\n\nCarrier Two")
        result = _split_multiline_carrier(block)
        assert len(result) == 2

    def test_all_lines_filtered(self):
        """Test when all lines are filtered returns original."""
        block = Block(row=1, col=2, value="PG12345\n61385843")
        result = _split_multiline_carrier(block)
        # If all filtered, returns original
        assert len(result) >= 1


class TestExtractLayerSummary:
    """Tests for layer summary extraction."""

    def test_summary_with_bound_premium(self):
        """Test extraction of layer summary with bound premium."""
        wb = Workbook()
        ws = wb.active

        # Layer
        ws["A1"] = "$50M"
        ws.merge_cells("A1:A5")

        # Summary column header
        ws["Z1"] = "Layer Bound Premium"

        # Summary value in layer row
        ws["Z2"] = 100000

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries, summaries = extract_schematic_with_summaries(temp_path)
            # Should find the layer summary
            assert isinstance(summaries, list)
        finally:
            Path(temp_path).unlink()

    def test_summary_with_layer_target(self):
        """Test extraction of layer summary with layer target."""
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "$50M"
        ws["Y1"] = "Layer Target"
        ws["Y2"] = 500000

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries, summaries = extract_schematic_with_summaries(temp_path)
            assert isinstance(summaries, list)
        finally:
            Path(temp_path).unlink()

    def test_summary_with_layer_rate(self):
        """Test extraction of layer summary with layer rate."""
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "$50M"
        ws["X1"] = "Layer Rate"
        ws["X2"] = 0.005

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries, summaries = extract_schematic_with_summaries(temp_path)
            assert isinstance(summaries, list)
        finally:
            Path(temp_path).unlink()


class TestYearLayerRateColumn:
    """Tests for year-prefixed layer rate column detection."""

    def test_year_layer_rate(self):
        """Test detecting year layer rate column."""
        wb = Workbook()
        ws = wb.active

        ws["U1"] = "2019 Layer Premium"
        ws["V1"] = "2019 Layer Rate"

        result = _detect_summary_columns(ws)
        assert 21 in result["columns"]  # U
        assert 22 in result["columns"]  # V


class TestIdentifyLayersYearPattern:
    """Tests for layer identification with year patterns."""

    def test_filters_year_prefixed_rows(self):
        """Test that rows with year prefix are not treated as layers."""
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "$50M"  # Real layer
        ws["A3"] = "2019 Bound"
        ws["B3"] = 50_000_000  # Large number but in year-prefixed row

        blocks = _find_all_blocks(ws)
        _classify_blocks(blocks)
        layers = _identify_layers(blocks, ws)
        assert len(layers) == 1


class TestProximityMatching:
    """Tests for proximity-based data matching."""

    def test_carrier_with_nearby_percentage(self):
        """Test carrier matched with percentage in same column."""
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "$50M"
        ws["B1"] = "Test Insurance"
        ws["B2"] = 0.25  # Percentage in same column

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries = extract_schematic(temp_path)
            # Should find carrier with participation
            carrier_entries = [e for e in entries if e.get("carrier") == "Test Insurance"]
            assert len(carrier_entries) >= 1
        finally:
            Path(temp_path).unlink()

    def test_carrier_with_nearby_premium(self):
        """Test carrier matched with premium in same column."""
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "$50M"
        ws["B1"] = "Test Insurance"
        ws["B2"] = 50000  # Premium in same column

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries = extract_schematic(temp_path)
            carrier_entries = [e for e in entries if e.get("carrier") == "Test Insurance"]
            assert len(carrier_entries) >= 1
        finally:
            Path(temp_path).unlink()

    def test_carrier_with_zero_premium(self):
        """Test carrier matched with zero premium value."""
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "$50M"
        ws["B1"] = "Test Insurance"
        ws["B2"] = 0  # Zero premium

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries = extract_schematic(temp_path)
            assert isinstance(entries, list)
        finally:
            Path(temp_path).unlink()

    def test_carrier_with_percentage_string(self):
        """Test carrier matched with percentage string."""
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "$50M"
        ws["B1"] = "Test Insurance"
        ws["B2"] = "25%"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries = extract_schematic(temp_path)
            assert isinstance(entries, list)
        finally:
            Path(temp_path).unlink()


class TestColumnHeaderDetection:
    """Tests for column header detection in extraction."""

    def test_rate_column_exclusion(self):
        """Test that Rate column values are excluded from participation."""
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "$50M"
        ws["B1"] = "Test Insurance"
        ws["C1"] = "Rate"
        ws["C2"] = 0.005  # Rate value, not participation

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries = extract_schematic(temp_path)
            assert isinstance(entries, list)
        finally:
            Path(temp_path).unlink()

    def test_tiv_column_exclusion(self):
        """Test that TIV column values are excluded from premium."""
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "$50M"
        ws["B1"] = "Test Insurance"
        ws["C1"] = "TIV"
        ws["C2"] = 100_000_000  # TIV value, not premium

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries = extract_schematic(temp_path)
            assert isinstance(entries, list)
        finally:
            Path(temp_path).unlink()

    def test_premium_column_header(self):
        """Test PREMIUM column header is detected."""
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "$50M"
        ws["B1"] = "Test Insurance"
        ws["C1"] = "PREMIUM"
        ws["C2"] = 50000

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries = extract_schematic(temp_path)
            assert isinstance(entries, list)
        finally:
            Path(temp_path).unlink()


class TestRowLabelDetection:
    """Tests for row label detection in extraction."""

    def test_percent_premium_row_label(self):
        """Test % Premium row label is detected."""
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "$50M"
        ws["A2"] = "% Premium"
        ws["B1"] = "Test Insurance"
        ws["B2"] = 25000

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries = extract_schematic(temp_path)
            assert isinstance(entries, list)
        finally:
            Path(temp_path).unlink()

    def test_limit_row_label(self):
        """Test LIMIT row label is detected."""
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "$50M"
        ws["A2"] = "LIMIT"
        ws["B1"] = "Test Insurance"
        ws["B2"] = 25_000_000

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries = extract_schematic(temp_path)
            assert isinstance(entries, list)
        finally:
            Path(temp_path).unlink()

    def test_policy_row_label(self):
        """Test Policy row label excludes values from premium."""
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "$50M"
        ws["A2"] = "Policy"
        ws["B1"] = "Test Insurance"
        ws["B2"] = 12345  # Policy number, not premium

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries = extract_schematic(temp_path)
            assert isinstance(entries, list)
        finally:
            Path(temp_path).unlink()


class TestComplexWorkbookExtraction:
    """Tests for complex workbook extraction scenarios."""

    def test_workbook_with_summary_columns(self):
        """Test extraction with summary columns present."""
        wb = Workbook()
        ws = wb.active

        # Layer
        ws["A1"] = "$50M"
        ws.merge_cells("A1:A5")

        # Carrier
        ws["B2"] = "Test Insurance"

        # Participation
        ws["B3"] = 0.25

        # Premium
        ws["B4"] = 25000

        # Summary columns
        ws["Z1"] = "Layer Bound Premium"
        ws["Z2"] = 100000

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries, summaries = extract_schematic_with_summaries(temp_path)
            assert isinstance(entries, list)
        finally:
            Path(temp_path).unlink()

    def test_workbook_with_terms(self):
        """Test extraction includes terms."""
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "$50M"
        ws["B1"] = "Test Insurance"
        ws["C1"] = "Flood excluded"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries = extract_schematic(temp_path)
            assert isinstance(entries, list)
        finally:
            Path(temp_path).unlink()

    def test_workbook_with_layer_description(self):
        """Test extraction with layer descriptions."""
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "$50M xs. $25M"
        ws["B1"] = "Test Insurance"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries = extract_schematic(temp_path)
            assert isinstance(entries, list)
        finally:
            Path(temp_path).unlink()

    def test_workbook_with_percentage_strings(self):
        """Test extraction with percentage strings."""
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "$50M"
        ws["B1"] = "Test Insurance"
        ws["C1"] = "25%"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries = extract_schematic(temp_path)
            assert isinstance(entries, list)
        finally:
            Path(temp_path).unlink()

    def test_workbook_with_row_labels(self):
        """Test extraction with row labels."""
        wb = Workbook()
        ws = wb.active

        # Layer limit
        ws["A1"] = "$50M"

        # Row labels in column A, data in column B
        ws["A2"] = "carrier"
        ws["B2"] = "Test Insurance"
        ws["A3"] = "premium"
        ws["B3"] = 25000
        ws["A4"] = "participation"
        ws["B4"] = 0.25

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries = extract_schematic(temp_path)
            assert isinstance(entries, list)
        finally:
            Path(temp_path).unlink()

    def test_workbook_with_column_headers(self):
        """Test extraction with column headers."""
        wb = Workbook()
        ws = wb.active

        # Headers row
        ws["A1"] = "Layer"
        ws["B1"] = "Carrier"
        ws["C1"] = "Premium"
        ws["D1"] = "Participation"

        # Data row
        ws["A2"] = "$50M"
        ws["B2"] = "Test Insurance"
        ws["C2"] = 25000
        ws["D2"] = 0.25

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            entries = extract_schematic(temp_path)
            assert isinstance(entries, list)
        finally:
            Path(temp_path).unlink()

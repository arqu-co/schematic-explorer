"""Tests for schematic_explorer.utils module."""

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from schematic_explorer.utils import (
    find_merged_range_at,
    get_cell_color,
    get_cell_value,
)


@pytest.fixture
def workbook():
    """Create a test workbook."""
    wb = Workbook()
    ws = wb.active
    return wb, ws


class TestFindMergedRangeAt:
    """Tests for find_merged_range_at function."""

    def test_no_merged_cells(self, workbook):
        """Test when there are no merged cells."""
        wb, ws = workbook
        ws["A1"] = "Test"
        result = find_merged_range_at(ws, 1, 1)
        assert result is None

    def test_find_merged_range_top_left(self, workbook):
        """Test finding merged range from top-left cell."""
        wb, ws = workbook
        ws.merge_cells("A1:C3")
        ws["A1"] = "Merged"
        result = find_merged_range_at(ws, 1, 1)
        assert result == "A1:C3"

    def test_find_merged_range_middle(self, workbook):
        """Test finding merged range from middle cell."""
        wb, ws = workbook
        ws.merge_cells("A1:C3")
        ws["A1"] = "Merged"
        result = find_merged_range_at(ws, 2, 2)
        assert result == "A1:C3"

    def test_find_merged_range_bottom_right(self, workbook):
        """Test finding merged range from bottom-right cell."""
        wb, ws = workbook
        ws.merge_cells("A1:C3")
        ws["A1"] = "Merged"
        result = find_merged_range_at(ws, 3, 3)
        assert result == "A1:C3"

    def test_cell_outside_merged_range(self, workbook):
        """Test cell outside merged range."""
        wb, ws = workbook
        ws.merge_cells("A1:C3")
        ws["A1"] = "Merged"
        result = find_merged_range_at(ws, 4, 4)
        assert result is None

    def test_multiple_merged_ranges(self, workbook):
        """Test with multiple merged ranges."""
        wb, ws = workbook
        ws.merge_cells("A1:B2")
        ws.merge_cells("D1:E2")
        ws["A1"] = "First"
        ws["D1"] = "Second"

        assert find_merged_range_at(ws, 1, 1) == "A1:B2"
        assert find_merged_range_at(ws, 1, 4) == "D1:E2"
        assert find_merged_range_at(ws, 1, 3) is None


class TestGetCellValue:
    """Tests for get_cell_value function."""

    def test_simple_cell_value(self, workbook):
        """Test getting value from simple cell."""
        wb, ws = workbook
        ws["A1"] = "Test Value"
        assert get_cell_value(ws, 1, 1) == "Test Value"

    def test_numeric_cell_value(self, workbook):
        """Test getting numeric value."""
        wb, ws = workbook
        ws["A1"] = 12345
        assert get_cell_value(ws, 1, 1) == 12345

    def test_float_cell_value(self, workbook):
        """Test getting float value."""
        wb, ws = workbook
        ws["A1"] = 123.45
        assert get_cell_value(ws, 1, 1) == 123.45

    def test_empty_cell(self, workbook):
        """Test getting value from empty cell."""
        wb, ws = workbook
        assert get_cell_value(ws, 1, 1) is None

    def test_merged_cell_value_from_top_left(self, workbook):
        """Test getting merged cell value from top-left."""
        wb, ws = workbook
        ws.merge_cells("A1:C3")
        ws["A1"] = "Merged Value"
        assert get_cell_value(ws, 1, 1) == "Merged Value"

    def test_merged_cell_value_from_middle(self, workbook):
        """Test getting merged cell value from middle cell."""
        wb, ws = workbook
        ws.merge_cells("A1:C3")
        ws["A1"] = "Merged Value"
        # Value should come from top-left cell
        assert get_cell_value(ws, 2, 2) == "Merged Value"

    def test_merged_cell_value_from_bottom_right(self, workbook):
        """Test getting merged cell value from bottom-right."""
        wb, ws = workbook
        ws.merge_cells("A1:C3")
        ws["A1"] = "Merged Value"
        assert get_cell_value(ws, 3, 3) == "Merged Value"


class TestGetCellColor:
    """Tests for get_cell_color function."""

    def test_no_fill(self, workbook):
        """Test cell with no fill."""
        wb, ws = workbook
        ws["A1"] = "Test"
        result = get_cell_color(ws, 1, 1)
        assert result is None

    def test_solid_fill_color(self, workbook):
        """Test cell with solid fill color."""
        wb, ws = workbook
        ws["A1"] = "Test"
        ws["A1"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        result = get_cell_color(ws, 1, 1)
        # openpyxl may return ARGB format (8 chars) or RGB format (6 chars)
        assert result in ("FF0000", "00FF0000")

    def test_yellow_fill_color(self, workbook):
        """Test cell with yellow fill."""
        wb, ws = workbook
        ws["A1"] = "Test"
        ws["A1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        result = get_cell_color(ws, 1, 1)
        # openpyxl may return ARGB format (8 chars) or RGB format (6 chars)
        assert result in ("FFFF00", "00FFFF00")

    def test_transparent_fill(self, workbook):
        """Test cell with transparent/no fill (00000000)."""
        wb, ws = workbook
        ws["A1"] = "Test"
        ws["A1"].fill = PatternFill(start_color="00000000", end_color="00000000", fill_type="solid")
        result = get_cell_color(ws, 1, 1)
        assert result is None

    def test_black_fill(self, workbook):
        """Test cell with black fill (000000) - treated as no fill."""
        wb, ws = workbook
        ws["A1"] = "Test"
        ws["A1"].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        result = get_cell_color(ws, 1, 1)
        assert result is None

    def test_merged_cell_color(self, workbook):
        """Test getting color from merged cell."""
        wb, ws = workbook
        ws.merge_cells("A1:C3")
        ws["A1"] = "Merged"
        ws["A1"].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        # Color should come from top-left cell
        result = get_cell_color(ws, 2, 2)
        # openpyxl may return ARGB format (8 chars) or RGB format (6 chars)
        assert result in ("00FF00", "0000FF00")

    def test_merged_cell_color_bottom_right(self, workbook):
        """Test getting color from bottom-right of merged cell."""
        wb, ws = workbook
        ws.merge_cells("A1:B2")
        ws["A1"] = "Merged"
        ws["A1"].fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        result = get_cell_color(ws, 2, 2)
        # openpyxl may return ARGB format (8 chars) or RGB format (6 chars)
        assert result in ("0000FF", "000000FF")
